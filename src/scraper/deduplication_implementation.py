"""
Deduplication Implementation - Multi-Stage Pipeline
====================================================

Dieses Skript implementiert eine vollständige Deduplication-Pipeline:

Stage 1: Export
    - Liest Roh-HTML/PDF aus content_database.db (gzip-komprimiert)
    - Dekomprimiert und bereinigt den Content (Naive Cleaning)
    - Speichert bereinigte Volltexte in data/deduplication/dedup_input.jsonl

Stage 2: Normalisierung (in deduplication.py)
    - normalize_text() für robuste Text-Normalisierung
    - compute_normalized_hash() für SHA256-Hashing

Stage 3: Exact-Deduplication
    - Berechnet Hash für jeden normalisierten Text
    - Gruppiert Dokumente mit identischem Hash
    - Entfernt exakte Duplikate (behält je 1 Dokument pro Hash)

KEIN Chunking, KEINE Embeddings, KEIN ChromaDB!

Output-Format (JSONL):
{"doc_id": "1", "source": "https://...", "content_type": "html", "title": "...", "text": "...", "char_count": 1234, "word_count": 200}

Autor: Deduplication Pipeline
Datum: 2025-12-09
"""
import sys
import os
import json

# Füge Projekt-Root zum Path hinzu
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
sys.path.insert(0, project_root)

import sqlite3
import gzip
import re
from pathlib import Path
from tqdm import tqdm
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import random

# ============================================================================
# KONFIGURATION
# ============================================================================

# Seed für reproduzierbare Zufallsstichprobe
RANDOM_SEED = 42
SAMPLE_SIZE_HTML = 20  # Anzahl HTML-Dokumente für manuelle Überprüfung
SAMPLE_SIZE_PDF = 2    # Anzahl PDF-Dokumente für manuelle Überprüfung

# Input: SQLite Content Database
CONTENT_DB = Path(project_root) / "data" / "content_database.db"

# Output: JSONL für Deduplication
DEDUP_OUTPUT_DIR = Path(project_root) / "data" / "deduplication"
DEDUP_INPUT_JSONL = DEDUP_OUTPUT_DIR / "dedup_input.jsonl"

# Output: Excel für Übersicht
DEDUP_EXCEL_OUTPUT_DIR = Path(project_root) / "src" / "advanced_rag" / "data"
DEDUP_OVERVIEW_EXCEL = DEDUP_EXCEL_OUTPUT_DIR / "deduplication_overview.xlsx"

# ============================================================================
# HILFSFUNKTIONEN (kopiert aus run_production_scraper.py)
# ============================================================================

def decompress_content(compressed_data: bytes) -> str:
    """Dekomprimiere gzip-Content."""
    return gzip.decompress(compressed_data).decode('utf-8')


def naive_extract_text_from_html(html: str) -> str:
    """
    Naive HTML-zu-Text Extraktion mit Strukturerhaltung.
    Konvertiert HTML zu Markdown-ähnlichem Text.
    
    Erhaltene Strukturen:
    - Überschriften (h1-h6) → # Markdown-Überschriften
    - Listen (ul/ol) → - oder 1. Listenelemente
    - Blockquotes → > Zitate
    - Absätze/Divs → Zeilenumbrüche
    
    Entfernt:
    - Script, Style, Head, Meta Tags
    - Navigation, Footer, Aside (Layout-Elemente)
    - Elemente mit menu/nav/sidebar/breadcrumb Klassen
    - UI-Texte wie "Menü schließen"
    - HTML-Tags selbst
    - Übermäßige Whitespaces
    """
    from bs4 import BeautifulSoup
    
    soup = BeautifulSoup(html, 'html.parser')
    
    # 1. Entferne unsichtbare Elemente komplett
    for element in soup(['script', 'style', 'head', 'meta', 'link', 'noscript', 'iframe']):
        element.decompose()
    
    # 2. Entferne Layout-Elemente ohne semantischen Inhalt
    for element in soup(['nav', 'footer', 'aside']):
        element.decompose()
    
    # 3. Entferne Elemente mit bestimmten Klassen/IDs (Navigation, Menüs, Sidebars)
    boilerplate_patterns = re.compile(r'menu|nav|sidebar|breadcrumb|cookie|banner|popup|modal', re.IGNORECASE)
    for element in soup.find_all(class_=boilerplate_patterns):
        element.decompose()
    for element in soup.find_all(id=boilerplate_patterns):
        element.decompose()
    
    # 4. Überschriften → Markdown
    for i, tag in enumerate(['h1', 'h2', 'h3', 'h4', 'h5', 'h6']):
        for h in soup.find_all(tag):
            prefix = '#' * (i + 1) + ' '
            h.insert_before('\n\n')
            h.insert_after('\n')
            if h.string:
                h.string = prefix + h.get_text().strip()
            else:
                h.insert_before(prefix)
    
    # 5. Listen → Markdown
    # Nummerierte Listen
    for ol in soup.find_all('ol'):
        ol.insert_before('\n')
        ol.insert_after('\n')
        for i, li in enumerate(ol.find_all('li', recursive=False), 1):
            li.insert_before(f'\n{i}. ')
    
    # Unnummerierte Listen
    for ul in soup.find_all('ul'):
        ul.insert_before('\n')
        ul.insert_after('\n')
        for li in ul.find_all('li', recursive=False):
            li.insert_before('\n- ')
    
    # 6. Blockquotes → Markdown
    for bq in soup.find_all('blockquote'):
        bq.insert_before('\n')
        bq.insert_after('\n')
        # Füge > vor dem Text ein
        text = bq.get_text().strip()
        quoted_lines = '\n'.join('> ' + line for line in text.split('\n'))
        bq.string = quoted_lines
    
    # 7. Block-Elemente → Zeilenumbrüche
    for tag in soup.find_all(['p', 'div', 'br', 'tr', 'article', 'section']):
        tag.insert_before('\n')
        if tag.name != 'br':
            tag.insert_after('\n')
    
    # 8. Tabellenzellen → Tab-getrennt
    for td in soup.find_all(['td', 'th']):
        td.insert_after('\t')
    
    # 9. Extrahiere Text
    text = soup.get_text()
    
    # 10. Entferne typische UI-Texte
    ui_patterns = [
        r'Menü schließen',
        r'Zur Übersichtsseite\s+\w+',
        r'zum Inhalt springen',
        r'Sprache wechseln',
        r'Suchbegriff eingeben',
        r'Abschicken',
        r'Finden',
        r'EnglishEnglish',
        r'Hauptnavigation\..*?anzuspringen\.',
    ]
    for pattern in ui_patterns:
        text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    
    # 11. Bereinige Whitespace
    text = re.sub(r'[ \t]+', ' ', text)  # Mehrere Spaces/Tabs zu einem Space
    text = re.sub(r'\n[ \t]+', '\n', text)  # Spaces am Zeilenanfang entfernen
    text = re.sub(r'[ \t]+\n', '\n', text)  # Spaces am Zeilenende entfernen
    text = re.sub(r'\n{3,}', '\n\n', text)  # Max 2 Zeilenumbrüche
    
    return text.strip()


def naive_clean_text(text: str) -> str:
    """Naive Text-Bereinigung für bereits extrahierten Text (z.B. PDFs)."""
    # Nur grundlegende Bereinigung
    text = re.sub(r'\s+', ' ', text)  # Normalisiere Leerzeichen
    text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)  # Entferne mehrfache Zeilenumbrüche
    return text.strip()


# ============================================================================
# DOKUMENT-VERARBEITUNG (vereinfacht - kein Chunking!)
# ============================================================================

def process_document_for_dedup(doc_id: int, url: str, title: str, 
                                content: bytes, content_type: str) -> dict | None:
    """
    Verarbeite ein einzelnes Dokument für Deduplication.
    Nur Dekompression + Cleaning, KEIN Chunking!
    
    Args:
        doc_id: Dokument-ID aus der Datenbank
        url: Quell-URL
        title: Dokumenttitel
        content: Komprimierter Inhalt (gzip)
        content_type: 'html' oder 'pdf'
    
    Returns:
        Dict mit doc_id, source, content_type, title, text
        oder None bei Fehler/leerem Dokument
    """
    try:
        # 1. Dekomprimiere Content
        raw_content = decompress_content(content)
        
        # 2. Extrahiere und bereinige Text (Naive Cleaning)
        if content_type == 'html':
            cleaned_text = naive_extract_text_from_html(raw_content)
        else:  # pdf
            cleaned_text = naive_clean_text(raw_content)
        
        # 3. Prüfe ob Text substantiell ist
        if not cleaned_text or len(cleaned_text.strip()) < 50:
            return None
        
        return {
            "doc_id": str(doc_id),
            "source": url,
            "content_type": content_type,
            "title": title,
            "text": cleaned_text,
            "char_count": len(cleaned_text),
            "word_count": len(cleaned_text.split())
        }
        
    except Exception as e:
        print(f"\n⚠️  Fehler bei Dokument {doc_id}: {e}")
        return None


# ============================================================================
# HAUPTFUNKTION
# ============================================================================

def run_dedup_export():
    """
    Exportiere alle bereinigten Dokumente in JSONL für Deduplication.
    
    Diese Funktion:
    1. Liest alle Dokumente aus content_database.db
    2. Dekomprimiert und bereinigt jeden Text (Naive Cleaning)
    3. Schreibt alle Dokumente in data/deduplication/dedup_input.jsonl
    """
    
    start_time = time.time()
    
    print("=" * 80)
    print("DEDUPLICATION STAGE 1: Bereinigte Texte exportieren")
    print("=" * 80)
    print(f"Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Modus: Naive Cleaning (identisch zu run_production_scraper.py)")
    print("=" * 80)
    
    # Prüfe ob Content Database existiert
    if not CONTENT_DB.exists():
        print(f"\n❌ Content Database nicht gefunden: {CONTENT_DB}")
        print("   Bitte zuerst den Crawler ausführen!")
        return
    
    # Erstelle Output-Verzeichnis
    DEDUP_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    print(f"\n📁 Input:  {CONTENT_DB}")
    print(f"📁 Output: {DEDUP_INPUT_JSONL}")
    
    # Verbinde zur Datenbank
    print("\n" + "=" * 80)
    print("SCHRITT 1: Dokumente laden")
    print("=" * 80)
    
    conn = sqlite3.connect(CONTENT_DB)
    
    # Hole Gesamtanzahl
    cursor = conn.execute("SELECT COUNT(*) FROM documents")
    total_docs = cursor.fetchone()[0]
    print(f"\n📊 Gefunden: {total_docs:,} Dokumente")
    
    # Zähle nach Content-Type
    cursor = conn.execute("SELECT content_type, COUNT(*) FROM documents GROUP BY content_type")
    for content_type, count in cursor.fetchall():
        print(f"   • {content_type.upper()}: {count:,}")
    
    # Statistiken
    stats = {
        'total': 0,
        'processed': 0,
        'skipped': 0,
        'errors': 0,
        'html': 0,
        'pdf': 0,
        'total_chars': 0,
        'total_words': 0
    }
    
    # Verarbeite alle Dokumente
    print("\n" + "=" * 80)
    print("SCHRITT 2: Dokumente verarbeiten")
    print("=" * 80)
    print("🔄 Workflow: Decompress → Naive Clean → JSONL Export")
    print()
    
    cursor = conn.execute("""
        SELECT id, url, title, content, content_type
        FROM documents
        ORDER BY id
    """)
    
    # Öffne Output-Datei
    with open(DEDUP_INPUT_JSONL, 'w', encoding='utf-8') as jsonl_file:
        
        # Progress Bar
        with tqdm(total=total_docs, desc="📄 Dokumente verarbeiten", unit="doc") as pbar:
            
            for row in cursor:
                doc_id, url, title, content, content_type = row
                stats['total'] += 1
                
                # Zeige aktuelles Dokument
                short_title = title[:35] + "..." if len(title) > 35 else title
                pbar.set_description(f"📄 [{content_type.upper()}] {short_title}")
                
                # Verarbeite Dokument
                result = process_document_for_dedup(
                    doc_id, url, title, content, content_type
                )
                
                if result is None:
                    stats['skipped'] += 1
                else:
                    # Schreibe in JSONL
                    json_line = json.dumps(result, ensure_ascii=False)
                    jsonl_file.write(json_line + '\n')
                    
                    # Update Stats
                    stats['processed'] += 1
                    stats[content_type] += 1
                    stats['total_chars'] += result['char_count']
                    stats['total_words'] += result['word_count']
                
                # Update Progress Bar
                pbar.set_postfix({
                    'OK': stats['processed'],
                    'Skip': stats['skipped'],
                    'Words': f"{stats['total_words']:,}"
                })
                pbar.update(1)
    
    conn.close()
    
    # Berechne Laufzeit
    elapsed = time.time() - start_time
    minutes = int(elapsed // 60)
    seconds = int(elapsed % 60)
    
    # Finale Statistiken
    print("\n" + "=" * 80)
    print("✅ EXPORT ABGESCHLOSSEN")
    print("=" * 80)
    
    print(f"\n⏱️  Laufzeit: {minutes} Minuten {seconds} Sekunden")
    
    print(f"\n📊 Dokument-Statistiken:")
    print(f"   • Gesamt: {stats['total']:,} Dokumente")
    print(f"   • Verarbeitet: {stats['processed']:,}")
    print(f"   • Übersprungen: {stats['skipped']:,} (leer oder zu kurz)")
    print(f"   • HTML: {stats['html']:,}")
    print(f"   • PDF: {stats['pdf']:,}")
    
    print(f"\n📝 Text-Statistiken:")
    print(f"   • Gesamtzeichen: {stats['total_chars']:,}")
    print(f"   • Gesamtwörter: {stats['total_words']:,}")
    if stats['processed'] > 0:
        print(f"   • Ø Zeichen/Dokument: {stats['total_chars'] // stats['processed']:,}")
        print(f"   • Ø Wörter/Dokument: {stats['total_words'] // stats['processed']:,}")
    
    print(f"\n💾 Output:")
    print(f"   • Datei: {DEDUP_INPUT_JSONL}")
    file_size_mb = DEDUP_INPUT_JSONL.stat().st_size / 1024 / 1024
    print(f"   • Größe: {file_size_mb:.2f} MB")
    
    print("\n" + "=" * 80)
    print("🎯 NÄCHSTER SCHRITT: Deduplication auf dedup_input.jsonl anwenden")
    print("=" * 80)
    print(f"\nEnde: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)


# ============================================================================
# HILFSFUNKTIONEN FÜR SPÄTERE DEDUP-STAGES
# ============================================================================

def load_dedup_input() -> list:
    """
    Lade alle Dokumente aus der JSONL-Datei.
    
    Returns:
        Liste von Dokumenten als Dictionaries
    """
    if not DEDUP_INPUT_JSONL.exists():
        print(f"❌ JSONL nicht gefunden: {DEDUP_INPUT_JSONL}")
        print("   Bitte zuerst run_dedup_export() ausführen!")
        return []
    
    documents = []
    with open(DEDUP_INPUT_JSONL, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                documents.append(json.loads(line))
    
    print(f"✅ {len(documents):,} Dokumente geladen aus {DEDUP_INPUT_JSONL}")
    return documents


def save_dedup_output(documents: list, output_path: Path):
    """
    Speichere deduplizierte Dokumente in JSONL.
    
    Args:
        documents: Liste von Dokumenten
        output_path: Zielpfad für JSONL
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        for doc in documents:
            json_line = json.dumps(doc, ensure_ascii=False)
            f.write(json_line + '\n')
    
    print(f"✅ {len(documents):,} Dokumente gespeichert: {output_path}")


# ============================================================================
# STAGE 3: EXACT-DEDUPLICATION
# ============================================================================

def run_exact_dedup(show_examples: bool = True) -> dict:
    """
    Stage 3: Exact-Deduplication mittels normalisiertem Hashing.
    
    Findet Dokumente mit identischem Inhalt nach Normalisierung.
    
    Args:
        show_examples: Wenn True, zeige Beispiele für gefundene Duplikate
        
    Returns:
        Dictionary mit Statistiken und Ergebnissen
    """
    from src.advanced_rag.pre_retrieval.deduplication import normalize_text, compute_normalized_hash
    from collections import defaultdict
    
    print("\n" + "=" * 80)
    print("STAGE 3: EXACT-DEDUPLICATION")
    print("=" * 80)
    
    # Lade Dokumente aus Stage 1
    documents = load_dedup_input()
    if not documents:
        return {"error": "Keine Dokumente geladen"}
    
    print(f"\n📊 Berechne Hashes für {len(documents):,} Dokumente...")
    
    # Hash-Berechnung mit Gruppierung
    hash_to_docs = defaultdict(list)
    
    for doc in tqdm(documents, desc="Hashing"):
        text = doc.get('text', '')
        doc_hash = compute_normalized_hash(text)
        doc['normalized_hash'] = doc_hash
        hash_to_docs[doc_hash].append(doc)
    
    # Analyse der Duplikat-Gruppen
    unique_hashes = len(hash_to_docs)
    duplicate_groups = {h: docs for h, docs in hash_to_docs.items() if len(docs) > 1}
    total_duplicates = sum(len(docs) - 1 for docs in duplicate_groups.values())
    
    print("\n" + "-" * 80)
    print("📈 ERGEBNISSE")
    print("-" * 80)
    print(f"   Gesamt-Dokumente:     {len(documents):,}")
    print(f"   Unique Hashes:        {unique_hashes:,}")
    print(f"   Duplikat-Gruppen:     {len(duplicate_groups):,}")
    print(f"   Exakte Duplikate:     {total_duplicates:,}")
    print(f"   Reduktion:            {total_duplicates / len(documents) * 100:.1f}%")
    
    # Zeige Beispiele für Duplikat-Gruppen
    if show_examples and duplicate_groups:
        print("\n" + "-" * 80)
        print("📋 BEISPIELE FÜR DUPLIKAT-GRUPPEN")
        print("-" * 80)
        
        # Sortiere nach Gruppengröße (größte zuerst)
        sorted_groups = sorted(duplicate_groups.items(), key=lambda x: len(x[1]), reverse=True)
        
        for i, (hash_val, docs) in enumerate(sorted_groups[:5]):  # Top 5 Gruppen
            print(f"\n🔗 Duplikat-Gruppe {i+1} ({len(docs)} Dokumente, Hash: {hash_val[:16]}...)")
            
            for j, doc in enumerate(docs[:3]):  # Max 3 Dokumente pro Gruppe zeigen
                source = doc.get('source', 'N/A')
                # Kürze URL für Anzeige
                if len(source) > 70:
                    source = source[:35] + "..." + source[-30:]
                title = (doc.get('title') or 'N/A')[:50]
                char_count = doc.get('char_count', 0)
                
                print(f"      [{j+1}] Doc {doc['doc_id']} ({doc['content_type'].upper()})")
                print(f"          URL: {source}")
                print(f"          Titel: {title}")
                print(f"          Zeichen: {char_count:,}")
            
            if len(docs) > 3:
                print(f"      ... und {len(docs) - 3} weitere Dokumente")
        
        if len(sorted_groups) > 5:
            print(f"\n   ... und {len(sorted_groups) - 5} weitere Duplikat-Gruppen")
    
    # Erstelle deduplizierte Liste (behalte jeweils erstes Dokument)
    unique_documents = []
    removed_documents = []
    seen_hashes = set()
    
    for doc in documents:
        doc_hash = doc['normalized_hash']
        if doc_hash not in seen_hashes:
            seen_hashes.add(doc_hash)
            unique_documents.append(doc)
        else:
            removed_documents.append(doc)
    
    # Speichere Ergebnisse
    output_unique = DEDUP_OUTPUT_DIR / "dedup_stage3_unique.jsonl"
    output_removed = DEDUP_OUTPUT_DIR / "dedup_stage3_removed.jsonl"
    
    save_dedup_output(unique_documents, output_unique)
    save_dedup_output(removed_documents, output_removed)
    
    # Detaillierte Statistik nach Content-Type
    print("\n" + "-" * 80)
    print("📊 STATISTIK NACH CONTENT-TYPE")
    print("-" * 80)
    
    for content_type in ['html', 'pdf']:
        original = sum(1 for d in documents if d['content_type'] == content_type)
        unique = sum(1 for d in unique_documents if d['content_type'] == content_type)
        removed = original - unique
        print(f"   {content_type.upper():4s}: {original:,} → {unique:,} (entfernt: {removed:,})")
    
    print("\n" + "=" * 80)
    print("✅ Stage 3 abgeschlossen!")
    print("=" * 80)
    
    return {
        "stage": 1,
        "stage_name": "Exact-Deduplication",
        "total_documents": len(documents),
        "unique_documents": len(unique_documents),
        "duplicate_groups": len(duplicate_groups),
        "total_duplicates": total_duplicates,
        "reduction_percent": total_duplicates / len(documents) * 100,
        "html_original": sum(1 for d in documents if d['content_type'] == 'html'),
        "html_unique": sum(1 for d in unique_documents if d['content_type'] == 'html'),
        "pdf_original": sum(1 for d in documents if d['content_type'] == 'pdf'),
        "pdf_unique": sum(1 for d in unique_documents if d['content_type'] == 'pdf'),
        "output_unique": str(output_unique),
        "output_removed": str(output_removed),
        "duplicate_details": duplicate_groups  # Für Excel-Export
    }


# ============================================================================
# EXCEL-ÜBERSICHT GENERIERUNG
# ============================================================================

def create_dedup_overview_excel(stage_results: list[dict]) -> Path:
    """
    Erstellt eine Übersichts-Excel mit Deduplication-Statistiken pro Stage.
    
    Args:
        stage_results: Liste von Dictionaries mit Ergebnissen pro Stage
        
    Returns:
        Path zur erstellten Excel-Datei
    """
    from collections import defaultdict
    
    print("\n" + "=" * 80)
    print("📊 ERSTELLE DEDUPLICATION-ÜBERSICHT EXCEL")
    print("=" * 80)
    
    # Erstelle Output-Verzeichnis
    DEDUP_EXCEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # Styles definieren
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_fill = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
    title_font = Font(bold=True, size=14)
    number_font = Font(name='Consolas', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================================================
    # SHEET 1: Übersicht
    # ========================================================================
    ws_overview = wb.active
    ws_overview.title = "Übersicht"
    
    # Titel
    ws_overview.cell(row=1, column=1, value="🔍 Deduplication Pipeline - Übersicht")
    ws_overview.cell(row=1, column=1).font = title_font
    ws_overview.cell(row=2, column=1, value=f"Erstellt am: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Stage-Tabelle
    row = 4
    headers = ["Stage", "Name", "Input", "Output", "Entfernt", "Reduktion %", "Duplikat-Gruppen"]
    for col, header in enumerate(headers, 1):
        cell = ws_overview.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Daten pro Stage
    for result in stage_results:
        row += 1
        values = [
            result.get('stage', 'N/A'),
            result.get('stage_name', 'N/A'),
            result.get('total_documents', 0),
            result.get('unique_documents', 0),
            result.get('total_duplicates', 0),
            f"{result.get('reduction_percent', 0):.1f}%",
            result.get('duplicate_groups', 0)
        ]
        for col, value in enumerate(values, 1):
            cell = ws_overview.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col >= 3:  # Zahlen rechtsbündig
                cell.alignment = Alignment(horizontal='right')
                cell.font = number_font
    
    # Gesamt-Zeile
    row += 1
    if stage_results:
        first_stage = stage_results[0]
        last_stage = stage_results[-1]
        total_removed = first_stage.get('total_documents', 0) - last_stage.get('unique_documents', 0)
        total_reduction = (total_removed / first_stage.get('total_documents', 1)) * 100
        
        gesamt_values = [
            "GESAMT",
            f"Alle {len(stage_results)} Stages",
            first_stage.get('total_documents', 0),
            last_stage.get('unique_documents', 0),
            total_removed,
            f"{total_reduction:.1f}%",
            sum(r.get('duplicate_groups', 0) for r in stage_results)
        ]
        for col, value in enumerate(gesamt_values, 1):
            cell = ws_overview.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True)
            cell.fill = subheader_fill
            cell.border = thin_border
            if col >= 3:
                cell.alignment = Alignment(horizontal='right')
    
    # Content-Type Statistik
    row += 3
    ws_overview.cell(row=row, column=1, value="📁 Statistik nach Content-Type")
    ws_overview.cell(row=row, column=1).font = title_font
    row += 1
    
    ct_headers = ["Content-Type", "Original", "Nach Dedup", "Entfernt", "Reduktion %"]
    for col, header in enumerate(ct_headers, 1):
        cell = ws_overview.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    if stage_results:
        first = stage_results[0]
        last = stage_results[-1]
        
        for ct in ['html', 'pdf']:
            row += 1
            orig = first.get(f'{ct}_original', 0)
            uniq = last.get(f'{ct}_unique', 0)
            removed = orig - uniq
            red_pct = (removed / orig * 100) if orig > 0 else 0
            
            values = [ct.upper(), orig, uniq, removed, f"{red_pct:.1f}%"]
            for col, value in enumerate(values, 1):
                cell = ws_overview.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col >= 2:
                    cell.alignment = Alignment(horizontal='right')
                    cell.font = number_font
    
    # Spaltenbreiten anpassen
    ws_overview.column_dimensions['A'].width = 12
    ws_overview.column_dimensions['B'].width = 25
    ws_overview.column_dimensions['C'].width = 12
    ws_overview.column_dimensions['D'].width = 12
    ws_overview.column_dimensions['E'].width = 12
    ws_overview.column_dimensions['F'].width = 14
    ws_overview.column_dimensions['G'].width = 18
    
    # ========================================================================
    # SHEET 2: Duplikat-Gruppen Detail (für Stage 1)
    # ========================================================================
    if stage_results and stage_results[0].get('duplicate_details'):
        ws_dups = wb.create_sheet("Duplikat-Gruppen")
        
        duplicate_groups = stage_results[0]['duplicate_details']
        
        # Titel
        ws_dups.cell(row=1, column=1, value="🔗 Exakte Duplikat-Gruppen (Stage 1)")
        ws_dups.cell(row=1, column=1).font = title_font
        ws_dups.cell(row=2, column=1, value=f"Gesamt: {len(duplicate_groups)} Gruppen mit {sum(len(docs)-1 for docs in duplicate_groups.values())} Duplikaten")
        
        row = 4
        headers = ["Gruppe", "Hash (kurz)", "Anzahl Docs", "Doc IDs", "URLs", "Content-Types"]
        for col, header in enumerate(headers, 1):
            cell = ws_dups.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Sortiere nach Größe
        sorted_groups = sorted(duplicate_groups.items(), key=lambda x: len(x[1]), reverse=True)
        
        for i, (hash_val, docs) in enumerate(sorted_groups, 1):
            row += 1
            doc_ids = ", ".join(d['doc_id'] for d in docs)
            urls = "\n".join(d['source'][:80] + "..." if len(d['source']) > 80 else d['source'] for d in docs[:5])
            if len(docs) > 5:
                urls += f"\n... +{len(docs)-5} weitere"
            content_types = ", ".join(set(d['content_type'].upper() for d in docs))
            
            values = [
                i,
                hash_val[:16] + "...",
                len(docs),
                doc_ids[:100] + "..." if len(doc_ids) > 100 else doc_ids,
                urls,
                content_types
            ]
            for col, value in enumerate(values, 1):
                cell = ws_dups.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 5:  # URLs
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Spaltenbreiten
        ws_dups.column_dimensions['A'].width = 8
        ws_dups.column_dimensions['B'].width = 20
        ws_dups.column_dimensions['C'].width = 12
        ws_dups.column_dimensions['D'].width = 30
        ws_dups.column_dimensions['E'].width = 80
        ws_dups.column_dimensions['F'].width = 15
    
    # ========================================================================
    # SHEET 3: Entfernte Dokumente
    # ========================================================================
    ws_removed = wb.create_sheet("Entfernte Dokumente")
    
    ws_removed.cell(row=1, column=1, value="🗑️ Alle entfernten Dokumente")
    ws_removed.cell(row=1, column=1).font = title_font
    ws_removed.cell(row=2, column=1, value=f"Stichprobe: {SAMPLE_SIZE_HTML} HTML + {SAMPLE_SIZE_PDF} PDF (Seed: {RANDOM_SEED})")
    
    # Sammle zuerst alle entfernten Dokumente für Stichprobenauswahl
    all_removed_docs = []
    for result in stage_results:
        stage_name = result.get('stage_name', 'Unknown')
        duplicate_groups = result.get('duplicate_details', {})
        
        sorted_groups = sorted(duplicate_groups.items(), key=lambda x: len(x[1]), reverse=True)
        
        for group_idx, (hash_val, docs) in enumerate(sorted_groups, 1):
            kept_doc = docs[0]
            kept_url = kept_doc.get('source', 'N/A')
            
            for doc in docs[1:]:
                all_removed_docs.append({
                    'stage_name': stage_name,
                    'group': group_idx,
                    'doc': doc,
                    'kept_url': kept_url
                })
    
    # Trenne nach Content-Type für stratifizierte Stichprobe
    html_indices = [i for i, item in enumerate(all_removed_docs) if item['doc'].get('content_type') == 'html']
    pdf_indices = [i for i, item in enumerate(all_removed_docs) if item['doc'].get('content_type') == 'pdf']
    
    # Wähle Stichprobe mit festem Seed
    # WICHTIG: Erst 20 aus ALLEN Dokumenten (wie in der alten Version), dann separat 2 PDFs
    random.seed(RANDOM_SEED)
    
    # Schritt 1: Ziehe 20 Dokumente aus allen (reproduziert alte Stichprobe)
    base_sample_size = min(SAMPLE_SIZE_HTML, len(all_removed_docs))
    base_sample_indices = set(random.sample(range(len(all_removed_docs)), base_sample_size))
    
    # Schritt 2: Ziehe zusätzlich 2 PDFs (die noch nicht in der Stichprobe sind)
    pdf_not_in_sample = [i for i in pdf_indices if i not in base_sample_indices]
    pdf_sample_size = min(SAMPLE_SIZE_PDF, len(pdf_not_in_sample))
    if pdf_not_in_sample:
        additional_pdf_sample = set(random.sample(pdf_not_in_sample, pdf_sample_size))
    else:
        additional_pdf_sample = set()
    
    sample_indices = base_sample_indices | additional_pdf_sample
    
    # Zähle HTML und PDF in der Stichprobe
    html_sample_size = sum(1 for i in sample_indices if i in html_indices)
    pdf_sample_size = sum(1 for i in sample_indices if i in pdf_indices)
    total_sample_size = len(sample_indices)
    
    # Styles für Stichprobe
    sample_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Gelb
    
    row = 4
    headers = ["Stichprobe", "Stage", "Gruppe", "Doc ID", "Content-Type", "Entfernte URL", "Titel", "Zeichen", "Ersetzt durch URL"]
    for col, header in enumerate(headers, 1):
        cell = ws_removed.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Schreibe alle entfernten Dokumente mit Stichproben-Markierung
    for idx, item in enumerate(all_removed_docs):
        row += 1
        doc = item['doc']
        is_sample = idx in sample_indices
        
        values = [
            "✓ PRÜFEN" if is_sample else "",
            item['stage_name'],
            item['group'],
            doc.get('doc_id', 'N/A'),
            doc.get('content_type', 'N/A').upper(),
            doc.get('source', 'N/A'),
            (doc.get('title') or 'N/A')[:60],
            doc.get('char_count', 0),
            item['kept_url']
        ]
        for col, value in enumerate(values, 1):
            cell = ws_removed.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if is_sample:
                cell.fill = sample_fill
            if col == 8:  # Zeichen
                cell.alignment = Alignment(horizontal='right')
                cell.font = number_font
    
    ws_removed.column_dimensions['A'].width = 12
    ws_removed.column_dimensions['B'].width = 20
    ws_removed.column_dimensions['C'].width = 8
    ws_removed.column_dimensions['D'].width = 10
    ws_removed.column_dimensions['E'].width = 12
    ws_removed.column_dimensions['F'].width = 70
    ws_removed.column_dimensions['G'].width = 40
    ws_removed.column_dimensions['H'].width = 10
    ws_removed.column_dimensions['I'].width = 70
    
    # ========================================================================
    # SHEET 4: Stichprobe zur Überprüfung
    # ========================================================================
    ws_sample = wb.create_sheet("Stichprobe")
    
    ws_sample.cell(row=1, column=1, value="🔍 Stichprobe zur manuellen Überprüfung")
    ws_sample.cell(row=1, column=1).font = title_font
    ws_sample.cell(row=2, column=1, value=f"Seed: {RANDOM_SEED} | Stichprobe: {html_sample_size} HTML + {pdf_sample_size} PDF = {total_sample_size} von {len(all_removed_docs)} Dokumenten")
    ws_sample.cell(row=3, column=1, value="Bitte überprüfen Sie, ob die entfernten Dokumente tatsächlich Duplikate der Ersatz-Dokumente sind.")
    
    row = 5
    headers = ["#", "Typ", "Doc ID", "Entfernte URL", "Ersetzt durch URL", "Korrekt? (J/N)", "Anmerkung"]
    for col, header in enumerate(headers, 1):
        cell = ws_sample.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Schreibe nur die Stichprobe
    sample_counter = 0
    for idx in sorted(sample_indices):
        sample_counter += 1
        row += 1
        item = all_removed_docs[idx]
        doc = item['doc']
        
        values = [
            sample_counter,
            doc.get('content_type', 'N/A').upper(),
            doc.get('doc_id', 'N/A'),
            doc.get('source', 'N/A'),
            item['kept_url'],
            "",  # Korrekt? (manuell auszufüllen)
            ""   # Anmerkung (manuell auszufüllen)
        ]
        for col, value in enumerate(values, 1):
            cell = ws_sample.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col >= 6:  # Eingabefelder
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws_sample.column_dimensions['A'].width = 5
    ws_sample.column_dimensions['B'].width = 8
    ws_sample.column_dimensions['C'].width = 10
    ws_sample.column_dimensions['D'].width = 80
    ws_sample.column_dimensions['E'].width = 80
    ws_sample.column_dimensions['F'].width = 15
    ws_sample.column_dimensions['G'].width = 40
    
    # Speichere
    wb.save(DEDUP_OVERVIEW_EXCEL)
    
    print(f"\n✅ Excel-Übersicht erstellt: {DEDUP_OVERVIEW_EXCEL}")
    print(f"   📋 Sheets: {wb.sheetnames}")
    
    return DEDUP_OVERVIEW_EXCEL


def run_full_dedup_pipeline() -> Path:
    """
    Führt die vollständige Deduplication-Pipeline aus und erstellt Übersichts-Excel.
    
    Returns:
        Path zur erstellten Excel-Datei
    """
    print("\n" + "=" * 80)
    print("🚀 VOLLSTÄNDIGE DEDUPLICATION-PIPELINE")
    print("=" * 80)
    
    stage_results = []
    
    # Stage 0: Export (Vorbereitung)
    print("\n📦 Stage 0: Export der bereinigten Dokumente...")
    run_dedup_export()
    
    # Stage 1: Exact-Deduplication
    print("\n🔍 Stage 1: Exact-Deduplication...")
    result_stage1 = run_exact_dedup(show_examples=False)
    stage_results.append(result_stage1)
    
    # Erstelle Übersichts-Excel
    excel_path = create_dedup_overview_excel(stage_results)
    
    print("\n" + "=" * 80)
    print("✅ PIPELINE ABGESCHLOSSEN")
    print("=" * 80)
    print(f"\n📊 Excel-Übersicht: {excel_path}")
    
    return excel_path


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Deduplication Pipeline")
    parser.add_argument(
        '--stage', 
        type=int, 
        choices=[0, 1], 
        default=None,
        help="Einzelne Stage ausführen: 0=Export, 1=Exact-Dedup"
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help="Führe vollständige Pipeline aus (alle Stages + Excel)"
    )
    parser.add_argument(
        '--excel-only',
        action='store_true',
        help="Nur Excel-Übersicht erstellen (aus vorhandenen Ergebnissen)"
    )
    
    args = parser.parse_args()
    
    if args.all:
        run_full_dedup_pipeline()
    elif args.excel_only:
        # Lade existierende Ergebnisse und erstelle Excel
        result = run_exact_dedup(show_examples=False)
        create_dedup_overview_excel([result])
    elif args.stage == 0:
        run_dedup_export()
    elif args.stage == 1:
        run_exact_dedup()
    else:
        # Default: Vollständige Pipeline
        print("Keine Option angegeben. Nutze --help für Optionen.")
        print("Starte vollständige Pipeline...")
        run_full_dedup_pipeline()
