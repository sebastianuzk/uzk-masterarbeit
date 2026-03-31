"""
Content Deduplication für Web Scraping
======================================

Entfernt near-duplicate Dokumente mithilfe von MinHash und Simhash.

Enthält:
- normalize_text(): Text-Normalisierung für Exact-Deduplication
- ContentDeduplicator: Near-Duplicate-Erkennung via Shingling/Jaccard
"""

import hashlib
import re
import unicodedata
from typing import List, Set, Tuple
from collections import defaultdict
from dataclasses import dataclass
import logging

# Excel-Formatierung
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


# ============================================================================
# TEXT-NORMALISIERUNG FÜR EXACT-DEDUPLICATION
# ============================================================================

def normalize_text(text: str) -> str:
    """
    Normalisiere Text für Exact-Deduplication (Hashing).
    
    Diese Funktion bereitet Text so auf, dass Dokumente als gleich erkannt werden,
    wenn sie sich nur in Typografie, Groß-/Kleinschreibung, Whitespace und 
    Aufzählungsmarkern unterscheiden.
    
    Normalisierungsschritte:
    1. Lowercasing
    2. Unicode-Normalisierung (NFKC)
    3. Typografische Vereinheitlichung (Anführungszeichen, Bindestriche, NBSP)
    4. Entfernung von Aufzählungsmarkern am Zeilenanfang
    5. Entfernung dekorativer Sequenzen (----, ====, etc.)
    6. Whitespace-Normalisierung
    
    Was NICHT gemacht wird:
    - Keine Entfernung/Vereinheitlichung von Zahlen und Datumsangaben
    - Kein Stemming oder Lemmatizing
    - Umlaute (ä, ö, ü) und ß bleiben erhalten
    
    Args:
        text: Eingabetext (kann None oder leer sein)
        
    Returns:
        Normalisierter Text für Hashing
    """
    # Robuste Behandlung von None/leerem Input
    if not text:
        return ""
    
    if not isinstance(text, str):
        text = str(text)
    
    # 1. Lowercasing
    text = text.lower()
    
    # 2. Unicode-Normalisierung (NFKC)
    # - Vereinheitlicht Kompatibilitätszeichen (z.B. ﬁ → fi)
    # - Kombiniert diakritische Zeichen
    text = unicodedata.normalize('NFKC', text)
    
    # 3. Typografische Vereinheitlichung
    # 3a. Anführungszeichen → einfaches "
    quote_chars = [
        '"', '"',  # Typografische doppelte Anführungszeichen
        '„', '‟',  # Deutsche Anführungszeichen
        ''', ''',  # Typografische einfache Anführungszeichen
        '‚', '‛',  # Weitere einfache Anführungszeichen
        '«', '»',  # Guillemets
        '‹', '›',  # Einfache Guillemets
    ]
    for char in quote_chars:
        text = text.replace(char, '"')
    
    # 3b. Bindestrich-Varianten → normaler Bindestrich
    dash_chars = [
        '–',  # En-Dash
        '—',  # Em-Dash
        '―',  # Horizontal Bar
        '‐',  # Hyphen
        '‑',  # Non-Breaking Hyphen
        '⁃',  # Hyphen Bullet
    ]
    for char in dash_chars:
        text = text.replace(char, '-')
    
    # 3c. Geschützte/spezielle Leerzeichen → normales Space
    space_chars = [
        '\u00A0',  # Non-Breaking Space
        '\u2007',  # Figure Space
        '\u2008',  # Punctuation Space
        '\u2009',  # Thin Space
        '\u200A',  # Hair Space
        '\u200B',  # Zero-Width Space
        '\u202F',  # Narrow No-Break Space
        '\u205F',  # Medium Mathematical Space
        '\u3000',  # Ideographic Space
    ]
    for char in space_chars:
        text = text.replace(char, ' ')
    
    # 4. Entfernung von Aufzählungsmarkern am Zeilenanfang
    # Arbeite zeilenweise, falls noch Zeilenumbrüche vorhanden sind
    lines = text.split('\n')
    normalized_lines = []
    
    for line in lines:
        # Entferne führende Whitespaces für Pattern-Matching
        stripped = line.lstrip()
        
        # Pattern für Aufzählungsmarker am Zeilenanfang
        # Bullets: -, *, •, +, >, #
        # Nummerierung: 1., 2., 3., ...
        # Buchstaben: a), b), c), ... oder a., b., c., ...
        
        # Bullet-Marker entfernen
        bullet_pattern = r'^[\-\*\•\+\>\#]\s+'
        stripped = re.sub(bullet_pattern, '', stripped)
        
        # Nummerierte Listen: 1. , 2. , etc.
        number_pattern = r'^\d+[\.\)]\s+'
        stripped = re.sub(number_pattern, '', stripped)
        
        # Buchstaben-Listen: a), b), a., b., etc.
        letter_pattern = r'^[a-z][\.\)]\s+'
        stripped = re.sub(letter_pattern, '', stripped)
        
        # Markdown-Überschriften: #, ##, ###, etc.
        heading_pattern = r'^#{1,6}\s+'
        stripped = re.sub(heading_pattern, '', stripped)
        
        normalized_lines.append(stripped)
    
    text = '\n'.join(normalized_lines)
    
    # 5. Entfernung dekorativer Sequenzen
    # Linien aus wiederholten Zeichen: ----, ====, ****, ~~~~, etc.
    decorative_pattern = r'[\-=\*~_]{3,}'
    text = re.sub(decorative_pattern, '', text)
    
    # 6. Whitespace-Normalisierung
    # Alle Whitespace-Arten (Space, Tab, Newline, etc.) auf einzelnes Space
    text = re.sub(r'\s+', ' ', text)
    
    # Führende und trailing Whitespaces entfernen
    text = text.strip()
    
    return text


def compute_normalized_hash(text: str) -> str:
    """
    Berechne Hash für normalisierten Text (für Exact-Dedup).
    
    Kombiniert normalize_text() mit SHA256-Hashing.
    
    Args:
        text: Eingabetext
        
    Returns:
        SHA256-Hash des normalisierten Textes
    """
    normalized = normalize_text(text)
    return hashlib.sha256(normalized.encode('utf-8')).hexdigest()


# ============================================================================
# EXACT-DEDUPLICATION AUF DOKUMENT-EBENE
# ============================================================================

def deduplicate_documents_exact(
    documents: list[dict],
    text_key: str = 'text',
    id_key: str = 'doc_id'
) -> tuple[list[dict], list[dict], dict]:
    """
    Exact-Deduplication auf Dokument-Ebene (VOR Chunking).
    
    Findet Dokumente mit identischem Inhalt nach Normalisierung und
    entfernt Duplikate. Behält jeweils das erste Dokument einer Gruppe.
    
    Args:
        documents: Liste von Dokumenten als Dictionaries
                   Erwartet mindestens: {text_key: "...", id_key: "..."}
        text_key: Schlüssel für den Textinhalt (default: 'text')
        id_key: Schlüssel für die Dokument-ID (default: 'doc_id')
    
    Returns:
        Tuple von:
        - unique_documents: Liste der behaltenen Dokumente
        - removed_documents: Liste der entfernten Duplikate
        - stats: Dictionary mit Statistiken
    
    Example:
        >>> docs = [
        ...     {"doc_id": "1", "text": "Hello World", "url": "a.html"},
        ...     {"doc_id": "2", "text": "hello world", "url": "b.html"},  # Duplikat!
        ...     {"doc_id": "3", "text": "Different text", "url": "c.html"},
        ... ]
        >>> unique, removed, stats = deduplicate_documents_exact(docs)
        >>> len(unique)
        2
        >>> stats['duplicates_removed']
        1
    """
    from collections import defaultdict
    
    if not documents:
        return [], [], {"total": 0, "unique": 0, "duplicates_removed": 0, "duplicate_groups": 0}
    
    # Hash-Berechnung mit Gruppierung
    hash_to_docs = defaultdict(list)
    
    for doc in documents:
        text = doc.get(text_key, '')
        doc_hash = compute_normalized_hash(text)
        # Speichere Hash im Dokument für spätere Referenz
        doc['_normalized_hash'] = doc_hash
        hash_to_docs[doc_hash].append(doc)
    
    # Trenne unique von Duplikaten
    unique_documents = []
    removed_documents = []
    duplicate_groups = {}
    
    for doc_hash, docs in hash_to_docs.items():
        # Erstes Dokument behalten
        unique_documents.append(docs[0])
        
        # Rest sind Duplikate
        if len(docs) > 1:
            duplicate_groups[doc_hash] = docs
            for dup_doc in docs[1:]:
                dup_doc['_kept_doc_id'] = docs[0].get(id_key, 'unknown')
                removed_documents.append(dup_doc)
    
    stats = {
        "total": len(documents),
        "unique": len(unique_documents),
        "duplicates_removed": len(removed_documents),
        "duplicate_groups": len(duplicate_groups),
        "reduction_percent": (len(removed_documents) / len(documents) * 100) if documents else 0
    }
    
    logger.info(
        f"Exact-Dedup: {stats['total']} → {stats['unique']} Dokumente "
        f"({stats['duplicates_removed']} entfernt, {stats['reduction_percent']:.1f}%)"
    )
    
    return unique_documents, removed_documents, stats


def create_dedup_excel(unique_docs: list, removed_docs: list, dedup_stats: dict, 
                       output_path: str = None) -> str:
    """
    Erstelle Excel-Übersicht für Exact-Deduplication.
    
    Formatierung entspricht der Vorlage aus deduplication_overview - Kopie.xlsx:
    - Übersicht: Header mit Stage-Tabelle und Content-Type-Statistik
    - Duplikat-Gruppen: Kompakte Darstellung mit Doc IDs als Liste
    - Entfernte Dokumente: Alle entfernten Docs mit Stichproben-Markierung
    - Stichprobe: Nur entfernte Docs zur manuellen Prüfung
    
    Args:
        unique_docs: Liste der behaltenen Dokumente
        removed_docs: Liste der entfernten Duplikate
        dedup_stats: Statistiken aus deduplicate_documents_exact()
        output_path: Optional: Pfad für Excel-Datei
    
    Returns:
        Pfad zur erstellten Excel-Datei
    """
    import pandas as pd
    from pathlib import Path
    from datetime import datetime
    import random
    
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = Path(f"data/deduplication/exact_deduplication_{timestamp}.xlsx")
    else:
        excel_path = Path(output_path)
    
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"\n   📊 Erstelle Exact-Deduplication Excel: {excel_path}")
    
    all_docs = unique_docs + removed_docs
    
    # Berechne Content-Type Statistiken
    html_original = sum(1 for d in all_docs if d.get('content_type') == 'html')
    html_unique = sum(1 for d in unique_docs if d.get('content_type') == 'html')
    html_removed = html_original - html_unique
    pdf_original = sum(1 for d in all_docs if d.get('content_type') == 'pdf')
    pdf_unique = sum(1 for d in unique_docs if d.get('content_type') == 'pdf')
    pdf_removed = pdf_original - pdf_unique
    
    # Baue Hash-Gruppen auf
    hash_to_docs = defaultdict(list)
    for doc in all_docs:
        doc_hash = doc.get('_normalized_hash', 'unknown')
        hash_to_docs[doc_hash].append(doc)
    duplicate_groups = {h: docs for h, docs in hash_to_docs.items() if len(docs) > 1}
    sorted_groups = sorted(duplicate_groups.items(), key=lambda x: len(x[1]), reverse=True)
    
    # Mapping: doc_id -> (group_index, kept_doc)
    doc_to_group = {}
    for group_idx, (hash_val, docs) in enumerate(sorted_groups, 1):
        kept_doc = docs[0]  # Erstes Dokument wird behalten
        for doc in docs:
            doc_to_group[doc.get('doc_id')] = (group_idx, kept_doc)
    
    # ================================================================
    # Sheet 1: Übersicht (mit Header und Formatierung)
    # ================================================================
    overview_rows = [
        ['🔍 Deduplication Pipeline - Übersicht', '', '', '', '', '', ''],
        [f'Erstellt am: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['Stage', 'Name', 'Input', 'Output', 'Entfernt', 'Reduktion %', 'Duplikat-Gruppen'],
        [1, 'Exact-Deduplication', dedup_stats['total'], dedup_stats['unique'], 
         dedup_stats['duplicates_removed'], f"{dedup_stats['reduction_percent']:.1f}%", 
         dedup_stats['duplicate_groups']],
        ['GESAMT', 'Alle 1 Stages', dedup_stats['total'], dedup_stats['unique'],
         dedup_stats['duplicates_removed'], f"{dedup_stats['reduction_percent']:.1f}%",
         dedup_stats['duplicate_groups']],
        ['', '', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['📁 Statistik nach Content-Type', '', '', '', '', '', ''],
        ['Content-Type', 'Original', 'Nach Dedup', 'Entfernt', 'Reduktion %', '', ''],
        ['HTML', html_original, html_unique, html_removed, 
         f"{html_removed/html_original*100:.1f}%" if html_original > 0 else '0%', '', ''],
        ['PDF', pdf_original, pdf_unique, pdf_removed,
         f"{pdf_removed/pdf_original*100:.1f}%" if pdf_original > 0 else '0%', '', ''],
    ]
    df_overview = pd.DataFrame(overview_rows)
    
    # ================================================================
    # Sheet 2: Duplikat-Gruppen (kompakte Darstellung)
    # ================================================================
    groups_rows = [
        ['🔗 Exakte Duplikat-Gruppen (Stage 1)', '', '', '', '', ''],
        [f'Gesamt: {len(duplicate_groups)} Gruppen mit {len(removed_docs)} Duplikaten', '', '', '', '', ''],
        ['', '', '', '', '', ''],
        ['Gruppe', 'Hash (kurz)', 'Anzahl Docs', 'Doc IDs', 'URLs', 'Content-Types'],
    ]
    
    for group_idx, (hash_val, docs) in enumerate(sorted_groups, 1):
        doc_ids = ', '.join(str(d.get('doc_id', '?')) for d in docs)
        urls = '\n'.join(d.get('url', '')[:80] + ('...' if len(d.get('url', '')) > 80 else '') for d in docs[:5])
        if len(docs) > 5:
            urls += f'\n... +{len(docs) - 5} weitere'
        content_types = ', '.join(set(d.get('content_type', '').upper() for d in docs))
        
        groups_rows.append([
            group_idx, 
            hash_val[:16] + '...', 
            len(docs), 
            doc_ids, 
            urls, 
            content_types
        ])
    
    df_groups = pd.DataFrame(groups_rows)
    
    # ================================================================
    # Sheet 3: Entfernte Dokumente (mit Stichproben-Markierung)
    # ================================================================
    # WICHTIG: Die entfernten Dokumente müssen in derselben Reihenfolge wie in
    # deduplication_implementation.py sein: Sortiert nach Duplikat-Gruppen (größte zuerst),
    # innerhalb jeder Gruppe ab dem zweiten Dokument.
    
    # Baue sortierte Liste der entfernten Dokumente (wie in deduplication_implementation.py)
    all_removed_sorted = []
    for group_idx, (hash_val, docs) in enumerate(sorted_groups, 1):
        kept_doc = docs[0]  # Erstes Dokument wird behalten
        for doc in docs[1:]:  # Rest sind Duplikate
            all_removed_sorted.append({
                'doc': doc,
                'group_idx': group_idx,
                'kept_doc': kept_doc
            })
    
    # Stichprobe bestimmen - WICHTIG: Algorithmus muss exakt deduplication_implementation.py entsprechen!
    # 1. Erst 20 aus ALLEN entfernten Dokumenten ziehen (nicht getrennt nach Typ!)
    # 2. Dann zusätzlich 2 PDFs, die noch nicht in der Stichprobe sind
    random.seed(42)
    
    # Index-Listen für PDF-Tracking (basierend auf sortierter Liste!)
    pdf_indices = {i for i, item in enumerate(all_removed_sorted) if item['doc'].get('content_type') == 'pdf'}
    
    # Schritt 1: Ziehe 20 Dokumente aus allen (reproduziert alte Stichprobe)
    base_sample_size = min(20, len(all_removed_sorted))
    base_sample_indices = set(random.sample(range(len(all_removed_sorted)), base_sample_size)) if all_removed_sorted else set()
    
    # Schritt 2: Ziehe zusätzlich 2 PDFs (die noch nicht in der Stichprobe sind)
    pdf_not_in_sample = [i for i in pdf_indices if i not in base_sample_indices]
    additional_pdf_sample = set()
    if pdf_not_in_sample:
        pdf_sample_size = min(2, len(pdf_not_in_sample))
        additional_pdf_sample = set(random.sample(pdf_not_in_sample, pdf_sample_size))
    
    sample_indices = base_sample_indices | additional_pdf_sample
    sample_doc_ids = {all_removed_sorted[i]['doc'].get('doc_id') for i in sample_indices}
    
    removed_rows = [
        ['🗑️ Alle entfernten Dokumente', '', '', '', '', '', '', '', ''],
        [f'Stichprobe: {len([i for i in sample_indices if all_removed_sorted[i]["doc"].get("content_type") == "html"])} HTML + {len([i for i in sample_indices if all_removed_sorted[i]["doc"].get("content_type") == "pdf"])} PDF (Seed: 42)', '', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '', ''],
        ['Stichprobe', 'Stage', 'Gruppe', 'Doc ID', 'Content-Type', 'Entfernte URL', 'Titel', 'Zeichen', 'Ersetzt durch URL'],
    ]
    
    for idx, item in enumerate(all_removed_sorted):
        doc = item['doc']
        doc_id = doc.get('doc_id')
        group_idx = item['group_idx']
        kept_doc = item['kept_doc']
        is_sample = idx in sample_indices
        
        removed_rows.append([
            '✓ PRÜFEN' if is_sample else '',
            'Exact-Deduplication',
            group_idx,
            doc_id,
            doc.get('content_type', '').upper(),
            doc.get('url', ''),
            (doc.get('title', '') or '')[:60],
            len(doc.get('text', '')),
            kept_doc.get('url', '') if kept_doc else ''
        ])
    
    df_removed = pd.DataFrame(removed_rows)
    
    # ================================================================
    # Sheet 4: Stichprobe (nur entfernte Docs zur manuellen Prüfung)
    # ================================================================
    # Sample-Docs aus der sortierten Liste extrahieren
    sample_items = [all_removed_sorted[i] for i in sorted(sample_indices)]
    sample_count_html = len([item for item in sample_items if item['doc'].get('content_type') == 'html'])
    sample_count_pdf = len([item for item in sample_items if item['doc'].get('content_type') == 'pdf'])
    
    sample_rows = [
        ['🔍 Stichprobe zur manuellen Überprüfung', '', '', '', '', ''],
        [f'Seed: 42 | Stichprobe: {sample_count_html} HTML + {sample_count_pdf} PDF = {len(sample_items)} von {len(all_removed_sorted)} Dokumenten', '', '', '', '', ''],
        ['Bitte überprüfen Sie, ob die entfernten Dokumente tatsächlich Duplikate der Ersatz-Dokumente sind.', '', '', '', '', ''],
        ['', '', '', '', '', ''],
        ['#', 'Typ', 'Doc ID', 'Entfernte URL', 'Ersetzt durch URL', 'Korrekt? (J/N)', 'Anmerkung'],
    ]
    
    for i, item in enumerate(sample_items, 1):
        doc = item['doc']
        kept_doc = item['kept_doc']
        
        sample_rows.append([
            i,
            doc.get('content_type', '').upper(),
            doc.get('doc_id', ''),
            doc.get('url', ''),
            kept_doc.get('url', '') if kept_doc else '',
            '',  # Korrekt? (J/N) - leer für manuelle Eingabe
            ''   # Anmerkung - leer für manuelle Eingabe
        ])
    
    df_sample = pd.DataFrame(sample_rows)
    
    # Speichere Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_overview.to_excel(writer, sheet_name='Übersicht', index=False, header=False)
        df_groups.to_excel(writer, sheet_name='Duplikat-Gruppen', index=False, header=False)
        df_removed.to_excel(writer, sheet_name='Entfernte Dokumente', index=False, header=False)
        df_sample.to_excel(writer, sheet_name='Stichprobe', index=False, header=False)
        
        # ================================================================
        # Formatierungen anwenden (wie im Template)
        # ================================================================
        wb = writer.book
        
        # Farben definieren (wie im Template)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Dunkelblau
        total_fill = PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')   # Hellblau
        sample_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Gelb
        header_font = Font(bold=True, color='FFFFFF')  # Weiß
        total_font = Font(bold=True)
        
        # --- Sheet 1: Übersicht ---
        ws_overview = wb['Übersicht']
        # Zeile 1: Titel (Bold)
        ws_overview.cell(row=1, column=1).font = Font(bold=True)
        # Zeile 4: Header (Stage, Name, Input, Output, Entfernt, ...)
        for col in range(1, 8):
            cell = ws_overview.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Zeile 6: GESAMT
        for col in range(1, 8):
            cell = ws_overview.cell(row=6, column=col)
            cell.fill = total_fill
            cell.font = total_font
        # Zeile 9: Statistik-Titel (Bold)
        ws_overview.cell(row=9, column=1).font = Font(bold=True)
        # Zeile 10: Content-Type Header
        for col in range(1, 6):
            cell = ws_overview.cell(row=10, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # --- Sheet 2: Duplikat-Gruppen ---
        ws_groups = wb['Duplikat-Gruppen']
        # Zeile 4: Header
        for col in range(1, 7):
            cell = ws_groups.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Zeile 1: Titel (Bold)
        ws_groups.cell(row=1, column=1).font = Font(bold=True)
        
        # --- Sheet 3: Entfernte Dokumente ---
        ws_removed = wb['Entfernte Dokumente']
        # Zeile 4: Header
        for col in range(1, 10):
            cell = ws_removed.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Zeile 1: Titel (Bold)
        ws_removed.cell(row=1, column=1).font = Font(bold=True)
        # Gelbe Markierung für Stichproben-Zeilen
        for row_idx in range(5, ws_removed.max_row + 1):
            if ws_removed.cell(row=row_idx, column=1).value == '✓ PRÜFEN':
                for col in range(1, 10):
                    ws_removed.cell(row=row_idx, column=col).fill = sample_fill
        
        # --- Sheet 4: Stichprobe ---
        ws_sample = wb['Stichprobe']
        # Zeile 5: Header
        for col in range(1, 8):
            cell = ws_sample.cell(row=5, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Zeile 1: Titel (Bold)
        ws_sample.cell(row=1, column=1).font = Font(bold=True)
        # Hellgrüne Eingabefelder für Spalten 6-7 (Korrekt? und Anmerkung)
        input_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for row_idx in range(6, ws_sample.max_row + 1):
            if ws_sample.cell(row=row_idx, column=1).value:  # Nur Datenzeilen
                ws_sample.cell(row=row_idx, column=6).fill = input_fill
                ws_sample.cell(row=row_idx, column=7).fill = input_fill
    
    print(f"   ✅ Excel erstellt: {len(duplicate_groups)} Gruppen, {len(all_removed_sorted)} entfernt, {len(sample_items)} in Stichprobe")
    
    return str(excel_path)

    """
    Dedupliziert Inhalte basierend auf Similarity-Hashing.
    
    Verwendet Shingling und MinHash für effiziente near-duplicate Erkennung.
    """
    
    def __init__(self, similarity_threshold: float = 0.85, shingle_size: int = 3):
        """
        Initialisiere den Deduplicator.
        
        Args:
            similarity_threshold: Schwellwert für Ähnlichkeit (0.0-1.0)
            shingle_size: Größe der Shingles für Vergleich
        """
        self.similarity_threshold = similarity_threshold
        self.shingle_size = shingle_size
        self.seen_fingerprints: Set[str] = set()
        self.url_to_fingerprint: dict = {}
        
        # Quick-Win Optimierungen
        self.shingle_cache: dict = {}  # Cache für Shingles
        self.chunks_by_size: dict = defaultdict(list)  # Size-Bucketing
        
    def create_shingles(self, text: str) -> Set[str]:
        """
        Erstelle Shingles (n-grams) aus Text mit Caching.
        
        Args:
            text: Eingabetext
            
        Returns:
            Set von Shingles
        """
        # Quick-Win 1: Shingle-Cache
        text_hash = hash(text)
        if text_hash in self.shingle_cache:
            return self.shingle_cache[text_hash]
        
        # Normalisiere Text
        text = text.lower().strip()
        words = text.split()
        
        # Erstelle Wort-Shingles
        shingles = set()
        for i in range(len(words) - self.shingle_size + 1):
            shingle = " ".join(words[i:i + self.shingle_size])
            shingles.add(shingle)
        
        # Cache speichern
        self.shingle_cache[text_hash] = shingles
        return shingles
    
    def compute_content_hash(self, text: str, use_full_normalization: bool = True) -> str:
        """
        Berechne eindeutigen Hash für Inhalt.
        
        Args:
            text: Eingabetext
            use_full_normalization: Wenn True, nutze normalize_text() für robuste Normalisierung.
                                    Wenn False, nur lower().strip() (Legacy-Verhalten).
            
        Returns:
            SHA256-Hash als Hex-String
        """
        if use_full_normalization:
            normalized = normalize_text(text)
        else:
            normalized = text.lower().strip()
        return hashlib.sha256(normalized.encode('utf-8')).hexdigest()
    
    def compute_shingles_hash(self, shingles: Set[str]) -> str:
        """
        Berechne Hash für Shingles-Set.
        
        Args:
            shingles: Set von Shingles
            
        Returns:
            Hash-Repräsentation
        """
        # Sortiere für konsistenten Hash
        sorted_shingles = sorted(shingles)
        combined = "".join(sorted_shingles)
        return hashlib.md5(combined.encode('utf-8')).hexdigest()
    
    def jaccard_similarity(self, set1: Set[str], set2: Set[str]) -> float:
        """
        Berechne Jaccard-Ähnlichkeit zwischen zwei Sets mit Early Exit.
        
        Args:
            set1: Erstes Set
            set2: Zweites Set
            
        Returns:
            Ähnlichkeit zwischen 0.0 und 1.0
        """
        if not set1 and not set2:
            return 1.0
        
        # Quick-Win 2: Early Exit - prüfe maximale mögliche Similarity
        min_size = min(len(set1), len(set2))
        max_size = max(len(set1), len(set2))
        
        if max_size > 0:
            max_possible_similarity = min_size / max_size
            if max_possible_similarity < self.similarity_threshold:
                return 0.0  # Kann nie Threshold erreichen
        
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        
        return intersection / union if union > 0 else 0.0
    
    def is_duplicate(self, text: str, url: str) -> Tuple[bool, str]:
        """
        Prüfe ob Text ein Duplikat ist mit Size-Bucketing.
        
        Args:
            text: Zu prüfender Text
            url: URL des Dokuments
            
        Returns:
            Tuple von (ist_duplikat, grund)
        """
        # Exakte Duplikate
        content_hash = self.compute_content_hash(text)
        if content_hash in self.seen_fingerprints:
            return True, "exact_duplicate"
        
        # Quick-Win 3: Size-Bucketing - nur ähnlich große Texte vergleichen
        text_size = len(text)
        size_bucket = text_size // 500  # Buckets von 500 Zeichen
        
        # Kandidaten: Aktueller Bucket ± 1
        candidates = []
        for bucket in [size_bucket - 1, size_bucket, size_bucket + 1]:
            candidates.extend(self.chunks_by_size.get(bucket, []))
        
        # Near-duplicates - nur gegen Kandidaten
        shingles = self.create_shingles(text)
        
        for candidate_url in candidates:
            if candidate_url not in self.url_to_fingerprint:
                continue
                
            candidate_text = self.url_to_fingerprint[candidate_url].get('text', '')
            seen_shingles = self.create_shingles(candidate_text)
            
            similarity = self.jaccard_similarity(shingles, seen_shingles)
            
            if similarity >= self.similarity_threshold:
                logger.info(
                    f"Near-duplicate gefunden: {url} ähnlich zu {candidate_url} "
                    f"(Similarity: {similarity:.2f})"
                )
                return True, f"near_duplicate_{similarity:.2f}"
        
        # Kein Duplikat - speichere Fingerprint UND Size-Bucket
        self.seen_fingerprints.add(content_hash)
        self.url_to_fingerprint[url] = {
            'content_hash': content_hash,
            'shingles_hash': self.compute_shingles_hash(shingles),
            'text': text[:5000],
            'word_count': len(text.split())
        }
        self.chunks_by_size[size_bucket].append(url)
        
        return False, "unique"
    
    def deduplicate_batch(self, documents: List[dict]) -> Tuple[List[dict], List[dict]]:
        """
        Dedupliziere eine Batch von Dokumenten.
        
        Args:
            documents: Liste von Dokumenten mit 'url' und 'content' Keys
            
        Returns:
            Tuple von (unique_documents, duplicate_documents)
        """
        unique = []
        duplicates = []
        
        for doc in documents:
            url = doc.get('url', '')
            content = doc.get('content', '')
            
            is_dup, reason = self.is_duplicate(content, url)
            
            if is_dup:
                doc['duplicate_reason'] = reason
                duplicates.append(doc)
            else:
                unique.append(doc)
        
        logger.info(
            f"Deduplication: {len(unique)} unique, {len(duplicates)} duplicates "
            f"von {len(documents)} gesamt"
        )
        
        return unique, duplicates
    
    def get_statistics(self) -> dict:
        """
        Erhalte Statistiken über gesehene Dokumente.
        
        Returns:
            Dictionary mit Statistiken
        """
        return {
            'total_seen': len(self.seen_fingerprints),
            'unique_urls': len(self.url_to_fingerprint),
            'similarity_threshold': self.similarity_threshold,
            'shingle_size': self.shingle_size
        }


# ============================================================================
# DOCUMENT-LEVEL NEAR-DEDUPLICATION
# ============================================================================

def _create_word_shingles(text: str, k: int = 5) -> Set[str]:
    """
    Erstelle Wort-Shingles (k-grams) aus normalisiertem Text.
    
    Args:
        text: Eingabetext
        k: Shingle-Größe (Anzahl Wörter pro Shingle)
        
    Returns:
        Set von Shingles
    """
    # Normalisiere Text
    normalized = normalize_text(text)
    words = normalized.split()
    
    if len(words) < k:
        # Wenn weniger Wörter als k, ganzen Text als ein Shingle
        return {normalized} if words else set()
    
    shingles = set()
    for i in range(len(words) - k + 1):
        shingle = " ".join(words[i:i + k])
        shingles.add(shingle)
    
    return shingles


def _jaccard_similarity(set1: Set[str], set2: Set[str]) -> float:
    """
    Berechne Jaccard-Ähnlichkeit zwischen zwei Sets.
    
    Args:
        set1: Erstes Set
        set2: Zweites Set
        
    Returns:
        Ähnlichkeit zwischen 0.0 und 1.0
    """
    if not set1 and not set2:
        return 1.0
    if not set1 or not set2:
        return 0.0
    
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    
    return intersection / union if union > 0 else 0.0


def _select_canonical_url(docs: list, id_key: str = 'doc_id') -> dict:
    """
    Wähle das kanonische Dokument aus einem Cluster.
    
    Kriterien (in Reihenfolge):
    1. Bevorzuge URLs ohne Query-Parameter (?), print, search, etc.
    2. Bevorzuge längeren Text
    3. Im Zweifelsfall: niedrigste doc_id
    
    Args:
        docs: Liste von Dokumenten im Cluster
        id_key: Schlüssel für die Dokument-ID
        
    Returns:
        Das kanonische Dokument
    """
    def url_quality_score(doc):
        """Höherer Score = bessere URL."""
        url = doc.get('url', '').lower()
        score = 0
        
        # Malus für problematische URL-Patterns
        if '?' in url:
            score -= 10
        if 'print' in url:
            score -= 5
        if 'search' in url:
            score -= 5
        if 'mobile' in url:
            score -= 3
        if 'redirect' in url:
            score -= 3
        
        # Bonus für kürzere, sauberere URLs
        score -= len(url) // 50  # Leichter Malus für sehr lange URLs
        
        return score
    
    def text_length(doc):
        """Textlänge als Sekundärkriterium."""
        return len(doc.get('text', ''))
    
    def doc_id_sortable(doc):
        """Doc-ID für deterministisches Tie-Breaking."""
        doc_id = doc.get(id_key, 'zzz')
        try:
            return int(doc_id)
        except (ValueError, TypeError):
            return float('inf')
    
    # Sortiere nach: URL-Qualität (absteigend), Textlänge (absteigend), Doc-ID (aufsteigend)
    sorted_docs = sorted(
        docs,
        key=lambda d: (-url_quality_score(d), -text_length(d), doc_id_sortable(d))
    )
    
    return sorted_docs[0]


class UnionFind:
    """
    Union-Find Datenstruktur für Cluster-Bildung.
    
    Verwendet für effizientes Clustering von Near-Duplicate-Paaren.
    """
    
    def __init__(self):
        self.parent = {}
        self.rank = {}
    
    def find(self, x):
        """Finde den Wurzelknoten mit Pfadkompression."""
        if x not in self.parent:
            self.parent[x] = x
            self.rank[x] = 0
        
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]
    
    def union(self, x, y):
        """Vereinige zwei Sets (Union by Rank)."""
        root_x = self.find(x)
        root_y = self.find(y)
        
        if root_x == root_y:
            return
        
        if self.rank[root_x] < self.rank[root_y]:
            self.parent[root_x] = root_y
        elif self.rank[root_x] > self.rank[root_y]:
            self.parent[root_y] = root_x
        else:
            self.parent[root_y] = root_x
            self.rank[root_x] += 1
    
    def get_clusters(self) -> dict:
        """Erhalte alle Cluster als Dictionary {root: [members]}."""
        clusters = defaultdict(list)
        for item in self.parent:
            root = self.find(item)
            clusters[root].append(item)
        return dict(clusters)


def deduplicate_documents_near(
    documents: list[dict],
    text_key: str = 'text',
    id_key: str = 'doc_id',
    shingle_k: int = 5,
    similarity_threshold: float = 0.90,
    min_words: int = 120
) -> tuple[list[dict], list[dict], dict]:
    """
    Near-Deduplication auf Dokument-Ebene (nach Exact-Dedup, vor Chunking).
    
    Findet inhaltlich nahezu identische Dokumente mittels Wort-Shingling und
    Jaccard-Ähnlichkeit.
    
    Vergleicht nur Dokumente mit gleichem content_type (HTML↔HTML, PDF↔PDF).
    Verwendet Size-Bucketing zur Reduzierung der Kandidatenpaare.
    
    Args:
        documents: Liste von Dokumenten als Dictionaries
                   Erwartet: {text_key: "...", id_key: "...", 'content_type': "..."}
        text_key: Schlüssel für den Textinhalt (default: 'text')
        id_key: Schlüssel für die Dokument-ID (default: 'doc_id')
        shingle_k: Größe der Wort-Shingles (default: 5)
        similarity_threshold: Schwellwert für Content Near-Duplicate (default: 0.90)
        min_words: Minimale Wortanzahl für Vergleich (default: 120)
    
    Returns:
        Tuple von:
        - unique_documents: Liste der behaltenen Dokumente
        - removed_documents: Liste der entfernten Near-Duplicates
        - stats: Dictionary mit Statistiken
    
    Example:
        >>> docs = [
        ...     {"doc_id": "1", "text": "...", "title": "PO BWL 2024", "content_type": "html"},
        ...     {"doc_id": "2", "text": "...", "title": "PO BWL 2024", "content_type": "html"},  # Near-dup!
        ... ]
        >>> unique, removed, stats = deduplicate_documents_near(docs)
    """
    if not documents:
        return [], [], {
            "total": 0, "unique": 0, "duplicates_removed": 0,
            "clusters": 0, "candidate_pairs": 0, "verified_pairs": 0,
            "reduction_percent": 0.0
        }
    
    logger.info(f"Near-Dedup: Starte mit {len(documents)} Dokumenten (k={shingle_k}, threshold={similarity_threshold})")
    
    # ================================================================
    # SCHRITT 1: Dokumente nach content_type gruppieren
    # ================================================================
    docs_by_type = defaultdict(list)
    for doc in documents:
        content_type = doc.get('content_type', 'unknown')
        docs_by_type[content_type].append(doc)
    
    logger.info(f"   Content-Types: {', '.join(f'{ct}={len(docs)}' for ct, docs in docs_by_type.items())}")
    
    # ================================================================
    # SCHRITT 2: Shingles berechnen und Size-Bucketing
    # ================================================================
    doc_shingles = {}  # doc_id -> shingles
    doc_word_counts = {}  # doc_id -> word_count
    
    # Size-Buckets pro content_type (Bucket-Größe: 500 Wörter)
    BUCKET_SIZE = 500
    size_buckets = defaultdict(lambda: defaultdict(list))  # content_type -> bucket -> [doc_ids]
    
    for doc in documents:
        doc_id = doc.get(id_key)
        text = doc.get(text_key, '')
        content_type = doc.get('content_type', 'unknown')
        
        # Normalisiere und zähle Wörter
        normalized = normalize_text(text)
        words = normalized.split()
        word_count = len(words)
        doc_word_counts[doc_id] = word_count
        
        # Nur Dokumente mit >= min_words für Vergleich
        if word_count >= min_words:
            shingles = _create_word_shingles(text, k=shingle_k)
            doc_shingles[doc_id] = shingles
            
            # Size-Bucket zuweisen
            bucket = word_count // BUCKET_SIZE
            size_buckets[content_type][bucket].append(doc_id)
    
    logger.info(f"   {len(doc_shingles)} Dokumente mit >= {min_words} Wörtern für Vergleich")
    
    # ================================================================
    # SCHRITT 3: Kandidatenpaare generieren (Size-Bucketing)
    # ================================================================
    candidate_pairs = set()
    
    for content_type, buckets in size_buckets.items():
        for bucket, doc_ids in buckets.items():
            # Vergleiche innerhalb des Buckets
            for i, doc_id1 in enumerate(doc_ids):
                for doc_id2 in doc_ids[i+1:]:
                    candidate_pairs.add((doc_id1, doc_id2))
            
            # Vergleiche mit benachbarten Buckets (±1)
            for neighbor_bucket in [bucket - 1, bucket + 1]:
                if neighbor_bucket in buckets:
                    for doc_id1 in doc_ids:
                        for doc_id2 in buckets[neighbor_bucket]:
                            if doc_id1 < doc_id2:  # Vermeidung von Duplikaten
                                candidate_pairs.add((doc_id1, doc_id2))
                            else:
                                candidate_pairs.add((doc_id2, doc_id1))
    
    logger.info(f"   {len(candidate_pairs):,} Kandidatenpaare generiert")
    
    # ================================================================
    # SCHRITT 4: Jaccard-Similarity prüfen und Cluster bilden
    # ================================================================
    union_find = UnionFind()
    verified_pairs = []
    
    # Mapping doc_id -> doc für schnellen Zugriff
    id_to_doc = {doc.get(id_key): doc for doc in documents}
    
    for doc_id1, doc_id2 in candidate_pairs:
        shingles1 = doc_shingles.get(doc_id1)
        shingles2 = doc_shingles.get(doc_id2)
        
        if shingles1 is None or shingles2 is None:
            continue
        
        # Early Exit: Prüfe ob Jaccard-Threshold überhaupt erreichbar
        min_size = min(len(shingles1), len(shingles2))
        max_size = max(len(shingles1), len(shingles2))
        
        if max_size > 0:
            max_possible = min_size / max_size
            if max_possible < similarity_threshold:
                continue
        
        # Berechne Content Jaccard-Similarity
        content_similarity = _jaccard_similarity(shingles1, shingles2)
        
        if content_similarity >= similarity_threshold:
            verified_pairs.append((doc_id1, doc_id2, content_similarity))
            union_find.union(doc_id1, doc_id2)
    
    logger.info(f"   {len(verified_pairs)} verifizierte Near-Duplicate-Paare")
    
    # ================================================================
    # SCHRITT 5: Cluster extrahieren und Canonical wählen
    # ================================================================
    raw_clusters = union_find.get_clusters()
    
    # Nur Cluster mit > 1 Element sind relevant
    duplicate_clusters = {k: v for k, v in raw_clusters.items() if len(v) > 1}
    
    # id_to_doc wurde bereits in SCHRITT 4 erstellt
    
    # Pro Cluster: Canonical wählen, Rest als Duplikate markieren
    canonical_doc_ids = set()
    removed_doc_ids = set()
    cluster_info = []  # Für Excel-Report
    
    for cluster_idx, (root, member_ids) in enumerate(sorted(duplicate_clusters.items()), 1):
        member_docs = [id_to_doc[mid] for mid in member_ids if mid in id_to_doc]
        
        if len(member_docs) < 2:
            continue
        
        # Wähle Canonical
        canonical = _select_canonical_url(member_docs, id_key=id_key)
        canonical_id = canonical.get(id_key)
        canonical_doc_ids.add(canonical_id)
        
        # Rest sind Duplikate
        for doc in member_docs:
            doc_id = doc.get(id_key)
            if doc_id != canonical_id:
                removed_doc_ids.add(doc_id)
                # Annotiere für spätere Dokumentation
                doc['_kept_doc_id'] = canonical_id
                doc['_near_duplicate_of'] = canonical.get('url', '')
        
        cluster_info.append({
            'cluster_idx': cluster_idx,
            'canonical_id': canonical_id,
            'canonical_url': canonical.get('url', ''),
            'member_ids': member_ids,
            'size': len(member_docs)
        })
    
    logger.info(f"   {len(duplicate_clusters)} Cluster mit {len(removed_doc_ids)} zu entfernenden Dokumenten")
    
    # ================================================================
    # SCHRITT 6: Unique und Removed Listen erstellen
    # ================================================================
    unique_documents = []
    removed_documents = []
    
    for doc in documents:
        doc_id = doc.get(id_key)
        if doc_id in removed_doc_ids:
            removed_documents.append(doc)
        else:
            unique_documents.append(doc)
    
    # ================================================================
    # SCHRITT 7: Statistiken
    # ================================================================
    stats = {
        "total": len(documents),
        "unique": len(unique_documents),
        "duplicates_removed": len(removed_documents),
        "clusters": len(duplicate_clusters),
        "candidate_pairs": len(candidate_pairs),
        "verified_pairs": len(verified_pairs),
        "reduction_percent": (len(removed_documents) / len(documents) * 100) if documents else 0,
        "shingle_k": shingle_k,
        "similarity_threshold": similarity_threshold,
        "min_words": min_words,
        "_cluster_info": cluster_info  # Für Excel-Report
    }
    
    logger.info(
        f"Near-Dedup: {stats['total']} → {stats['unique']} Dokumente "
        f"({stats['duplicates_removed']} entfernt, {stats['reduction_percent']:.1f}%)"
    )
    
    return unique_documents, removed_documents, stats


def create_near_dedup_excel(unique_docs: list, removed_docs: list, near_dedup_stats: dict,
                            output_path: str = None) -> str:
    """
    Erstelle Excel-Übersicht für Near-Deduplication.
    
    Formatierung analog zu create_dedup_excel():
    - Übersicht: Header mit Stage-Tabelle und Content-Type-Statistik
    - Near-Duplicate-Cluster: Kompakte Darstellung mit Canonical und Members
    - Entfernte Dokumente: Alle entfernten Docs mit Stichproben-Markierung
    - Stichprobe: Nur entfernte Docs zur manuellen Prüfung
    
    Args:
        unique_docs: Liste der behaltenen Dokumente
        removed_docs: Liste der entfernten Near-Duplicates
        near_dedup_stats: Statistiken aus deduplicate_documents_near()
        output_path: Optional: Pfad für Excel-Datei
    
    Returns:
        Pfad zur erstellten Excel-Datei
    """
    import pandas as pd
    from pathlib import Path
    from datetime import datetime
    import random
    
    if output_path is None:
        excel_path = Path("src/advanced_rag/data/near_deduplication_overview.xlsx")
    else:
        excel_path = Path(output_path)
    
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"\n   📊 Erstelle Near-Deduplication Excel: {excel_path}")
    
    all_docs = unique_docs + removed_docs
    
    # Berechne Content-Type Statistiken
    html_original = sum(1 for d in all_docs if d.get('content_type') == 'html')
    html_unique = sum(1 for d in unique_docs if d.get('content_type') == 'html')
    html_removed = html_original - html_unique
    pdf_original = sum(1 for d in all_docs if d.get('content_type') == 'pdf')
    pdf_unique = sum(1 for d in unique_docs if d.get('content_type') == 'pdf')
    pdf_removed = pdf_original - pdf_unique
    
    cluster_info = near_dedup_stats.get('_cluster_info', [])
    
    # ================================================================
    # Sheet 1: Übersicht (mit Header und Formatierung)
    # ================================================================
    overview_rows = [
        ['🔍 Near-Deduplication Pipeline - Übersicht', '', '', '', '', '', ''],
        [f'Erstellt am: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['Parameter', 'Wert', '', '', '', '', ''],
        ['Shingle-Größe (k)', near_dedup_stats.get('shingle_k', 5), '', '', '', '', ''],
        ['Content-Similarity-Threshold', near_dedup_stats.get('similarity_threshold', 0.90), '', '', '', '', ''],
        ['Min. Wörter', near_dedup_stats.get('min_words', 120), '', '', '', '', ''],
        ['Kandidatenpaare', f"{near_dedup_stats.get('candidate_pairs', 0):,}", '', '', '', '', ''],
        ['Verifizierte Paare', near_dedup_stats.get('verified_pairs', 0), '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['Stage', 'Name', 'Input', 'Output', 'Entfernt', 'Reduktion %', 'Cluster'],
        [2, 'Near-Deduplication', near_dedup_stats['total'], near_dedup_stats['unique'], 
         near_dedup_stats['duplicates_removed'], f"{near_dedup_stats['reduction_percent']:.1f}%", 
         near_dedup_stats['clusters']],
        ['', '', '', '', '', '', ''],
        ['📁 Statistik nach Content-Type', '', '', '', '', '', ''],
        ['Content-Type', 'Original', 'Nach Dedup', 'Entfernt', 'Reduktion %', '', ''],
        ['HTML', html_original, html_unique, html_removed, 
         f"{html_removed/html_original*100:.1f}%" if html_original > 0 else '0%', '', ''],
        ['PDF', pdf_original, pdf_unique, pdf_removed,
         f"{pdf_removed/pdf_original*100:.1f}%" if pdf_original > 0 else '0%', '', ''],
    ]
    df_overview = pd.DataFrame(overview_rows)
    
    # ================================================================
    # Sheet 2: Near-Duplicate-Cluster (kompakte Darstellung)
    # ================================================================
    clusters_rows = [
        ['🔗 Near-Duplicate-Cluster (Stage 2)', '', '', '', '', ''],
        [f'Gesamt: {len(cluster_info)} Cluster mit {len(removed_docs)} entfernten Dokumenten', '', '', '', '', ''],
        ['', '', '', '', '', ''],
        ['Cluster', 'Canonical Doc ID', 'Canonical URL', 'Anzahl Docs', 'Member IDs', 'Content-Types'],
    ]
    
    # Mapping doc_id -> doc
    id_to_doc = {doc.get('doc_id'): doc for doc in all_docs}
    
    for info in cluster_info:
        member_ids_str = ', '.join(str(mid) for mid in info['member_ids'])
        member_docs = [id_to_doc.get(mid) for mid in info['member_ids'] if mid in id_to_doc]
        content_types = ', '.join(set(d.get('content_type', '').upper() for d in member_docs if d))
        
        clusters_rows.append([
            info['cluster_idx'],
            info['canonical_id'],
            info['canonical_url'][:80] + ('...' if len(info['canonical_url']) > 80 else ''),
            info['size'],
            member_ids_str,
            content_types
        ])
    
    df_clusters = pd.DataFrame(clusters_rows)
    
    # ================================================================
    # Sheet 3: Entfernte Dokumente (mit Stichproben-Markierung)
    # ================================================================
    # Stichprobe bestimmen (deterministisch mit seed=42)
    random.seed(42)
    
    sample_size = min(20, len(removed_docs))
    sample_indices = set(random.sample(range(len(removed_docs)), sample_size)) if removed_docs else set()
    
    removed_rows = [
        ['🗑️ Alle entfernten Near-Duplicate-Dokumente', '', '', '', '', '', '', '', ''],
        [f'Stichprobe: {len(sample_indices)} Dokumente (Seed: 42)', '', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '', ''],
        ['Stichprobe', 'Stage', 'Cluster', 'Doc ID', 'Content-Type', 'Entfernte URL', 'Titel', 'Zeichen', 'Ersetzt durch URL'],
    ]
    
    for idx, doc in enumerate(removed_docs):
        doc_id = doc.get('doc_id')
        is_sample = idx in sample_indices
        
        # Finde Cluster-Index
        cluster_idx = 0
        for info in cluster_info:
            if doc_id in info['member_ids']:
                cluster_idx = info['cluster_idx']
                break
        
        removed_rows.append([
            '✓ PRÜFEN' if is_sample else '',
            'Near-Deduplication',
            cluster_idx,
            doc_id,
            doc.get('content_type', '').upper(),
            doc.get('url', ''),
            (doc.get('title', '') or '')[:60],
            len(doc.get('text', '')),
            doc.get('_near_duplicate_of', '')
        ])
    
    df_removed = pd.DataFrame(removed_rows)
    
    # ================================================================
    # Sheet 4: Stichprobe (nur entfernte Docs zur manuellen Prüfung)
    # ================================================================
    sample_docs = [removed_docs[i] for i in sorted(sample_indices)]
    
    sample_rows = [
        ['🔍 Stichprobe zur manuellen Überprüfung', '', '', '', '', ''],
        [f'Seed: 42 | Stichprobe: {len(sample_docs)} von {len(removed_docs)} Dokumenten', '', '', '', '', ''],
        ['Bitte überprüfen Sie, ob die entfernten Dokumente tatsächlich Near-Duplicates der Canonical-Dokumente sind.', '', '', '', '', ''],
        ['', '', '', '', '', ''],
        ['#', 'Typ', 'Doc ID', 'Entfernte URL', 'Ersetzt durch URL', 'Korrekt? (J/N)', 'Anmerkung'],
    ]
    
    for i, doc in enumerate(sample_docs, 1):
        sample_rows.append([
            i,
            doc.get('content_type', '').upper(),
            doc.get('doc_id', ''),
            doc.get('url', ''),
            doc.get('_near_duplicate_of', ''),
            '',  # Korrekt? (J/N) - leer für manuelle Eingabe
            ''   # Anmerkung - leer für manuelle Eingabe
        ])
    
    df_sample = pd.DataFrame(sample_rows)
    
    # Speichere Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_overview.to_excel(writer, sheet_name='Übersicht', index=False, header=False)
        df_clusters.to_excel(writer, sheet_name='Near-Duplicate-Cluster', index=False, header=False)
        df_removed.to_excel(writer, sheet_name='Entfernte Dokumente', index=False, header=False)
        df_sample.to_excel(writer, sheet_name='Stichprobe', index=False, header=False)
        
        # ================================================================
        # Formatierungen anwenden (wie im Template)
        # ================================================================
        wb = writer.book
        
        # Farben definieren (wie im Template)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Dunkelblau
        total_fill = PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')   # Hellblau
        sample_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Gelb
        header_font = Font(bold=True, color='FFFFFF')  # Weiß
        total_font = Font(bold=True)
        
        # --- Sheet 1: Übersicht ---
        ws_overview = wb['Übersicht']
        ws_overview.cell(row=1, column=1).font = Font(bold=True)
        # Parameter-Header
        for col in range(1, 3):
            cell = ws_overview.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Stage-Header (verschoben um 2 Zeilen wegen neuer Titel-Parameter)
        for col in range(1, 8):
            cell = ws_overview.cell(row=13, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Statistik-Titel
        ws_overview.cell(row=16, column=1).font = Font(bold=True)
        # Content-Type Header
        for col in range(1, 6):
            cell = ws_overview.cell(row=17, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # --- Sheet 2: Cluster ---
        ws_clusters = wb['Near-Duplicate-Cluster']
        ws_clusters.cell(row=1, column=1).font = Font(bold=True)
        for col in range(1, 7):
            cell = ws_clusters.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # --- Sheet 3: Entfernte Dokumente ---
        ws_removed = wb['Entfernte Dokumente']
        ws_removed.cell(row=1, column=1).font = Font(bold=True)
        for col in range(1, 10):
            cell = ws_removed.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Gelbe Markierung für Stichproben-Zeilen
        for row_idx in range(5, ws_removed.max_row + 1):
            if ws_removed.cell(row=row_idx, column=1).value == '✓ PRÜFEN':
                for col in range(1, 10):
                    ws_removed.cell(row=row_idx, column=col).fill = sample_fill
        
        # --- Sheet 4: Stichprobe ---
        ws_sample = wb['Stichprobe']
        ws_sample.cell(row=1, column=1).font = Font(bold=True)
        for col in range(1, 8):
            cell = ws_sample.cell(row=5, column=col)
            cell.fill = header_fill
            cell.font = header_font
        # Hellgrüne Eingabefelder
        input_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for row_idx in range(6, ws_sample.max_row + 1):
            if ws_sample.cell(row=row_idx, column=1).value:
                ws_sample.cell(row=row_idx, column=6).fill = input_fill
                ws_sample.cell(row=row_idx, column=7).fill = input_fill
    
    print(f"   ✅ Excel erstellt: {len(cluster_info)} Cluster, {len(removed_docs)} entfernt")
    
    return str(excel_path)