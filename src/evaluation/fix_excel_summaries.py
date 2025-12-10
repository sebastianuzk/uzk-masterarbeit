"""
Korrigiere alle Excel-Dateien: Berechne Zusammenfassung neu aus Detaillierte Ergebnisse
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

def recalculate_summary(xlsx_path: Path):
    """
    Berechnet das Zusammenfassung-Blatt komplett neu aus Detaillierte Ergebnisse.
    Verwendet NUR Zeilen 2-117 (116 echte Fragen, ohne AVG-Zeilen).
    """
    print(f"\n{'='*80}")
    print(f"Verarbeite: {xlsx_path.name}")
    print('='*80)
    
    # Lade Detaillierte Ergebnisse
    try:
        df = pd.read_excel(xlsx_path, sheet_name='Detaillierte Ergebnisse')
    except Exception as e:
        print(f"  ❌ Fehler beim Laden: {e}")
        return False
    
    # NUR Zeilen 2-117 verwenden (Index 0-115, also erste 116 Zeilen)
    # Filtere AVG-Zeilen und andere Nicht-Frage-Zeilen aus
    df_original_len = len(df)
    df = df.head(116)  # Nur erste 116 Zeilen (Zeile 2-117 in Excel = Index 0-115)
    
    print(f"  📊 Original Zeilen: {df_original_len}")
    print(f"  📊 Verwendet Zeilen: {len(df)} (Zeile 2-117)")
    
    # Definiere Metriken
    metrics = ['faithfulness', 'context_recall', 'context_precision', 'gold_doc_rank']
    available_metrics = [m for m in metrics if m in df.columns]
    
    print(f"  📊 Verfügbare Metriken: {available_metrics}")
    
    # Berechne Gesamt-Durchschnitte
    print("\n  === GESAMT-DURCHSCHNITTE (NEU berechnet) ===")
    avg_values = {}
    for metric in available_metrics:
        avg_values[metric] = df[metric].mean()
        print(f"    {metric}: {avg_values[metric]:.6f}")
    
    # Berechne pro Kategorie
    cat_stats = None
    if 'category' in df.columns:
        cat_stats = df.groupby('category')[available_metrics].mean()
        print(f"\n  === KATEGORIEN: {len(cat_stats)} ===")
    
    # Berechne pro Schwierigkeit
    diff_stats = None
    if 'difficulty' in df.columns:
        diff_stats = df.groupby('difficulty')[available_metrics].mean()
        print(f"  === SCHWIERIGKEITEN: {len(diff_stats)} ===")
    
    # Lade Workbook und lösche altes Zusammenfassung-Blatt
    wb = load_workbook(xlsx_path)
    
    if 'Zusammenfassung' in wb.sheetnames:
        del wb['Zusammenfassung']
    
    ws = wb.create_sheet('Zusammenfassung')
    
    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    metric_font = Font(bold=True)
    
    row = 1
    
    # Titel
    ws.cell(row=row, column=1, value='📊 RAGAS-Evaluation Zusammenfassung')
    ws.cell(row=row, column=1).font = header_font
    row += 2
    
    # Durchschnittliche Scores
    ws.cell(row=row, column=1, value='Durchschnittliche Scores')
    ws.cell(row=row, column=1).font = subheader_font
    row += 1
    
    for metric in available_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=avg_values[metric])
        row += 1
    
    row += 2
    
    # Scores nach Kategorie
    if cat_stats is not None:
        ws.cell(row=row, column=1, value='Scores nach Kategorie')
        ws.cell(row=row, column=1).font = subheader_font
        row += 1
        
        # Header
        ws.cell(row=row, column=1, value='Kategorie')
        ws.cell(row=row, column=1).font = metric_font
        for i, metric in enumerate(available_metrics):
            col_name = metric.replace('_', ' ').title()
            ws.cell(row=row, column=i+2, value=col_name)
            ws.cell(row=row, column=i+2).font = metric_font
        row += 1
        
        # Daten
        for cat_name in sorted(cat_stats.index):
            ws.cell(row=row, column=1, value=cat_name)
            for i, metric in enumerate(available_metrics):
                value = cat_stats.loc[cat_name, metric]
                if pd.notna(value):
                    ws.cell(row=row, column=i+2, value=value)
            row += 1
        
        row += 2
    
    # Scores nach Schwierigkeit
    if diff_stats is not None:
        ws.cell(row=row, column=1, value='Scores nach Schwierigkeit')
        ws.cell(row=row, column=1).font = subheader_font
        row += 1
        
        # Header
        ws.cell(row=row, column=1, value='Schwierigkeit')
        ws.cell(row=row, column=1).font = metric_font
        for i, metric in enumerate(available_metrics):
            col_name = metric.replace('_', ' ').title()
            ws.cell(row=row, column=i+2, value=col_name)
            ws.cell(row=row, column=i+2).font = metric_font
        row += 1
        
        # Daten - in fester Reihenfolge (lowercase im DataFrame, UPPERCASE in Ausgabe)
        for diff_name in ['easy', 'medium', 'hard']:
            if diff_name in diff_stats.index:
                ws.cell(row=row, column=1, value=diff_name.upper())
                for i, metric in enumerate(available_metrics):
                    value = diff_stats.loc[diff_name, metric]
                    if pd.notna(value):
                        ws.cell(row=row, column=i+2, value=value)
                row += 1
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # Speichern
    wb.save(xlsx_path)
    print(f"\n  💾 Gespeichert: {xlsx_path.name}")
    return True


# Hauptprogramm
if __name__ == "__main__":
    backup_dir = Path('src/evaluation/data/Backup')
    
    # Liste der zu verarbeitenden Dateien
    target_files = [
        'naive_1250_300.xlsx',
        'naive_1500_300.xlsx',
        'naive_1750_200.xlsx',
        'naive_1750_300.xlsx',
        'semantic_1750_400_200_07.xlsx',
        'semantic_1750_400_300_07.xlsx',
        'semantic_2000_500_300_08.xlsx',
        'semantic_70.xlsx',
    ]
    
    print("=" * 80)
    print("KORREKTUR: Zusammenfassung-Blätter neu berechnen")
    print("=" * 80)
    
    success_count = 0
    for filename in target_files:
        xlsx_path = backup_dir / filename
        if xlsx_path.exists():
            if recalculate_summary(xlsx_path):
                success_count += 1
        else:
            print(f"\n⚠️  Datei nicht gefunden: {filename}")
    
    print("\n" + "=" * 80)
    print(f"ABGESCHLOSSEN: {success_count}/{len(target_files)} Dateien korrigiert")
    print("=" * 80)
