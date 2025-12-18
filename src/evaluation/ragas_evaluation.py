"""
RAGAS-Evaluation für WiSo-Chatbot

Evaluiert den Chatbot mit RAGAS-Framework:
- Verwendet Ollama (qwen3:8b) als LLM-Judge
- Lädt Testfragen aus Testset.CSV
- Generiert Antworten mit dem Chatbot
- Extrahiert RAG-Kontexte aus LangSmith
- Berechnet RAGAS-Metriken (Faithfulness, Context Recall)
"""

import sys
import warnings

# Unterdrücke DeprecationWarnings von RAGAS (LangchainEmbeddingsWrapper ist deprecated,
# aber notwendig für ResponseRelevancy mit lokalen HuggingFace-Embeddings in RAGAS 0.3.x)
warnings.filterwarnings("ignore", category=DeprecationWarning, module="ragas")

import random
import pandas as pd
import numpy as np
from pathlib import Path
from typing import List
import time
import requests
import gc

# Reproduzierbarkeit: Seeds werden aus config.settings geladen
import random
import numpy as np

# Projekt-Root
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from ragas import evaluate
from ragas.metrics import faithfulness, context_recall, context_precision, SemanticSimilarity, ResponseRelevancy, ContextEntityRecall
from ragas.dataset_schema import SingleTurnSample, EvaluationDataset
from ragas.embeddings import HuggingFaceEmbeddings, LangchainEmbeddingsWrapper
from ragas.run_config import RunConfig
from langchain_ollama import ChatOllama
from langchain_huggingface import HuggingFaceEmbeddings as LangchainHFEmbeddings
from langsmith import Client
from config.settings import (
    OLLAMA_MODEL,
    OLLAMA_BASE_URL,
    LANGSMITH_API_KEY,
    LANGSMITH_PROJECT,
    RAGAS_EVAL_MODEL,
    TEMPERATURE,
    CONTEXT_WINDOW,
    RANDOM_SEED,
    SENTENCE_TRANSFORMER_MODEL
)
from src.agent.react_agent import create_react_agent

# Setze Seeds für Reproduzierbarkeit
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)


def calculate_RR_at5(context_hint: str, retrieved_urls: list) -> float:
    """
    Berechnet RR@5 (Reciprocal Rank): An welcher Position erscheint die Referenz-URL?
    
    Args:
        context_hint: Die erwartete Referenz-URL aus dem Testset
        retrieved_urls: Liste der URLs der retrieved contexts
        
    Returns:
        float: 1/rank wenn gefunden (1.0 für Platz 1, 0.5 für Platz 2, etc.), 0.0 wenn nicht gefunden
    """
    if not context_hint or not retrieved_urls:
        return 0.0
    
    context_hint_str = str(context_hint)
    
    for i, url in enumerate(retrieved_urls):
        if url is None:
            continue
        url_str = str(url)
        
        # Exakte Übereinstimmung
        if context_hint_str == url_str:
            return 1.0 / (i + 1)
        
        # Für Web-URLs: Prüfe ob context_hint in der URL enthalten ist
        # z.B. https://wiso.uni-koeln.de/de/studium -> file://...html_cache/html/wiso.uni-koeln.de_de_studium...
        if context_hint_str.startswith('https://'):
            # Konvertiere https://wiso.uni-koeln.de/de/... zu wiso.uni-koeln.de_de_...
            url_part = context_hint_str.replace('https://', '').replace('/', '_')
            if url_part in url_str:
                return 1.0 / (i + 1)
        
        # Für file:// URLs: Direkter Vergleich
        if context_hint_str.startswith('file://') and url_str.startswith('file://'):
            if context_hint_str == url_str:
                return 1.0 / (i + 1)
    
    # Nicht gefunden
    return 0.0


def load_testset(csv_path: str = "data/Testset.CSV", limit: int = None) -> pd.DataFrame:
    """Lädt Testset.CSV"""
    full_path = Path(__file__).parent / csv_path
    df = pd.read_csv(full_path, sep=';', encoding='utf-8')
    
    if limit:
        df = df.head(limit)
    
    print(f"✅ {len(df)} Testfragen geladen")
    
    return df


def get_rag_context_from_langsmith(client: Client, trace_id: str) -> tuple:
    """
    Holt RAG-Kontext, URLs und Content-Types aus LangSmith für eine spezifische Trace-ID.
    
    Die Documents befinden sich im Retriever-Output unter dem Key 'output' (nicht 'documents'!).
    Jedes Document hat 'page_content' und 'metadata' (mit 'url' und 'content_type').
    
    Args:
        client: LangSmith Client
        trace_id: Die Trace-ID der Session
        
    Returns:
        Tuple (contexts, urls, content_types): Listen von RAG-Context-Chunks, URLs und Content-Types
    """
    try:
        # Hole alle Child-Runs für diese Trace
        child_runs = list(client.list_runs(
            project_name=LANGSMITH_PROJECT,
            trace_id=trace_id,
            is_root=False
        ))
        
        # Suche nach Retriever-Run
        contexts = []
        urls = []
        content_types = []
        for child in child_runs:
            if child.run_type == "retriever":
                if child.outputs and isinstance(child.outputs, dict):
                    # Documents sind unter 'output' Key (nicht 'documents')!
                    documents = child.outputs.get('output', [])
                    for doc in documents:
                        if isinstance(doc, dict) and 'page_content' in doc:
                            contexts.append(doc['page_content'])
                            # Metadata extrahieren (ohne Fallbacks)
                            metadata = doc.get('metadata', {})
                            # URL und Content-Type direkt aus metadata
                            urls.append(metadata.get('url'))
                            content_types.append(metadata.get('content_type'))
        
        if contexts:
            return contexts, urls, content_types  # Tuple von 3 Listen zurückgeben
        
        return [], [], []  # Leere Listen wenn keine Kontexte gefunden
    
    except Exception as e:
        print(f"      ⚠️ LangSmith-Fehler: {str(e)[:100]}")
        return [], [], []  # Leere Listen bei Fehler


def stop_ollama_model(model_name: str):
    """
    Stoppt ein Ollama-Modell via CLI-Befehl, um GPU-Speicher freizugeben.
    Der Ollama-Server läuft weiter und kann andere Modelle laden.
    
    Args:
        model_name: Name des zu stoppenden Modells (z.B. 'llama3.1:8b')
    """
    import subprocess
    
    try:
        result = subprocess.run(
            ["ollama", "stop", model_name],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0:
            print(f"   ✅ Modell {model_name} gestoppt (GPU-Speicher freigegeben)")
        else:
            # Fehlerausgabe prüfen
            if "not running" in result.stderr.lower():
                print(f"   ℹ️ Modell {model_name} war nicht geladen")
            else:
                print(f"   ⚠️ ollama stop: {result.stderr.strip()}")
                
    except subprocess.TimeoutExpired:
        print(f"   ⚠️ Timeout beim Stoppen von {model_name}")
    except FileNotFoundError:
        print(f"   ⚠️ 'ollama' Befehl nicht gefunden - ist Ollama installiert?")
    except Exception as e:
        print(f"   ⚠️ Fehler beim Stoppen von {model_name}: {e}")


def generate_chatbot_responses(df: pd.DataFrame, agent, langsmith_client: Client) -> tuple:
    """
    Generiert Chatbot-Antworten für alle Fragen und sammelt RAG-Kontexte.
    Speichert nach jeder Frage einen inkrementellen Checkpoint.
    Bei Timeout (3 Min) wird der Agent neu gestartet.
    
    Returns:
        Tuple (dataset, response_times, urls_list, content_types_list): 
        EvaluationDataset, Liste der Antwortzeiten, Liste der URL-Listen, Liste der Content-Type-Listen
    """
    import pickle
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
    
    TIMEOUT_SECONDS = 180  # 3 Minuten
    
    print("\n🤖 Generiere Chatbot-Antworten...")
    print(f"   ⏱️ Timeout pro Frage: {TIMEOUT_SECONDS}s")
    print("=" * 80)
    
    # Checkpoint-Pfad
    checkpoint_path = Path(__file__).parent / "data" / "responses_checkpoint.pkl"
    checkpoint_path.parent.mkdir(exist_ok=True)
    
    # Prüfe ob inkrementeller Checkpoint existiert (für Fortsetzung nach Abbruch)
    samples = []
    response_times = []
    urls_list = []
    content_types_list = []
    start_idx = 0
    
    if checkpoint_path.exists():
        try:
            with open(checkpoint_path, 'rb') as f:
                checkpoint_data = pickle.load(f)
            
            # Prüfe ob Checkpoint zum aktuellen Testset passt
            if isinstance(checkpoint_data, dict) and 'test_df' in checkpoint_data:
                saved_df = checkpoint_data['test_df']
                if len(saved_df) == len(df):
                    # Lade bisherige Samples
                    saved_dataset = checkpoint_data.get('dataset')
                    if saved_dataset and hasattr(saved_dataset, 'samples'):
                        samples = list(saved_dataset.samples)
                        response_times = checkpoint_data.get('response_times', [])
                        urls_list = checkpoint_data.get('urls_list', [])
                        content_types_list = checkpoint_data.get('content_types_list', [])
                        start_idx = len(samples)
                        
                        if start_idx > 0 and start_idx < len(df):
                            print(f"📂 Checkpoint geladen: {start_idx}/{len(df)} Fragen bereits beantwortet")
                            print(f"   → Setze fort ab Frage {start_idx + 1}")
                        elif start_idx >= len(df):
                            print(f"📂 Checkpoint vollständig: Alle {len(df)} Fragen beantwortet")
                            dataset = EvaluationDataset(samples=samples)
                            return dataset, response_times, urls_list, content_types_list
        except Exception as e:
            print(f"⚠️ Checkpoint-Ladefehler: {e} - Starte neu")
            samples, response_times, urls_list, content_types_list, start_idx = [], [], [], [], 0
    
    # Iteriere über verbleibende Fragen (ab start_idx)
    total_questions = len(df)
    for i in range(start_idx, total_questions):
        row = df.iloc[i]
        question = row['question']
        expected_answer = row['expected_answer']
        
        print(f"\n[{i + 1}/{total_questions}] {question[:70]}...")
        
        # Memory löschen für isolierte Evaluation
        agent.clear_memory()
        
        # Chatbot fragen - mit Session-ID für LangSmith-Tracking
        print(f"   💬 Chatbot fragen...")
        import uuid
        session_id = str(uuid.uuid4())
        
        # Zeit messen für Antwortgenerierung
        response_start = time.time()
        
        # Agent.chat() mit Timeout
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(agent.chat, question, session_id)
                answer = future.result(timeout=TIMEOUT_SECONDS)
        except FuturesTimeoutError:
            print(f"   ⏰ TIMEOUT nach {TIMEOUT_SECONDS}s - Agent wird neu gestartet...")
            agent = create_react_agent()
            print(f"   🔄 Agent neu gestartet - überspringe diese Frage")
            continue
        
        response_time = time.time() - response_start
        response_times.append(response_time)
        
        print(f"   ✅ Antwort: {answer[:80]}... ({response_time:.2f}s)")
        
        # Warten damit LangSmith Trace vollständig ist
        time.sleep(1)  # Reduziert von 3s auf 1s
        
        # RAG-Kontext aus LangSmith holen - nur den letzten Run abrufen
        print(f"   🔍 Hole RAG-Kontext aus LangSmith...")
        
        # Optimiert: Nur den letzten Run holen (statt alle)
        recent_runs = list(langsmith_client.list_runs(
            project_name=LANGSMITH_PROJECT,
            is_root=True,
            limit=1  # Nur den letzten Run
        ))
        
        contexts = []  # Leere Liste als Default
        urls = []  # Leere Liste als Default
        content_types = []  # Leere Liste als Default
        matching_run = None
        
        # Der letzte Run sollte unser Run sein
        if recent_runs:
            matching_run = recent_runs[0]
        
        if matching_run:
            trace_id = matching_run.trace_id
            contexts, urls, content_types = get_rag_context_from_langsmith(langsmith_client, trace_id)
            print(f"   ✅ Run gefunden mit Session-ID: {session_id[:8]}...")
        else:
            print(f"   ⚠️ Kein Run mit Session-ID {session_id[:8]}... gefunden")
        
        urls_list.append(urls)
        content_types_list.append(content_types)
        
        total_chars = sum(len(c) for c in contexts)
        print(f"   📄 Kontext: {len(contexts)} chunks, {total_chars} Zeichen")
        print(f"   🔗 URLs: {len(urls)} Quellen")
        print(f"   📁 Content-Types: {set(content_types)}")
        
        # RAGAS-Sample erstellen
        sample = SingleTurnSample(
            user_input=question,
            response=answer,
            retrieved_contexts=contexts,  # Jetzt bereits eine Liste von Chunks
            reference=expected_answer
        )
        samples.append(sample)
        
        # 💾 INKREMENTELLER CHECKPOINT nach jeder Frage
        dataset = EvaluationDataset(samples=samples)
        checkpoint_data = {
            'dataset': dataset,
            'test_df': df,
            'response_times': response_times,
            'urls_list': urls_list,
            'content_types_list': content_types_list
        }
        with open(checkpoint_path, 'wb') as f:
            pickle.dump(checkpoint_data, f)
        print(f"   💾 Checkpoint: {len(samples)}/{len(df)} Fragen gespeichert")
    
    print("\n" + "=" * 80)
    print(f"✅ {len(samples)} Antworten generiert\n")
    print(f"   ⏱️ Durchschn. Antwortzeit: {sum(response_times)/len(response_times):.2f}s")
    print(f"   ⏱️ Gesamt Antwortzeit: {sum(response_times):.2f}s\n")
    
    dataset = EvaluationDataset(samples=samples)
    return dataset, response_times, urls_list, content_types_list


# ============================================================================
# KONFIGURATION
# ============================================================================
# Limit für Testfragen (None = alle, z.B. 5 für Test)
TEST_LIMIT = 2  # None = alle Fragen evaluieren


def run_ragas_evaluation(dataset: EvaluationDataset) -> tuple:
    """
    Führt RAGAS-Evaluation durch.
    Verwendet 3 Standard-RAGAS-Metriken: faithfulness, context_recall, context_precision.
    (answer_relevancy auskommentiert - benötigt qwen3-embedding:8b)
    
    Returns:
        Tuple (results_df, evaluation_time): DataFrame mit Ergebnissen und Evaluationszeit in Sekunden
    """
    print("🚀 Starte RAGAS-Evaluation...")
    print("=" * 80)
    
    # Separates LLM für RAGAS-Evaluation (gleiches Setup wie Chatbot, nur anderes Modell)
    llm = ChatOllama(
        model=RAGAS_EVAL_MODEL,  # Separates Modell für Evaluation
        base_url=OLLAMA_BASE_URL,
        temperature=TEMPERATURE,  # Gleiche Parameter wie Chatbot
        seed=RANDOM_SEED,
        num_ctx=CONTEXT_WINDOW
    )
    print(f"   RAGAS-LLM: {RAGAS_EVAL_MODEL} @ {OLLAMA_BASE_URL} (ctx={CONTEXT_WINDOW}, temp={TEMPERATURE}, seed={RANDOM_SEED})")
    print(f"   (Chatbot verwendet: {OLLAMA_MODEL})")
    
    # Embeddings für RAGAS-Metriken
    # - SemanticSimilarity: Verwendet RAGAS HuggingFaceEmbeddings (embed_text Interface)
    # - ResponseRelevancy: Benötigt LangChain-Interface (embed_query), daher LangchainEmbeddingsWrapper
    ragas_embeddings = HuggingFaceEmbeddings(model=SENTENCE_TRANSFORMER_MODEL)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        langchain_embeddings = LangchainEmbeddingsWrapper(LangchainHFEmbeddings(model_name=SENTENCE_TRANSFORMER_MODEL))
    print(f"   Embeddings: {SENTENCE_TRANSFORMER_MODEL} (für semantic_similarity, answer_relevancy)")
    
    # SemanticSimilarity mit RAGAS HuggingFaceEmbeddings
    semantic_similarity = SemanticSimilarity(embeddings=ragas_embeddings)
    
    # ResponseRelevancy benötigt LangChain-Interface (embed_query/embed_documents)
    response_relevancy = ResponseRelevancy(embeddings=langchain_embeddings)
    
    # ContextEntityRecall Metrik (misst Entitäten-Recall zwischen Referenz und Kontext)
    context_entity_recall = ContextEntityRecall()
    
    # Standard RAGAS-Metriken
    metrics = [
        faithfulness,           # Ist Antwort treu zum Kontext?
        context_recall,         # Wurden alle relevanten Infos abgerufen?
        context_precision,      # Sind relevante Chunks höher gerankt?
        semantic_similarity,    # Semantische Ähnlichkeit zwischen Antwort und Referenz
        context_entity_recall,  # Entitäten-Recall zwischen Referenz und Kontext
        response_relevancy      # Relevanz der Antwort zur gestellten Frage
    ]
    print(f"   Metriken: {[m.name for m in metrics]}")
    print(f"\n   ⏳ Evaluiere {len(dataset.samples)} Samples...")
    print(f"   💡 Dies kann mehrere Minuten dauern (ca. 1-2 Min pro Sample)\n")
    
    # RunConfig für parallele Requests an Ollama
    # seed=RANDOM_SEED für Reproduzierbarkeit (RAGAS verwendet numpy RNG intern)
    # timeout=240 für längere Metrik-Berechnungen (response_relevancy)
    run_config = RunConfig(
        max_workers=4,
        seed=RANDOM_SEED,
        timeout=240
    )
    
    # Zeit messen für Evaluation
    eval_start = time.time()
    
    # Evaluation durchführen
    results = evaluate(
        dataset=dataset,
        metrics=metrics,
        llm=llm,
        run_config=run_config,
        raise_exceptions=False  # Weiter bei Fehlern
    )
    
    evaluation_time = time.time() - eval_start
    
    print(f"\n   ✅ Evaluation abgeschlossen in {evaluation_time:.2f}s")
    print(f"   ⏱️ Durchschn. pro Sample: {evaluation_time/len(dataset.samples):.2f}s\n")
    
    return results.to_pandas(), evaluation_time


def display_and_save_results(results_df: pd.DataFrame, test_df: pd.DataFrame, 
                              response_times: List[float] = None, urls_list: List[List[str]] = None,
                              content_types_list: List[List[str]] = None, evaluation_time: float = None):
    """
    Zeigt Ergebnisse an und speichert sie.
    
    Args:
        results_df: DataFrame mit RAGAS-Ergebnissen
        test_df: DataFrame mit Testdaten
        response_times: Liste der Antwortzeiten pro Frage (optional)
        urls_list: Liste von URL-Listen pro Frage (optional)
        content_types_list: Liste von Content-Type-Listen pro Frage (optional)
        evaluation_time: Gesamtzeit für RAGAS-Evaluation in Sekunden (optional)
    """
    from datetime import datetime
    
    # IDs, Kategorien und Schwierigkeiten hinzufügen
    results_df['id'] = test_df['id'].values[:len(results_df)]
    results_df['category'] = test_df['category'].values[:len(results_df)]
    results_df['difficulty'] = test_df['difficulty'].values[:len(results_df)]
    
    # Antwortzeiten hinzufügen (falls vorhanden)
    if response_times:
        results_df['response_time_seconds'] = response_times[:len(results_df)]
    else:
        results_df['response_time_seconds'] = None
    
    # URLs hinzufügen (falls vorhanden)
    if urls_list:
        results_df['retrieved_urls'] = [str(urls) for urls in urls_list[:len(results_df)]]
    else:
        results_df['retrieved_urls'] = None
    
    # Content-Types hinzufügen (falls vorhanden)
    if content_types_list:
        results_df['retrieved_content_types'] = [str(types) for types in content_types_list[:len(results_df)]]
    else:
        results_df['retrieved_content_types'] = None
    
    # RR_at5 berechnen (Reciprocal Rank: Position der Referenz-URL in retrieved contexts)
    if urls_list:
        rr_at5_values = []
        hit_at5_values = []
        for i in range(len(results_df)):
            context_hint = test_df['context_hint'].iloc[i] if i < len(test_df) else None
            retrieved_urls = urls_list[i] if i < len(urls_list) else []
            rr = calculate_RR_at5(context_hint, retrieved_urls)
            rr_at5_values.append(rr)
            hit_at5_values.append(1.0 if rr > 0 else 0.0)
        results_df['RR_at5'] = rr_at5_values
        results_df['hit_at5'] = hit_at5_values
    else:
        results_df['RR_at5'] = None
        results_df['hit_at5'] = None
    
    # Latency als Kopie von response_time_seconds
    if 'response_time_seconds' in results_df.columns:
        results_df['latency'] = results_df['response_time_seconds']
    else:
        results_df['latency'] = None
    
    # ============================================================================
    # SOFORT SPEICHERN - Rohdaten CSV (bevor irgendwas schiefgehen kann)
    # ============================================================================
    output_path_raw = Path(__file__).parent / "data" / "ragas_results_raw.csv"
    try:
        results_df.to_csv(output_path_raw, index=False, encoding='utf-8-sig')
        print(f"\n💾 ROHDATEN GESPEICHERT: {output_path_raw}")
    except Exception as e:
        print(f"\n⚠️ Fehler beim Speichern der Rohdaten: {e}")
    
    print("\n" + "=" * 80)
    print("📊 RAGAS-EVALUATION ERGEBNISSE")
    print("=" * 80)
    
    # Gesamtscores
    print("\n📈 Durchschnittliche Scores:")
    print("-" * 80)
    for metric in ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency']:
        if metric in results_df.columns and results_df[metric].notna().any():
            avg = results_df[metric].mean()
            # Anzeigename für Zusammenfassung
            display_name = 'MRR@5' if metric == 'RR_at5' else ('Hit@5' if metric == 'hit_at5' else metric)
            print(f"   {display_name:20s}: {avg:.3f}")
    
    # Nach Kategorie (NaN ausfiltern)
    print("\n📁 Scores nach Kategorie:")
    print("-" * 80)
    display_categories = [c for c in results_df['category'].unique() if pd.notna(c)]
    for category in sorted(display_categories):
        cat_df = results_df[results_df['category'] == category]
        print(f"\n   {category}:")
        for metric in ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency']:
            if metric in cat_df.columns and cat_df[metric].notna().any():
                avg = cat_df[metric].mean()
                display_name = 'MRR@5' if metric == 'RR_at5' else ('Hit@5' if metric == 'hit_at5' else metric)
                print(f"      {display_name:20s}: {avg:.3f}")
    
    # Nach Schwierigkeit
    print("\n⚡ Scores nach Schwierigkeit:")
    print("-" * 80)
    for difficulty in ['easy', 'medium', 'hard']:
        diff_df = results_df[results_df['difficulty'] == difficulty]
        if len(diff_df) > 0:
            print(f"\n   {difficulty.upper()}:")
            for metric in ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency']:
                if metric in diff_df.columns and diff_df[metric].notna().any():
                    avg = diff_df[metric].mean()
                    display_name = 'MRR@5' if metric == 'RR_at5' else ('Hit@5' if metric == 'hit_at5' else metric)
                    print(f"      {display_name:20s}: {avg:.3f}")
    
    # Speichern in CSV (alle Spalten)
    output_path_csv = Path(__file__).parent / "data" / "ragas_results.csv"
    
    # Berechne Anzahl der Context-Chunks
    results_df['context_count'] = results_df['retrieved_contexts'].apply(lambda x: len(x) if isinstance(x, list) else 0)
    
    # Entferne Zeilenumbrüche aus Textfeldern für saubere CSV
    text_columns = ['user_input', 'response', 'reference']
    for col in text_columns:
        if col in results_df.columns:
            results_df[col] = results_df[col].apply(lambda x: x.replace('\n', ' ').replace('\r', ' ') if isinstance(x, str) else x)
    
    # Konvertiere retrieved_contexts zu String ohne Zeilenumbrüche
    results_df['retrieved_contexts'] = results_df['retrieved_contexts'].apply(
        lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if isinstance(x, list) else str(x)
    )
    
    # Konvertiere retrieved_urls zu String ohne Zeilenumbrüche (falls vorhanden)
    if 'retrieved_urls' in results_df.columns:
        results_df['retrieved_urls'] = results_df['retrieved_urls'].apply(
            lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if x else ''
        )
    
    # Konvertiere retrieved_content_types zu String ohne Zeilenumbrüche (falls vorhanden)
    if 'retrieved_content_types' in results_df.columns:
        results_df['retrieved_content_types'] = results_df['retrieved_content_types'].apply(
            lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if x else ''
        )
    
    # CSV mit allen wichtigen Spalten (erweitert um response_time, urls, content_types und alle Metriken)
    csv_columns = ['id', 'category', 'difficulty', 'user_input', 'response', 
                   'reference', 'retrieved_contexts', 'retrieved_urls', 'retrieved_content_types',
                   'faithfulness', 'context_recall', 'context_precision', 'semantic_similarity',
                   'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency',
                   'context_count', 'response_time_seconds']
    
    # Nur vorhandene Spalten verwenden
    csv_columns = [col for col in csv_columns if col in results_df.columns]
    csv_df = results_df[csv_columns].copy()
    
    # ============================================================================
    # METADATEN-ZEILEN: Modelle, Zeitstempel, Dauern
    # ============================================================================
    metadata_rows = []
    
    # Metadaten-Zeile 1: Allgemeine Infos
    meta1 = {col: '' for col in csv_columns}
    meta1['id'] = 'META'
    meta1['category'] = 'Evaluation Metadaten'
    meta1['difficulty'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta1['user_input'] = f'Chatbot: {OLLAMA_MODEL} (ctx=dynamisch, temp={TEMPERATURE}, seed={RANDOM_SEED})'
    meta1['response'] = f'RAGAS-LLM: {RAGAS_EVAL_MODEL} (ctx={CONTEXT_WINDOW}, temp={TEMPERATURE}, seed={RANDOM_SEED})'
    meta1['reference'] = f'Testset: {len(results_df)} Fragen'
    if evaluation_time:
        meta1['retrieved_contexts'] = f'Eval-Zeit: {evaluation_time:.2f}s'
    if response_times:
        meta1['retrieved_urls'] = f'Antwort-Zeit gesamt: {sum(response_times):.2f}s'
    metadata_rows.append(meta1)
    
    # ============================================================================
    # DURCHSCHNITTE: Gesamt, pro Kategorie, pro Schwierigkeit, kombiniert
    # ============================================================================
    metric_cols = ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency']
    
    # Gesamtdurchschnitt
    avg_row = {col: '' for col in csv_columns}
    avg_row['id'] = 'AVG'
    avg_row['category'] = 'GESAMT'
    avg_row['difficulty'] = f'n={len(results_df)}'
    for metric in metric_cols:
        if metric in results_df.columns:
            avg_row[metric] = results_df[metric].mean()
    if 'response_time_seconds' in results_df.columns and results_df['response_time_seconds'].notna().any():
        avg_row['response_time_seconds'] = results_df['response_time_seconds'].mean()
    metadata_rows.append(avg_row)
    
    # Durchschnitte pro Kategorie (NaN-Werte ausfiltern)
    categories = [c for c in results_df['category'].unique() if pd.notna(c)]
    for category in sorted(categories):
        cat_df = results_df[results_df['category'] == category]
        cat_row = {col: '' for col in csv_columns}
        cat_row['id'] = 'AVG'
        cat_row['category'] = category
        cat_row['difficulty'] = f'n={len(cat_df)}'
        for metric in metric_cols:
            if metric in cat_df.columns:
                cat_row[metric] = cat_df[metric].mean()
        if 'response_time_seconds' in cat_df.columns and cat_df['response_time_seconds'].notna().any():
            cat_row['response_time_seconds'] = cat_df['response_time_seconds'].mean()
        metadata_rows.append(cat_row)
    
    # Durchschnitte pro Schwierigkeit
    for difficulty in ['easy', 'medium', 'hard']:
        diff_df = results_df[results_df['difficulty'] == difficulty]
        if len(diff_df) > 0:
            diff_row = {col: '' for col in csv_columns}
            diff_row['id'] = 'AVG'
            diff_row['category'] = f'Schwierigkeit: {difficulty.upper()}'
            diff_row['difficulty'] = f'n={len(diff_df)}'
            for metric in metric_cols:
                if metric in diff_df.columns:
                    diff_row[metric] = diff_df[metric].mean()
            if 'response_time_seconds' in diff_df.columns and diff_df['response_time_seconds'].notna().any():
                diff_row['response_time_seconds'] = diff_df['response_time_seconds'].mean()
            metadata_rows.append(diff_row)
    
    # Durchschnitte pro Kategorie + Schwierigkeit (kombiniert)
    for category in sorted(categories):  # Verwende bereits gefilterte categories-Liste
        for difficulty in ['easy', 'medium', 'hard']:
            combo_df = results_df[(results_df['category'] == category) & (results_df['difficulty'] == difficulty)]
            if len(combo_df) > 0:
                combo_row = {col: '' for col in csv_columns}
                combo_row['id'] = 'AVG'
                combo_row['category'] = f'{category} / {difficulty.upper()}'
                combo_row['difficulty'] = f'n={len(combo_df)}'
                for metric in metric_cols:
                    if metric in combo_df.columns:
                        combo_row[metric] = combo_df[metric].mean()
                if 'response_time_seconds' in combo_df.columns and combo_df['response_time_seconds'].notna().any():
                    combo_row['response_time_seconds'] = combo_df['response_time_seconds'].mean()
                metadata_rows.append(combo_row)
    
    # Metadaten-DataFrame erstellen und anhängen
    meta_df = pd.DataFrame(metadata_rows)
    csv_df = pd.concat([csv_df, meta_df], ignore_index=True)
    
    # Speichere mit UTF-8-BOM für korrekte Umlaut-Darstellung
    csv_df.to_csv(output_path_csv, index=False, encoding='utf-8-sig', sep=',', quoting=1)
    
    # Excel mit Formatierung erstellen
    output_path_excel = Path(__file__).parent / "data" / "ragas_results.xlsx"
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Sheet 1: Detaillierte Ergebnisse
        ws_details = wb.active
        ws_details.title = "Detaillierte Ergebnisse"
        
        # Header-Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Daten schreiben - NUR results_df (ohne AVG-Zeilen), csv_df enthält Metadaten
        # Verwende die gleichen Spalten wie csv_df, aber aus results_df
        excel_details_df = results_df[csv_columns].copy()
        for r_idx, row in enumerate(dataframe_to_rows(excel_details_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatieren
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Metriken-Spalten (faithfulness, context_recall, context_precision) farbig
                    if c_idx in [8, 9, 10]:  # Metrik-Spalten
                        if isinstance(value, (int, float)):
                            if value >= 0.8:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value >= 0.6:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.number_format = '0.000'
                    
                    # Text-Wrap für lange Texte
                    if c_idx in [4, 5, 6, 7]:  # user_input, response, reference, contexts
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Spaltenbreiten anpassen
        ws_details.column_dimensions['A'].width = 8   # id
        ws_details.column_dimensions['B'].width = 20  # category
        ws_details.column_dimensions['C'].width = 12  # difficulty
        ws_details.column_dimensions['D'].width = 50  # user_input
        ws_details.column_dimensions['E'].width = 60  # response
        ws_details.column_dimensions['F'].width = 50  # reference
        ws_details.column_dimensions['G'].width = 40  # contexts
        ws_details.column_dimensions['H'].width = 30  # retrieved_urls
        ws_details.column_dimensions['I'].width = 20  # retrieved_content_types
        ws_details.column_dimensions['J'].width = 13  # faithfulness
        ws_details.column_dimensions['K'].width = 13  # context_recall
        ws_details.column_dimensions['L'].width = 15  # context_precision
        ws_details.column_dimensions['M'].width = 17  # semantic_similarity
        ws_details.column_dimensions['N'].width = 20  # context_entity_recall
        ws_details.column_dimensions['O'].width = 18  # answer_relevancy
        ws_details.column_dimensions['P'].width = 10  # RR_at5
        ws_details.column_dimensions['Q'].width = 10  # hit_at5
        ws_details.column_dimensions['R'].width = 12  # context_count
        ws_details.column_dimensions['S'].width = 15  # response_time_seconds
        
        # Sheet 2: Zusammenfassung
        ws_summary = wb.create_sheet("Zusammenfassung")
        
        # Titel
        ws_summary['A1'] = "📊 RAGAS-Evaluation Zusammenfassung"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')
        
        # Durchschnittliche Scores
        row = 3
        ws_summary[f'A{row}'] = "Durchschnittliche Scores"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for metric in ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency']:
            if metric in results_df.columns and results_df[metric].notna().any():
                avg = results_df[metric].mean()
                # Anzeigename für Excel
                display_name = 'MRR@5' if metric == 'RR_at5' else ('Hit@5' if metric == 'hit_at5' else metric)
                ws_summary[f'A{row}'] = display_name
                ws_summary[f'B{row}'] = avg
                ws_summary[f'B{row}'].number_format = '0.000'
                
                # Farbe basierend auf Score (nicht für latency)
                if metric != 'latency':
                    if avg >= 0.8:
                        ws_summary[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif avg >= 0.6:
                        ws_summary[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1
        
        # Nach Kategorie
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Kategorie"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header für Kategorie-Tabelle
        ws_summary[f'A{row}'] = "Kategorie"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        ws_summary[f'E{row}'] = "Semantic Similarity"
        ws_summary[f'F{row}'] = "Context Entity Recall"
        ws_summary[f'G{row}'] = "Answer Relevancy"
        ws_summary[f'H{row}'] = "MRR@5"
        ws_summary[f'I{row}'] = "Hit@5"
        ws_summary[f'J{row}'] = "Latency"
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        # NaN-Kategorien ausfiltern für Excel
        excel_categories = [c for c in results_df['category'].unique() if pd.notna(c)]
        for category in sorted(excel_categories):
            cat_df = results_df[results_df['category'] == category]
            ws_summary[f'A{row}'] = category
            
            for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency'], 1):
                if metric in cat_df.columns:
                    avg = cat_df[metric].mean()
                    col_letter = chr(65 + idx)  # B, C, D, E, F, G, H, I, J
                    ws_summary[f'{col_letter}{row}'] = avg
                    ws_summary[f'{col_letter}{row}'].number_format = '0.000'
            
            row += 1
        
        # Nach Schwierigkeit
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Schwierigkeit"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header
        ws_summary[f'A{row}'] = "Schwierigkeit"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        ws_summary[f'E{row}'] = "Semantic Similarity"
        ws_summary[f'F{row}'] = "Context Entity Recall"
        ws_summary[f'G{row}'] = "Answer Relevancy"
        ws_summary[f'H{row}'] = "MRR@5"
        ws_summary[f'I{row}'] = "Hit@5"
        ws_summary[f'J{row}'] = "Latency"
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        for difficulty in ['easy', 'medium', 'hard']:
            diff_df = results_df[results_df['difficulty'] == difficulty]
            if len(diff_df) > 0:
                ws_summary[f'A{row}'] = difficulty.upper()
                
                for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'RR_at5', 'hit_at5', 'latency'], 1):
                    if metric in diff_df.columns:
                        avg = diff_df[metric].mean()
                        col_letter = chr(65 + idx)  # B, C, D, E, F, G, H, I, J
                        ws_summary[f'{col_letter}{row}'] = avg
                        ws_summary[f'{col_letter}{row}'].number_format = '0.000'
                
                row += 1
        
        # Spaltenbreiten für Zusammenfassung
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 18
        ws_summary.column_dimensions['E'].width = 20
        ws_summary.column_dimensions['F'].width = 22
        ws_summary.column_dimensions['G'].width = 18
        ws_summary.column_dimensions['H'].width = 10
        ws_summary.column_dimensions['I'].width = 10
        ws_summary.column_dimensions['J'].width = 10
        
        # Speichern
        wb.save(output_path_excel)
        
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV:   {output_path_csv}")
        print(f"   Excel: {output_path_excel}")
        print("=" * 80 + "\n")
        
    except ImportError:
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV: {output_path_csv}")
        print(f"   ⚠️ Excel-Export nicht verfügbar (openpyxl nicht installiert)")
        print("=" * 80 + "\n")


def main():
    """Hauptfunktion"""
    
    print("\n" + "=" * 80)
    print("🎯 RAGAS-EVALUATION - WiSo-Chatbot")
    print("=" * 80 + "\n")
    
    # Checkpoint-Pfad
    checkpoint_path = Path(__file__).parent / "data" / "responses_checkpoint.pkl"
    
    # Variablen für Timing, URLs und Content-Types initialisieren
    response_times = None
    urls_list = None
    content_types_list = None
    dataset = None
    checkpoint_complete = False
    
    try:
        # 1. LangSmith Client (immer initialisieren)
        print("🔗 Initialisiere LangSmith...")
        langsmith_client = Client(api_key=LANGSMITH_API_KEY)
        print(f"   ✅ Projekt: {LANGSMITH_PROJECT}\n")
        
        # 2. Testset laden (mit optionalem Limit)
        print("📂 Lade Testset...")
        test_df = load_testset(limit=TEST_LIMIT)
        print()
        
        # 3. Prüfe ob Checkpoint existiert und vollständig ist
        if checkpoint_path.exists():
            print("📂 Prüfe Checkpoint...")
            import pickle
            with open(checkpoint_path, 'rb') as f:
                checkpoint_data = pickle.load(f)
            
            # Checkpoint kann EvaluationDataset oder dict sein
            if isinstance(checkpoint_data, dict):
                saved_dataset = checkpoint_data.get('dataset')
                saved_df = checkpoint_data.get('test_df')
                response_times = checkpoint_data.get('response_times', None)
                urls_list = checkpoint_data.get('urls_list', None)
                content_types_list = checkpoint_data.get('content_types_list', None)
                
                # Prüfe ob Checkpoint vollständig ist
                if saved_dataset and hasattr(saved_dataset, 'samples'):
                    num_saved = len(saved_dataset.samples)
                    num_expected = len(test_df)
                    
                    if num_saved >= num_expected:
                        # Vollständig → direkt zur Evaluation
                        print(f"   ✅ Checkpoint vollständig: {num_saved}/{num_expected} Antworten")
                        dataset = saved_dataset
                        checkpoint_complete = True
                    else:
                        # Unvollständig → Fortsetzung nötig
                        print(f"   ⏳ Checkpoint unvollständig: {num_saved}/{num_expected} Antworten")
                        print(f"   → generate_chatbot_responses() wird fortsetzen\n")
            else:
                # Alter Checkpoint-Format (nur Dataset)
                dataset = checkpoint_data
                if hasattr(dataset, 'samples') and len(dataset.samples) >= len(test_df):
                    checkpoint_complete = True
                    print(f"   ✅ Alter Checkpoint vollständig: {len(dataset.samples)} Antworten")
        
        # 4. Falls Checkpoint unvollständig oder nicht vorhanden → Antworten generieren
        if not checkpoint_complete:
            # Chatbot initialisieren
            print("🤖 Initialisiere Chatbot...")
            agent = create_react_agent()
            print()
            
            # Antworten generieren (setzt bei Checkpoint fort)
            dataset, response_times, urls_list, content_types_list = generate_chatbot_responses(test_df, agent, langsmith_client)
            
            # ====================================================================
            # CHATBOT-MODELL ENTLADEN (GPU-Speicher freigeben vor RAGAS)
            # ====================================================================
            print("\n🧹 Räume GPU-Speicher auf (Chatbot-Modell entladen)...")
            del agent  # Python-Referenz löschen
            gc.collect()  # Garbage Collection
            stop_ollama_model(OLLAMA_MODEL)  # Modell via CLI stoppen
            print()
        
        # 5. RAGAS-Evaluation (immer ausführen, jetzt mit Timing)
        results_df, evaluation_time = run_ragas_evaluation(dataset)
        
        # 6. Ergebnisse anzeigen und speichern (mit allen neuen Daten)
        display_and_save_results(results_df, test_df, response_times, urls_list, content_types_list, evaluation_time)
        
        print("✅ Evaluation erfolgreich abgeschlossen!")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Evaluation abgebrochen!\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Fehler: {e}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
