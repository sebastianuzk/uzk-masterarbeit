"""
RAGAS-Evaluation für WiSo-Chatbot

Evaluiert den Chatbot mit RAGAS-Framework:
- Unterstützt zwei LLM-Modi für den Judge: LOKAL (Ollama) oder CLOUD (OpenAI)
- Embeddings werden IMMER lokal mit Ollama (embeddinggemma) berechnet
- Konfiguration über RUN_EVALUATION_LOCAL in .env (true/false)
- Lädt Testfragen aus Testset.CSV
- Generiert Antworten mit dem Chatbot
- Extrahiert RAG-Kontexte aus LangSmith
- Berechnet RAGAS-Metriken (Faithfulness, Context Recall, etc.)

Modi:
- RUN_EVALUATION_LOCAL=true  → Ollama LLM (RAGAS_EVAL_MODEL) + Ollama Embeddings
- RUN_EVALUATION_LOCAL=false → OpenAI LLM (OPENAI_EVAL_MODEL) + Ollama Embeddings
"""

import sys
import warnings

# Unterdrücke DeprecationWarnings von RAGAS (LangchainEmbeddingsWrapper ist deprecated,
# aber notwendig für ResponseRelevancy mit lokalen HuggingFace-Embeddings in RAGAS 0.3.x)
warnings.filterwarnings("ignore", category=DeprecationWarning, module="ragas")

import random
import pandas as pd
import numpy as np
from pathlib import Path
from typing import List
import time
import requests
import gc

# BERT-Score für Token-Level semantische Ähnlichkeit
try:
    from bert_score import score as bert_score_fn
    BERT_SCORE_AVAILABLE = True
except ImportError:
    BERT_SCORE_AVAILABLE = False
    print("⚠️ bert-score nicht installiert. Installiere mit: pip install bert-score")

# Reproduzierbarkeit: Seeds werden aus config.settings geladen
import random
import numpy as np

# Projekt-Root
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from ragas import evaluate
from ragas.metrics import faithfulness, context_recall, context_precision, SemanticSimilarity, ResponseRelevancy, ContextEntityRecall
from ragas.dataset_schema import SingleTurnSample, EvaluationDataset
from ragas.embeddings import HuggingFaceEmbeddings, LangchainEmbeddingsWrapper
from ragas.run_config import RunConfig
from langchain_ollama import ChatOllama, OllamaEmbeddings
from langchain_huggingface import HuggingFaceEmbeddings as LangchainHFEmbeddings
from langsmith import Client
from config.settings import (
    OLLAMA_MODEL,
    OLLAMA_BASE_URL,
    LANGSMITH_API_KEY,
    LANGSMITH_PROJECT,
    RAGAS_EVAL_MODEL,
    TEMPERATURE,
    CONTEXT_WINDOW,
    RANDOM_SEED,
    SENTENCE_TRANSFORMER_MODEL,
    OPENAI_API_KEY,
    OPENAI_EVAL_MODEL,
    RUN_EVALUATION_LOCAL
)
from src.agent.react_agent import create_react_agent

# Setze Seeds für Reproduzierbarkeit
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

# Timestamps für Batch-Evaluation (Array wird in main() iteriert)
from datetime import datetime
#EVAL_TIMESTAMPS = [datetime.now().strftime("%Y%m%d_%H%M%S")]  # Für neue Evaluation


EVAL_TIMESTAMPS = [
    "20260128_183659",

    # Weitere Timestamps hier hinzufügen...
]

# IDs der Testfragen die evaluiert werden sollen (None = alle Fragen)
# Beispiel: EVAL_IDS = [1, 5, 10, 15]  → Nur diese IDs evaluieren
# Beispiel: EVAL_IDS = None  → Alle Fragen aus dem Testset evaluieren
EVAL_IDS = [1, 8, 22, 23, 29, 32, 40, 46, 47, 49, 52
]

  # None = alle Fragen, oder Liste von IDs z.B. [1, 2, 3]


# Aktueller Timestamp (wird in main() pro Iteration gesetzt)
EVAL_TIMESTAMP = None


def calculate_RR_at5(context_hint: str, retrieved_urls: list) -> float:
    """
    Berechnet RR@5 (Reciprocal Rank): An welcher Position erscheint die Referenz-URL?
    
    Args:
        context_hint: Die erwartete Referenz-URL aus dem Testset
        retrieved_urls: Liste der URLs der retrieved contexts
        
    Returns:
        float: 1/rank wenn gefunden (1.0 für Platz 1, 0.5 für Platz 2, etc.), 0.0 wenn nicht gefunden
    """
    if not context_hint or not retrieved_urls:
        return 0.0
    
    context_hint_str = str(context_hint)
    
    for i, url in enumerate(retrieved_urls):
        if url is None:
            continue
        url_str = str(url)
        
        # Exakte Übereinstimmung
        if context_hint_str == url_str:
            return 1.0 / (i + 1)
        
        # Für Web-URLs: Prüfe ob context_hint in der URL enthalten ist
        # z.B. https://wiso.uni-koeln.de/de/studium -> file://...html_cache/html/wiso.uni-koeln.de_de_studium...
        if context_hint_str.startswith('https://'):
            # Konvertiere https://wiso.uni-koeln.de/de/... zu wiso.uni-koeln.de_de_...
            url_part = context_hint_str.replace('https://', '').replace('/', '_')
            if url_part in url_str:
                return 1.0 / (i + 1)
        
        # Für file:// URLs: Direkter Vergleich
        if context_hint_str.startswith('file://') and url_str.startswith('file://'):
            if context_hint_str == url_str:
                return 1.0 / (i + 1)
    
    # Nicht gefunden
    return 0.0


def load_testset(csv_path: str = "data/Testset.CSV", limit: int = None, ids = None) -> pd.DataFrame:
    """
    Lädt Testset.CSV mit optionaler Filterung.
    
    Args:
        csv_path: Pfad zur CSV-Datei relativ zum Evaluations-Verzeichnis
        limit: Maximale Anzahl an Fragen (optional)
        ids: Einzelne ID (int) oder Liste von Frage-IDs (List[int]) die geladen werden sollen (optional)
        
    Returns:
        DataFrame mit den Testfragen
    """
    full_path = Path(__file__).parent / csv_path
    df = pd.read_csv(full_path, sep=';', encoding='utf-8')
    
    # Nach IDs filtern (falls angegeben)
    if ids is not None:
        # Konvertiere einzelne ID zu Liste
        if isinstance(ids, int):
            ids = [ids]
        df = df[df['id'].isin(ids)]
        print(f"🎯 Gefiltert auf IDs: {ids}")
    
    if limit:
        df = df.head(limit)
    
    print(f"✅ {len(df)} Testfragen geladen")
    
    return df


def get_rag_context_from_langsmith(client: Client, trace_id: str) -> tuple:
    """
    Holt RAG-Kontext, URLs und Content-Types aus LangSmith für eine spezifische Trace-ID.
    
    Die Documents befinden sich im Retriever-Output unter dem Key 'output' (nicht 'documents'!).
    Jedes Document hat 'page_content' und 'metadata' (mit 'url' und 'content_type').
    
    WICHTIG: Bei mehreren Retriever-Runs (z.B. _naive_retrieve + _advanced_retrieve)
    wird nur der Run mit den WENIGSTEN Dokumenten genommen, da das die finale Auswahl ist.
    
    Args:
        client: LangSmith Client
        trace_id: Die Trace-ID der Session
        
    Returns:
        Tuple (contexts, urls, content_types): Listen von RAG-Context-Chunks, URLs und Content-Types
    """
    try:
        # Hole alle Child-Runs für diese Trace
        child_runs = list(client.list_runs(
            project_name=LANGSMITH_PROJECT,
            trace_id=trace_id,
            is_root=False
        ))
        
        # Sammle alle Retriever-Runs mit ihren Dokumenten
        retriever_runs = []
        for child in child_runs:
            if child.run_type == "retriever":
                if child.outputs and isinstance(child.outputs, dict):
                    # Documents sind unter 'output' Key (nicht 'documents')!
                    documents = child.outputs.get('output', [])
                    if documents:
                        retriever_runs.append({
                            'name': child.name,
                            'documents': documents,
                            'doc_count': len(documents)
                        })
        
        if not retriever_runs:
            return [], [], []
        
        # Bei mehreren Retriever-Runs: Nimm den mit den WENIGSTEN Dokumenten
        # Das ist der finale Run (z.B. _advanced_retrieve mit Top-K nach ReRanking)
        # Bei gleichem Count: Nimm den letzten (zeitlich neuesten)
        final_run = min(retriever_runs, key=lambda x: x['doc_count'])
        
        contexts = []
        urls = []
        content_types = []
        
        for doc in final_run['documents']:
            if isinstance(doc, dict) and 'page_content' in doc:
                contexts.append(doc['page_content'])
                # Metadata extrahieren (ohne Fallbacks)
                metadata = doc.get('metadata', {})
                # URL und Content-Type direkt aus metadata
                urls.append(metadata.get('url'))
                content_types.append(metadata.get('content_type'))
        
        if contexts:
            return contexts, urls, content_types  # Tuple von 3 Listen zurückgeben
        
        return [], [], []  # Leere Listen wenn keine Kontexte gefunden
    
    except Exception as e:
        print(f"      ⚠️ LangSmith-Fehler: {str(e)[:100]}")
        return [], [], []  # Leere Listen bei Fehler


def get_token_usage_from_langsmith(client: Client, trace_id: str) -> dict:
    """
    Holt Token-Usage aus LangSmith für eine spezifische Trace-ID.
    
    Summiert alle Tokens aus LLM-Runs (ChatOllama/ChatOpenAI) innerhalb der Trace.
    Erfasst ReRanking-Tokens (Voyage) separat.
    
    Args:
        client: LangSmith Client
        trace_id: Die Trace-ID der Session
        
    Returns:
        dict: {
            'prompt_tokens': int, 
            'completion_tokens': int, 
            'total_tokens': int,
            'reranking_tokens': int  # NEU: Voyage ReRanking Tokens
        }
    """
    try:
        # Hole alle Child-Runs für diese Trace
        child_runs = list(client.list_runs(
            project_name=LANGSMITH_PROJECT,
            trace_id=trace_id,
            is_root=False
        ))
        
        total_prompt = 0
        total_completion = 0
        total_tokens = 0
        reranking_tokens = 0  # NEU: Separat für ReRanking
        
        for child in child_runs:
            # LLM-Runs haben run_type="llm"
            if child.run_type == "llm":
                # Prüfe ob es ein Reranker Run ist (Voyage, Cohere oder Local)
                is_reranking = child.name in ("VoyageReranker", "CohereReranker", "LocalReranker")
                
                # Token-Usage kann in verschiedenen Stellen sein:
                # 1. Direkt als Attribute: total_tokens, prompt_tokens, completion_tokens
                # 2. In extra['usage'] oder outputs['usage']
                
                run_tokens = 0
                run_prompt = 0
                run_completion = 0
                
                # Token-Extraktion mit Priorität (nur EINE Quelle verwenden, nicht addieren!)
                # Priorität: usage_metadata > outputs.usage > outputs.token_usage > direkte Attribute
                
                token_found = False
                
                # Methode 1: usage_metadata (Voyage ReRanking Format)
                if not token_found and child.outputs and isinstance(child.outputs, dict):
                    usage_metadata = child.outputs.get('usage_metadata', {})
                    if usage_metadata and usage_metadata.get('total_tokens'):
                        run_tokens = usage_metadata.get('total_tokens', 0) or 0
                        run_prompt = usage_metadata.get('input_tokens', 0) or 0
                        token_found = True
                
                # Methode 2: outputs.usage (OpenAI Format)
                if not token_found and child.outputs and isinstance(child.outputs, dict):
                    usage = child.outputs.get('usage', {})
                    if usage and (usage.get('total_tokens') or usage.get('prompt_tokens')):
                        run_prompt = usage.get('prompt_tokens', 0) or 0
                        run_completion = usage.get('completion_tokens', 0) or 0
                        run_tokens = usage.get('total_tokens', 0) or 0
                        token_found = True
                
                # Methode 3: outputs.token_usage (LangChain Format)
                if not token_found and child.outputs and isinstance(child.outputs, dict):
                    token_usage = child.outputs.get('token_usage', {})
                    if token_usage and (token_usage.get('total_tokens') or token_usage.get('prompt_tokens')):
                        run_prompt = token_usage.get('prompt_tokens', 0) or 0
                        run_completion = token_usage.get('completion_tokens', 0) or 0
                        run_tokens = token_usage.get('total_tokens', 0) or 0
                        token_found = True
                
                # Methode 4: outputs.llm_output.token_usage (älteres Format)
                if not token_found and child.outputs and isinstance(child.outputs, dict):
                    llm_output = child.outputs.get('llm_output', {})
                    if llm_output and isinstance(llm_output, dict):
                        tu = llm_output.get('token_usage', {})
                        if tu and (tu.get('total_tokens') or tu.get('prompt_tokens')):
                            run_prompt = tu.get('prompt_tokens', 0) or 0
                            run_completion = tu.get('completion_tokens', 0) or 0
                            run_tokens = tu.get('total_tokens', 0) or 0
                            token_found = True
                
                # Methode 5: Direkte Attribute
                if not token_found:
                    if hasattr(child, 'total_tokens') and child.total_tokens:
                        run_tokens = child.total_tokens
                    if hasattr(child, 'prompt_tokens') and child.prompt_tokens:
                        run_prompt = child.prompt_tokens
                    if hasattr(child, 'completion_tokens') and child.completion_tokens:
                        run_completion = child.completion_tokens
                    if run_tokens or run_prompt or run_completion:
                        token_found = True
                
                # Methode 6: extra.usage (Fallback)
                if not token_found and hasattr(child, 'extra') and child.extra and isinstance(child.extra, dict):
                    usage = child.extra.get('usage', {})
                    if usage and (usage.get('total_tokens') or usage.get('prompt_tokens')):
                        run_prompt = usage.get('prompt_tokens', 0) or 0
                        run_completion = usage.get('completion_tokens', 0) or 0
                        run_tokens = usage.get('total_tokens', 0) or 0
                
                # Zuordnung: ReRanking oder LLM
                if is_reranking:
                    reranking_tokens += run_tokens
                else:
                    total_prompt += run_prompt
                    total_completion += run_completion
                    total_tokens += run_tokens
        
        # Falls total_tokens nicht direkt verfügbar, berechne aus prompt + completion
        if total_tokens == 0 and (total_prompt > 0 or total_completion > 0):
            total_tokens = total_prompt + total_completion
        
        return {
            'prompt_tokens': total_prompt,
            'completion_tokens': total_completion,
            'total_tokens': total_tokens,
            'reranking_tokens': reranking_tokens  # NEU
        }
    
    except Exception as e:
        print(f"      ⚠️ Token-Usage Fehler: {str(e)[:100]}")
        return {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0, 'reranking_tokens': 0}


def stop_ollama_model(model_name: str):
    """
    Stoppt ein Ollama-Modell via CLI-Befehl, um GPU-Speicher freizugeben.
    Der Ollama-Server läuft weiter und kann andere Modelle laden.
    
    Args:
        model_name: Name des zu stoppenden Modells (z.B. 'llama3.1:8b')
    """
    import subprocess
    
    try:
        result = subprocess.run(
            ["ollama", "stop", model_name],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0:
            print(f"   ✅ Modell {model_name} gestoppt (GPU-Speicher freigegeben)")
        else:
            # Fehlerausgabe prüfen
            if "not running" in result.stderr.lower():
                print(f"   ℹ️ Modell {model_name} war nicht geladen")
            else:
                print(f"   ⚠️ ollama stop: {result.stderr.strip()}")
                
    except subprocess.TimeoutExpired:
        print(f"   ⚠️ Timeout beim Stoppen von {model_name}")
    except FileNotFoundError:
        print(f"   ⚠️ 'ollama' Befehl nicht gefunden - ist Ollama installiert?")
    except Exception as e:
        print(f"   ⚠️ Fehler beim Stoppen von {model_name}: {e}")


def stop_embedding_model():
    """
    Gibt das Embedding-Modell (SentenceTransformer) frei, um GPU/RAM-Speicher freizugeben.
    Löscht alle gecachten Modell-Instanzen aus dem RAG-Tool.
    """
    import torch
    
    try:
        # Lösche gecachte Embedding-Modelle aus dem RAG-Tool
        from src.tools.rag_tool import UniversityRAGTool
        
        # Iteriere über alle Instanzen und lösche Embedding-Modelle
        # Da _embedding_model ein Klassenattribut ist, setzen wir es auf None
        UniversityRAGTool._embedding_model = None
        
        # Garbage Collection und CUDA Cache leeren
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            torch.cuda.synchronize()
        
        print("   ✅ Embedding-Modell freigegeben (GPU/RAM-Speicher)")
        
    except Exception as e:
        print(f"   ⚠️ Fehler beim Freigeben des Embedding-Modells: {e}")


def generate_chatbot_responses(df: pd.DataFrame, agent, langsmith_client: Client) -> tuple:
    """
    Generiert Chatbot-Antworten für alle Fragen und sammelt RAG-Kontexte.
    Speichert nach jeder Frage einen inkrementellen Checkpoint.
    Bei Timeout (3 Min) wird der Agent neu gestartet.
    
    Returns:
        Tuple (dataset, response_times, urls_list, content_types_list, token_usage_list): 
        EvaluationDataset, Liste der Antwortzeiten, Liste der URL-Listen, Liste der Content-Type-Listen, Liste der Token-Usages
    """
    import pickle
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
    
    TIMEOUT_SECONDS = 180  # 3 Minuten
    
    print("\n🤖 Generiere Chatbot-Antworten...")
    print(f"   ⏱️ Timeout pro Frage: {TIMEOUT_SECONDS}s")
    print("=" * 80)
    
    # Checkpoint-Pfad mit Timestamp
    checkpoint_path = Path(__file__).parent / "data" / f"responses_checkpoint_{EVAL_TIMESTAMP}.pkl"
    checkpoint_path.parent.mkdir(exist_ok=True)
    
    # Prüfe ob inkrementeller Checkpoint existiert (für Fortsetzung nach Abbruch)
    samples = []
    response_times = []
    urls_list = []
    content_types_list = []
    token_usage_list = []  # Token-Usage pro Anfrage
    processed_ids = set()  # IDs die bereits im Checkpoint sind
    checkpoint_df = pd.DataFrame()  # DataFrame mit verarbeiteten Zeilen
    
    if checkpoint_path.exists():
        try:
            with open(checkpoint_path, 'rb') as f:
                checkpoint_data = pickle.load(f)
            
            # Prüfe ob Checkpoint gültig ist
            if isinstance(checkpoint_data, dict) and 'test_df' in checkpoint_data:
                saved_df = checkpoint_data['test_df']
                saved_dataset = checkpoint_data.get('dataset')
                
                if saved_dataset and hasattr(saved_dataset, 'samples'):
                    # Ermittle bereits verarbeitete IDs aus saved_df
                    processed_ids_in_checkpoint = set(saved_df['id'].tolist())
                    
                    # Prüfe welche IDs aus df (gefiltert!) noch fehlen
                    required_ids = set(df['id'].tolist())
                    
                    # Nur IDs nehmen, die SOWOHL im aktuellen df als AUCH im Checkpoint sind
                    ids_to_load = required_ids & processed_ids_in_checkpoint
                    
                    if len(ids_to_load) == len(required_ids):
                        # Alle benötigten IDs sind bereits beantwortet
                        # Lade nur die entsprechenden Daten aus dem Checkpoint
                        samples = [s for i, s in enumerate(saved_dataset.samples) if saved_df.iloc[i]['id'] in required_ids]
                        response_times = [rt for i, rt in enumerate(checkpoint_data.get('response_times', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in required_ids]
                        urls_list = [u for i, u in enumerate(checkpoint_data.get('urls_list', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in required_ids]
                        content_types_list = [ct for i, ct in enumerate(checkpoint_data.get('content_types_list', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in required_ids]
                        token_usage_list = [t for i, t in enumerate(checkpoint_data.get('token_usage_list', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in required_ids]
                        
                        print(f"📂 Checkpoint vollständig: Alle {len(samples)} gefilterten Fragen beantwortet")
                        dataset = EvaluationDataset(samples=samples)
                        return dataset, response_times, urls_list, content_types_list, token_usage_list
                    elif len(ids_to_load) > 0:
                        # Teilweise im Checkpoint vorhanden - lade diese Daten
                        samples = [s for i, s in enumerate(saved_dataset.samples) if i < len(saved_df) and saved_df.iloc[i]['id'] in ids_to_load]
                        response_times = [rt for i, rt in enumerate(checkpoint_data.get('response_times', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in ids_to_load]
                        urls_list = [u for i, u in enumerate(checkpoint_data.get('urls_list', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in ids_to_load]
                        content_types_list = [ct for i, ct in enumerate(checkpoint_data.get('content_types_list', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in ids_to_load]
                        token_usage_list = [t for i, t in enumerate(checkpoint_data.get('token_usage_list', [])) if i < len(saved_df) and saved_df.iloc[i]['id'] in ids_to_load]
                        checkpoint_df = saved_df[saved_df['id'].isin(ids_to_load)].copy()
                        processed_ids = ids_to_load
                        
                        print(f"📂 Checkpoint teilweise geladen: {len(samples)} Fragen bereits beantwortet")
                        print(f"   → Fehlende IDs: {sorted(required_ids - ids_to_load)}")
                    else:
                        # Keine der benötigten IDs im Checkpoint - Starte neu
                        print(f"📂 Checkpoint existiert, aber für andere IDs - Starte neu")
        except Exception as e:
            print(f"⚠️ Checkpoint-Ladefehler: {e} - Starte neu")
            samples, response_times, urls_list, content_types_list, token_usage_list, processed_ids = [], [], [], [], [], set()
            checkpoint_df = pd.DataFrame()
    
    # Iteriere über alle Fragen, aber überspringe bereits verarbeitete
    total_questions = len(df)
    for i in range(total_questions):
        row = df.iloc[i]
        question_id = row['id']
        
        # Überspringe bereits verarbeitete IDs
        if question_id in processed_ids:
            continue
        
        question = row['question']
        expected_answer = row['expected_answer']
        
        print(f"\n[ID {question_id}] {question[:70]}...")
        
        # Memory löschen für isolierte Evaluation
        agent.clear_memory()
        
        # Chatbot fragen - mit Session-ID für LangSmith-Tracking
        print(f"   💬 Chatbot fragen...")
        import uuid
        session_id = str(uuid.uuid4())
        
        # Zeit messen für Antwortgenerierung
        response_start = time.time()
        
        # Direkte Anfrage an den Chatbot (ohne ThreadPoolExecutor)
        answer = agent.chat(question, session_id)
        
        response_time = time.time() - response_start
        response_times.append(response_time)
        
        print(f"   ✅ Antwort: {answer[:80]}... ({response_time:.2f}s)")
        
        # Warten damit LangSmith Trace vollständig ist
        time.sleep(2)  # Erhöht auf 2s für zuverlässigere Synchronisation
        
        # RAG-Kontext aus LangSmith holen - nach Session-ID suchen
        print(f"   🔍 Hole RAG-Kontext aus LangSmith...")
        
        # Hole die letzten Runs und suche manuell nach unserer Session-ID
        matching_run = None
        max_retries = 3
        
        for retry in range(max_retries):
            recent_runs = list(langsmith_client.list_runs(
                project_name=LANGSMITH_PROJECT,
                is_root=True,
                limit=10  # Mehr Runs holen für Sicherheit
            ))
            
            # Suche nach der Session-ID in Metadata
            for run in recent_runs:
                if run.extra and isinstance(run.extra, dict):
                    metadata = run.extra.get('metadata', {})
                    if metadata.get('session_id') == session_id:
                        matching_run = run
                        break
            
            if matching_run:
                break
            
            # Warte und versuche erneut
            if retry < max_retries - 1:
                time.sleep(1)
        
        # Fallback: Nehme den neuesten Run wenn Session-ID nicht gefunden
        if not matching_run and recent_runs:
            matching_run = recent_runs[0]
            print(f"   ⚠️ Session-ID {session_id[:8]}... nicht in Metadata gefunden, verwende neuesten Run")
        
        contexts = []  # Leere Liste als Default
        urls = []  # Leere Liste als Default
        content_types = []  # Leere Liste als Default
        
        if matching_run:
            trace_id = matching_run.trace_id
            contexts, urls, content_types = get_rag_context_from_langsmith(langsmith_client, trace_id)
            # Token-Usage aus LangSmith holen
            token_usage = get_token_usage_from_langsmith(langsmith_client, trace_id)
            print(f"   ✅ Run gefunden mit Session-ID: {session_id[:8]}...")
        else:
            print(f"   ⚠️ Kein Run mit Session-ID {session_id[:8]}... gefunden")
            token_usage = {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0, 'reranking_tokens': 0}
        
        urls_list.append(urls)
        content_types_list.append(content_types)
        token_usage_list.append(token_usage)
        
        total_chars = sum(len(c) for c in contexts)
        print(f"   📄 Kontext: {len(contexts)} chunks, {total_chars} Zeichen")
        print(f"   🔗 URLs: {len(urls)} Quellen")
        print(f"   📁 Content-Types: {set(content_types)}")
        if token_usage['total_tokens'] > 0 or token_usage.get('reranking_tokens', 0) > 0:
            llm_tokens = token_usage['total_tokens']
            rerank_tokens = token_usage.get('reranking_tokens', 0)
            if rerank_tokens > 0:
                print(f"   📊 LLM-Tokens: {llm_tokens} (Prompt: {token_usage['prompt_tokens']}, Completion: {token_usage['completion_tokens']})")
                print(f"   🔄 ReRanking-Tokens: {rerank_tokens}")
            else:
                print(f"   📊 Tokens: {llm_tokens} (Prompt: {token_usage['prompt_tokens']}, Completion: {token_usage['completion_tokens']})")
        
        # RAGAS-Sample erstellen
        sample = SingleTurnSample(
            user_input=question,
            response=answer,
            retrieved_contexts=contexts,  # Jetzt bereits eine Liste von Chunks
            reference=expected_answer
        )
        samples.append(sample)
        
        # Füge aktuelle Zeile zum checkpoint_df hinzu
        new_row = row.to_frame().T.copy()
        new_row['id'] = int(question_id)  # Sicherstellen dass ID als int gespeichert wird
        checkpoint_df = pd.concat([checkpoint_df, new_row], ignore_index=True)
        processed_ids.add(question_id)
        
        # 💾 INKREMENTELLER CHECKPOINT nach jeder Frage
        dataset = EvaluationDataset(samples=samples)
        checkpoint_data = {
            'dataset': dataset,
            'test_df': checkpoint_df,
            'response_times': response_times,
            'urls_list': urls_list,
            'content_types_list': content_types_list,
            'token_usage_list': token_usage_list
        }
        with open(checkpoint_path, 'wb') as f:
            pickle.dump(checkpoint_data, f)
        print(f"   💾 Checkpoint: {len(samples)}/{len(df)} Fragen gespeichert (IDs: {len(processed_ids)})")
    
    print("\n" + "=" * 80)
    print(f"✅ {len(samples)} Antworten generiert\n")
    if response_times:
        print(f"   ⏱️ Durchschn. Antwortzeit: {sum(response_times)/len(response_times):.2f}s")
        print(f"   ⏱️ Gesamt Antwortzeit: {sum(response_times):.2f}s\n")
    if token_usage_list:
        total_llm_tokens = sum(t.get('total_tokens', 0) for t in token_usage_list)
        total_rerank_tokens = sum(t.get('reranking_tokens', 0) for t in token_usage_list)
        if total_rerank_tokens > 0:
            print(f"   📊 Gesamt LLM-Tokens: {total_llm_tokens}")
            print(f"   🔄 Gesamt ReRanking-Tokens: {total_rerank_tokens}\n")
        else:
            print(f"   📊 Gesamt Tokens: {total_llm_tokens}\n")
    
    dataset = EvaluationDataset(samples=samples)
    return dataset, response_times, urls_list, content_types_list, token_usage_list


# ============================================================================
# KONFIGURATION
# ============================================================================
# Limit für Testfragen (None = alle, z.B. 5 für Test)
TEST_LIMIT = None # None = alle Fragen evaluieren


def run_ragas_evaluation(dataset: EvaluationDataset, run_local: bool = None) -> tuple:
    """
    Führt RAGAS-Evaluation durch.
    Unterstützt sowohl lokale Ollama-Modelle als auch OpenAI Cloud-Modelle.
    
    Args:
        dataset: EvaluationDataset mit Samples
        run_local: True = lokales Ollama, False = OpenAI Cloud, None = aus Config
    
    Returns:
        Tuple (results_df, evaluation_time): DataFrame mit Ergebnissen und Evaluationszeit in Sekunden
    """
    # Bestimme Modus (Parameter überschreibt Config)
    use_local = run_local if run_local is not None else RUN_EVALUATION_LOCAL
    
    print("🚀 Starte RAGAS-Evaluation...")
    print("=" * 80)
    
    if use_local:
        # ====================================================================
        # LOKALE EVALUATION mit Ollama LLM
        # ====================================================================
        print("   🏠 Modus: LOKAL (Ollama LLM)")
        
        # Separates LLM für RAGAS-Evaluation (gleiches Setup wie Chatbot, nur anderes Modell)
        llm = ChatOllama(
            model=RAGAS_EVAL_MODEL,  # Separates Modell für Evaluation
            base_url=OLLAMA_BASE_URL,
            temperature=TEMPERATURE,  # Gleiche Parameter wie Chatbot
            seed=RANDOM_SEED,
            num_ctx=CONTEXT_WINDOW
        )
        print(f"   RAGAS-LLM: {RAGAS_EVAL_MODEL} @ {OLLAMA_BASE_URL} (ctx={CONTEXT_WINDOW}, temp={TEMPERATURE}, seed={RANDOM_SEED})")
        print(f"   (Chatbot verwendet: {OLLAMA_MODEL})")
        
        # RunConfig für Ollama-Evaluation
        # max_workers=4: Ollama verarbeitet Requests sequentiell, parallele Worker verursachen Timeouts
        # timeout=300: 5 Minuten pro Metrik-Berechnung (erhöht wegen LLM-Latenz)
        run_config = RunConfig(
            max_workers=4,
            seed=RANDOM_SEED,
            timeout=300
        )
    else:
        # ====================================================================
        # CLOUD EVALUATION mit OpenAI (nur LLM, Embeddings bleiben lokal!)
        # ====================================================================
        print("   ☁️ Modus: CLOUD (OpenAI LLM)")
        
        # Prüfe API-Key
        if not OPENAI_API_KEY:
            raise ValueError("OPENAI_API_KEY nicht gesetzt! Bitte in .env konfigurieren.")
        
        # OpenAI LLM für RAGAS-Evaluation (nur der Judge!)
        from langchain_openai import ChatOpenAI
        from ragas.llms import LangchainLLMWrapper
        
        openai_llm = ChatOpenAI(
            model=OPENAI_EVAL_MODEL,
            api_key=OPENAI_API_KEY,
            temperature=0.0,  # Deterministische Evaluation
            seed=RANDOM_SEED,  # Reproduzierbarkeit
            max_retries=5,  # Automatisches Retry bei Rate-Limits (429)
            max_tokens=4096  # Erhöht für RAGAS-Metriken (Faithfulness benötigt lange Outputs)
        )
        llm = LangchainLLMWrapper(openai_llm)
        print(f"   RAGAS-LLM: {OPENAI_EVAL_MODEL} (OpenAI Cloud, temp=0.0, seed={RANDOM_SEED}, retries=5, max_tokens=4096)")
        print(f"   (Chatbot verwendet: {OLLAMA_MODEL})")
        
        # RunConfig für OpenAI-Evaluation
        # max_workers=150: Hohe Parallelität
        # timeout=1800: Ausreichend Zeit für komplexe Metriken
        run_config = RunConfig(
            max_workers=150,
            seed=RANDOM_SEED,
            timeout=1800,
            max_retries=5
        )
    
    # ========================================================================
    # EMBEDDINGS: IMMER lokal mit Ollama (unabhängig vom LLM-Modus)
    # ========================================================================
    RAGAS_EMBEDDING_MODEL = "embeddinggemma"
    RAGAS_MAX_SEQ_LENGTH = 1024
    
    ollama_embeddings = OllamaEmbeddings(
        model=RAGAS_EMBEDDING_MODEL,
        base_url=OLLAMA_BASE_URL,
        num_ctx=RAGAS_MAX_SEQ_LENGTH
    )
    
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        langchain_embeddings = LangchainEmbeddingsWrapper(ollama_embeddings)
    
    print(f"   Embeddings: {RAGAS_EMBEDDING_MODEL} @ {OLLAMA_BASE_URL} (LOKAL, max_seq_length={RAGAS_MAX_SEQ_LENGTH})")
    print(f"   (RAG verwendet: {SENTENCE_TRANSFORMER_MODEL})")
    
    # SemanticSimilarity mit Embeddings
    semantic_similarity = SemanticSimilarity(embeddings=langchain_embeddings)
    
    # ResponseRelevancy benötigt LangChain-Interface (embed_query/embed_documents)
    response_relevancy = ResponseRelevancy(embeddings=langchain_embeddings)
    
    # ContextEntityRecall Metrik (misst Entitäten-Recall zwischen Referenz und Kontext)
    context_entity_recall = ContextEntityRecall()
    
    # Standard RAGAS-Metriken
    metrics = [
        faithfulness,           # Ist Antwort treu zum Kontext?
        context_recall,         # Wurden alle relevanten Infos abgerufen?
        context_precision,      # Sind relevante Chunks höher gerankt?
        semantic_similarity,    # Semantische Ähnlichkeit zwischen Antwort und Referenz
        context_entity_recall,  # Entitäten-Recall zwischen Referenz und Kontext
        response_relevancy      # Relevanz der Antwort zur gestellten Frage
    ]
    print(f"   Metriken: {[m.name for m in metrics]}")
    print(f"\n   ⏳ Evaluiere {len(dataset.samples)} Samples...")
    if use_local:
        print(f"   💡 Dies kann mehrere Minuten dauern (ca. 1-2 Min pro Sample)\n")
    else:
        print(f"   💡 OpenAI Cloud: ca. 10-30 Sek pro Sample\n")
    
    # Zeit messen für Evaluation
    eval_start = time.time()
    
    # Evaluation durchführen
    results = evaluate(
        dataset=dataset,
        metrics=metrics,
        llm=llm,
        run_config=run_config,
        raise_exceptions=False  # Weiter bei Fehlern
    )
    
    evaluation_time = time.time() - eval_start
    
    print(f"\n   ✅ Evaluation abgeschlossen in {evaluation_time:.2f}s")
    print(f"   ⏱️ Durchschn. pro Sample: {evaluation_time/len(dataset.samples):.2f}s\n")
    
    # ========================================================================
    # BERT-SCORE: Token-Level semantische Ähnlichkeit (nach RAGAS-Evaluation)
    # ========================================================================
    results_df = results.to_pandas()
    
    if BERT_SCORE_AVAILABLE:
        print("📊 Berechne BERT-Score...")
        bert_start = time.time()
        
        try:
            # Extrahiere responses und references aus dataset
            responses = [s.response for s in dataset.samples]
            references = [s.reference for s in dataset.samples]
            
            # BERT-Score berechnen (Multilingual: xlm-roberta-large)
            # Unterstützt Deutsch + Englisch gemischt
            # Unterdrücke Tokenizer-Warnungen
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore")
                P, R, F1 = bert_score_fn(
                    responses, 
                    references, 
                    model_type="xlm-roberta-large",  # Multilingual (100+ Sprachen, inkl. DE + EN)
                    lang="de",  # Sprache für Baseline-Rescaling (DE funktioniert auch für gemischte Texte)
                    verbose=False,
                    rescale_with_baseline=True  # Bessere Interpretierbarkeit
                )
            
            # Zu results_df hinzufügen
            results_df['bert_precision'] = P.numpy()
            results_df['bert_recall'] = R.numpy()
            results_df['bert_f1'] = F1.numpy()
            
            bert_time = time.time() - bert_start
            print(f"   ✅ BERT-Score berechnet in {bert_time:.2f}s")
            print(f"   📈 Durchschn. BERT-F1: {F1.mean():.3f}")
            
        except Exception as e:
            print(f"   ⚠️ BERT-Score Fehler: {e}")
            results_df['bert_precision'] = None
            results_df['bert_recall'] = None
            results_df['bert_f1'] = None
    else:
        print("⚠️ BERT-Score übersprungen (nicht installiert)")
        results_df['bert_precision'] = None
        results_df['bert_recall'] = None
        results_df['bert_f1'] = None
    
    return results_df, evaluation_time


def display_and_save_results(results_df: pd.DataFrame, test_df: pd.DataFrame, 
                              response_times: List[float] = None, urls_list: List[List[str]] = None,
                              content_types_list: List[List[str]] = None, evaluation_time: float = None,
                              token_usage_list: List[dict] = None):
    """
    Zeigt Ergebnisse an und speichert sie.
    
    Args:
        results_df: DataFrame mit RAGAS-Ergebnissen
        test_df: DataFrame mit Testdaten
        response_times: Liste der Antwortzeiten pro Frage (optional)
        urls_list: Liste von URL-Listen pro Frage (optional)
        content_types_list: Liste von Content-Type-Listen pro Frage (optional)
        evaluation_time: Gesamtzeit für RAGAS-Evaluation in Sekunden (optional)
        token_usage_list: Liste von Token-Usage-Dicts pro Frage (optional)
    """
    from datetime import datetime
    
    # IDs, Kategorien und Schwierigkeiten hinzufügen
    results_df['id'] = test_df['id'].values[:len(results_df)]
    results_df['category'] = test_df['category'].values[:len(results_df)]
    results_df['difficulty'] = test_df['difficulty'].values[:len(results_df)]
    
    # Antwortzeiten hinzufügen (falls vorhanden)
    if response_times:
        results_df['response_time_seconds'] = response_times[:len(results_df)]
    else:
        results_df['response_time_seconds'] = None
    
    # URLs hinzufügen (falls vorhanden)
    if urls_list:
        results_df['retrieved_urls'] = [str(urls) for urls in urls_list[:len(results_df)]]
    else:
        results_df['retrieved_urls'] = None
    
    # Content-Types hinzufügen (falls vorhanden)
    if content_types_list:
        results_df['retrieved_content_types'] = [str(types) for types in content_types_list[:len(results_df)]]
    else:
        results_df['retrieved_content_types'] = None
    
    # RR_at5 berechnen (Reciprocal Rank: Position der Referenz-URL in retrieved contexts)
    if urls_list:
        rr_at5_values = []
        hit_at5_values = []
        for i in range(len(results_df)):
            context_hint = test_df['context_hint'].iloc[i] if i < len(test_df) else None
            retrieved_urls = urls_list[i] if i < len(urls_list) else []
            rr = calculate_RR_at5(context_hint, retrieved_urls)
            rr_at5_values.append(rr)
            hit_at5_values.append(1.0 if rr > 0 else 0.0)
        results_df['RR_at5'] = rr_at5_values
        results_df['hit_at5'] = hit_at5_values
    else:
        results_df['RR_at5'] = None
        results_df['hit_at5'] = None
    
    # Latency als Kopie von response_time_seconds
    if 'response_time_seconds' in results_df.columns:
        results_df['latency'] = results_df['response_time_seconds']
    else:
        results_df['latency'] = None
    
    # Token-Usage hinzufügen (falls vorhanden)
    if token_usage_list:
        results_df['prompt_tokens'] = [t.get('prompt_tokens', 0) for t in token_usage_list[:len(results_df)]]
        results_df['completion_tokens'] = [t.get('completion_tokens', 0) for t in token_usage_list[:len(results_df)]]
        results_df['total_tokens'] = [t.get('total_tokens', 0) for t in token_usage_list[:len(results_df)]]
        results_df['reranking_tokens'] = [t.get('reranking_tokens', 0) for t in token_usage_list[:len(results_df)]]
    else:
        results_df['prompt_tokens'] = None
        results_df['completion_tokens'] = None
        results_df['total_tokens'] = None
        results_df['reranking_tokens'] = None
    
    # ============================================================================
    # SOFORT SPEICHERN - Rohdaten CSV mit Timestamp (bevor irgendwas schiefgehen kann)
    # ============================================================================
    output_path_raw = Path(__file__).parent / "data" / f"ragas_results_raw_{EVAL_TIMESTAMP}_rerun.csv"
    try:
        results_df.to_csv(output_path_raw, index=False, encoding='utf-8-sig')
        print(f"\n💾 ROHDATEN GESPEICHERT: {output_path_raw}")
    except Exception as e:
        print(f"\n⚠️ Fehler beim Speichern der Rohdaten: {e}")
    
    print("\n" + "=" * 80)
    print("📊 RAGAS-EVALUATION ERGEBNISSE")
    print("=" * 80)
    
    # Alle Metriken (inkl. BERT-Score und Tokens)
    all_metrics = ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 
                   'context_entity_recall', 'answer_relevancy', 'bert_f1', 'bert_precision', 'bert_recall',
                   'RR_at5', 'hit_at5', 'latency', 'total_tokens']
    
    # Gesamtscores
    print("\n📈 Durchschnittliche Scores:")
    print("-" * 80)
    for metric in all_metrics:
        if metric in results_df.columns and results_df[metric].notna().any():
            avg = results_df[metric].mean()
            # Anzeigename für Zusammenfassung
            display_names = {'RR_at5': 'MRR@5', 'hit_at5': 'Hit@5', 'total_tokens': 'Tokens (avg)'}
            display_name = display_names.get(metric, metric)
            if metric == 'total_tokens':
                print(f"   {display_name:20s}: {avg:.0f}")
            else:
                print(f"   {display_name:20s}: {avg:.3f}")
    
    # Nach Kategorie (NaN ausfiltern)
    print("\n📁 Scores nach Kategorie:")
    print("-" * 80)
    display_categories = [c for c in results_df['category'].unique() if pd.notna(c)]
    for category in sorted(display_categories):
        cat_df = results_df[results_df['category'] == category]
        print(f"\n   {category}:")
        for metric in all_metrics:
            if metric in cat_df.columns and cat_df[metric].notna().any():
                avg = cat_df[metric].mean()
                display_name = 'MRR@5' if metric == 'RR_at5' else ('Hit@5' if metric == 'hit_at5' else metric)
                print(f"      {display_name:20s}: {avg:.3f}")
        # Token-Statistiken pro Kategorie
        if 'prompt_tokens' in cat_df.columns and cat_df['prompt_tokens'].notna().any():
            avg_input = cat_df['prompt_tokens'].mean()
            print(f"      {'Tokens (Input)':20s}: {avg_input:.0f}")
        if 'completion_tokens' in cat_df.columns and cat_df['completion_tokens'].notna().any():
            avg_output = cat_df['completion_tokens'].mean()
            print(f"      {'Tokens (Output)':20s}: {avg_output:.0f}")
        if 'total_tokens' in cat_df.columns and cat_df['total_tokens'].notna().any():
            avg_total = cat_df['total_tokens'].mean()
            print(f"      {'Tokens (Gesamt)':20s}: {avg_total:.0f}")
    
    # Nach Schwierigkeit
    print("\n⚡ Scores nach Schwierigkeit:")
    print("-" * 80)
    for difficulty in ['easy', 'medium', 'hard']:
        diff_df = results_df[results_df['difficulty'] == difficulty]
        if len(diff_df) > 0:
            print(f"\n   {difficulty.upper()}:")
            for metric in all_metrics:
                if metric in diff_df.columns and diff_df[metric].notna().any():
                    avg = diff_df[metric].mean()
                    display_name = 'MRR@5' if metric == 'RR_at5' else ('Hit@5' if metric == 'hit_at5' else metric)
                    print(f"      {display_name:20s}: {avg:.3f}")
            # Token-Statistiken pro Schwierigkeit
            if 'prompt_tokens' in diff_df.columns and diff_df['prompt_tokens'].notna().any():
                avg_input = diff_df['prompt_tokens'].mean()
                print(f"      {'Tokens (Input)':20s}: {avg_input:.0f}")
            if 'completion_tokens' in diff_df.columns and diff_df['completion_tokens'].notna().any():
                avg_output = diff_df['completion_tokens'].mean()
                print(f"      {'Tokens (Output)':20s}: {avg_output:.0f}")
            if 'total_tokens' in diff_df.columns and diff_df['total_tokens'].notna().any():
                avg_total = diff_df['total_tokens'].mean()
                print(f"      {'Tokens (Gesamt)':20s}: {avg_total:.0f}")

    # Speichern in CSV mit Timestamp (alle Spalten)
    output_path_csv = Path(__file__).parent / "data" / f"ragas_results_{EVAL_TIMESTAMP}_rerun.csv"
    
    # Berechne Anzahl der Context-Chunks
    results_df['context_count'] = results_df['retrieved_contexts'].apply(lambda x: len(x) if isinstance(x, list) else 0)
    
    # Entferne Zeilenumbrüche aus Textfeldern für saubere CSV
    text_columns = ['user_input', 'response', 'reference']
    for col in text_columns:
        if col in results_df.columns:
            results_df[col] = results_df[col].apply(lambda x: x.replace('\n', ' ').replace('\r', ' ') if isinstance(x, str) else x)
    
    # Konvertiere retrieved_contexts zu String ohne Zeilenumbrüche
    results_df['retrieved_contexts'] = results_df['retrieved_contexts'].apply(
        lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if isinstance(x, list) else str(x)
    )
    
    # Konvertiere retrieved_urls zu String ohne Zeilenumbrüche (falls vorhanden)
    if 'retrieved_urls' in results_df.columns:
        results_df['retrieved_urls'] = results_df['retrieved_urls'].apply(
            lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if x else ''
        )
    
    # Konvertiere retrieved_content_types zu String ohne Zeilenumbrüche (falls vorhanden)
    if 'retrieved_content_types' in results_df.columns:
        results_df['retrieved_content_types'] = results_df['retrieved_content_types'].apply(
            lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if x else ''
        )
    
    # CSV mit allen wichtigen Spalten (erweitert um response_time, urls, content_types, alle Metriken inkl. BERT-Score und Tokens)
    csv_columns = ['id', 'category', 'difficulty', 'user_input', 'response', 
                   'reference', 'retrieved_contexts', 'retrieved_urls', 'retrieved_content_types',
                   'faithfulness', 'context_recall', 'context_precision', 'semantic_similarity',
                   'context_entity_recall', 'answer_relevancy', 'bert_f1', 'bert_precision', 'bert_recall',
                   'RR_at5', 'hit_at5', 'latency', 'prompt_tokens', 'completion_tokens', 'total_tokens',
                   'reranking_tokens', 'context_count', 'response_time_seconds']
    
    # Nur vorhandene Spalten verwenden
    csv_columns = [col for col in csv_columns if col in results_df.columns]
    csv_df = results_df[csv_columns].copy()
    
    # ============================================================================
    # METADATEN-ZEILEN: Modelle, Zeitstempel, Dauern
    # ============================================================================
    metadata_rows = []
    
    # Metadaten-Zeile 1: Allgemeine Infos
    meta1 = {col: '' for col in csv_columns}
    meta1['id'] = 'META'
    meta1['category'] = 'Evaluation Metadaten'
    meta1['difficulty'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta1['user_input'] = f'Chatbot: {OLLAMA_MODEL} (ctx=dynamisch, temp={TEMPERATURE}, seed={RANDOM_SEED})'
    meta1['response'] = f'RAGAS-LLM: {RAGAS_EVAL_MODEL} (ctx={CONTEXT_WINDOW}, temp={TEMPERATURE}, seed={RANDOM_SEED})'
    meta1['reference'] = f'Testset: {len(results_df)} Fragen'
    if evaluation_time:
        meta1['retrieved_contexts'] = f'Eval-Zeit: {evaluation_time:.2f}s'
    if response_times:
        meta1['retrieved_urls'] = f'Antwort-Zeit gesamt: {sum(response_times):.2f}s'
    metadata_rows.append(meta1)
    
    # ============================================================================
    # DURCHSCHNITTE: Gesamt, pro Kategorie, pro Schwierigkeit, kombiniert
    # ============================================================================
    metric_cols = ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 
                   'context_entity_recall', 'answer_relevancy', 'bert_f1', 'bert_precision', 'bert_recall',
                   'RR_at5', 'hit_at5', 'latency', 'prompt_tokens', 'completion_tokens', 'total_tokens', 'reranking_tokens']
    
    # Gesamtdurchschnitt
    avg_row = {col: '' for col in csv_columns}
    avg_row['id'] = 'AVG'
    avg_row['category'] = 'GESAMT'
    avg_row['difficulty'] = f'n={len(results_df)}'
    for metric in metric_cols:
        if metric in results_df.columns:
            avg_row[metric] = results_df[metric].mean()
    if 'response_time_seconds' in results_df.columns and results_df['response_time_seconds'].notna().any():
        avg_row['response_time_seconds'] = results_df['response_time_seconds'].mean()
    metadata_rows.append(avg_row)
    
    # Durchschnitte pro Kategorie (NaN-Werte ausfiltern)
    categories = [c for c in results_df['category'].unique() if pd.notna(c)]
    for category in sorted(categories):
        cat_df = results_df[results_df['category'] == category]
        cat_row = {col: '' for col in csv_columns}
        cat_row['id'] = 'AVG'
        cat_row['category'] = category
        cat_row['difficulty'] = f'n={len(cat_df)}'
        for metric in metric_cols:
            if metric in cat_df.columns:
                cat_row[metric] = cat_df[metric].mean()
        if 'response_time_seconds' in cat_df.columns and cat_df['response_time_seconds'].notna().any():
            cat_row['response_time_seconds'] = cat_df['response_time_seconds'].mean()
        metadata_rows.append(cat_row)
    
    # Durchschnitte pro Schwierigkeit
    for difficulty in ['easy', 'medium', 'hard']:
        diff_df = results_df[results_df['difficulty'] == difficulty]
        if len(diff_df) > 0:
            diff_row = {col: '' for col in csv_columns}
            diff_row['id'] = 'AVG'
            diff_row['category'] = f'Schwierigkeit: {difficulty.upper()}'
            diff_row['difficulty'] = f'n={len(diff_df)}'
            for metric in metric_cols:
                if metric in diff_df.columns:
                    diff_row[metric] = diff_df[metric].mean()
            if 'response_time_seconds' in diff_df.columns and diff_df['response_time_seconds'].notna().any():
                diff_row['response_time_seconds'] = diff_df['response_time_seconds'].mean()
            metadata_rows.append(diff_row)
    
    # Durchschnitte pro Kategorie + Schwierigkeit (kombiniert)
    for category in sorted(categories):  # Verwende bereits gefilterte categories-Liste
        for difficulty in ['easy', 'medium', 'hard']:
            combo_df = results_df[(results_df['category'] == category) & (results_df['difficulty'] == difficulty)]
            if len(combo_df) > 0:
                combo_row = {col: '' for col in csv_columns}
                combo_row['id'] = 'AVG'
                combo_row['category'] = f'{category} / {difficulty.upper()}'
                combo_row['difficulty'] = f'n={len(combo_df)}'
                for metric in metric_cols:
                    if metric in combo_df.columns:
                        combo_row[metric] = combo_df[metric].mean()
                if 'response_time_seconds' in combo_df.columns and combo_df['response_time_seconds'].notna().any():
                    combo_row['response_time_seconds'] = combo_df['response_time_seconds'].mean()
                metadata_rows.append(combo_row)
    
    # Metadaten-DataFrame erstellen und anhängen
    meta_df = pd.DataFrame(metadata_rows)
    csv_df = pd.concat([csv_df, meta_df], ignore_index=True)
    
    # Speichere mit UTF-8-BOM für korrekte Umlaut-Darstellung
    csv_df.to_csv(output_path_csv, index=False, encoding='utf-8-sig', sep=',', quoting=1)
    
    # Excel mit Formatierung erstellen (mit Timestamp)
    output_path_excel = Path(__file__).parent / "data" / f"ragas_results_{EVAL_TIMESTAMP}_rerun.xlsx"
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Sheet 1: Detaillierte Ergebnisse
        ws_details = wb.active
        ws_details.title = "Detaillierte Ergebnisse"
        
        # Header-Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Daten schreiben - NUR results_df (ohne AVG-Zeilen), csv_df enthält Metadaten
        # Verwende die gleichen Spalten wie csv_df, aber aus results_df
        excel_details_df = results_df[csv_columns].copy()
        for r_idx, row in enumerate(dataframe_to_rows(excel_details_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatieren
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Metriken-Spalten (faithfulness, context_recall, context_precision) farbig
                    if c_idx in [8, 9, 10]:  # Metrik-Spalten
                        if isinstance(value, (int, float)):
                            if value >= 0.8:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value >= 0.6:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.number_format = '0.000'
                    
                    # Text-Wrap für lange Texte
                    if c_idx in [4, 5, 6, 7]:  # user_input, response, reference, contexts
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Spaltenbreiten anpassen
        ws_details.column_dimensions['A'].width = 8   # id
        ws_details.column_dimensions['B'].width = 20  # category
        ws_details.column_dimensions['C'].width = 12  # difficulty
        ws_details.column_dimensions['D'].width = 50  # user_input
        ws_details.column_dimensions['E'].width = 60  # response
        ws_details.column_dimensions['F'].width = 50  # reference
        ws_details.column_dimensions['G'].width = 40  # contexts
        ws_details.column_dimensions['H'].width = 30  # retrieved_urls
        ws_details.column_dimensions['I'].width = 20  # retrieved_content_types
        ws_details.column_dimensions['J'].width = 13  # faithfulness
        ws_details.column_dimensions['K'].width = 13  # context_recall
        ws_details.column_dimensions['L'].width = 15  # context_precision
        ws_details.column_dimensions['M'].width = 17  # semantic_similarity
        ws_details.column_dimensions['N'].width = 20  # context_entity_recall
        ws_details.column_dimensions['O'].width = 18  # response_relevancy
        ws_details.column_dimensions['P'].width = 10  # RR_at5
        ws_details.column_dimensions['Q'].width = 10  # hit_at5
        ws_details.column_dimensions['R'].width = 12  # context_count
        ws_details.column_dimensions['S'].width = 15  # response_time_seconds
        
        # Sheet 2: Zusammenfassung
        ws_summary = wb.create_sheet("Zusammenfassung")
        
        # Titel
        ws_summary['A1'] = "📊 RAGAS-Evaluation Zusammenfassung"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')
        
        # Durchschnittliche Scores
        row = 3
        ws_summary[f'A{row}'] = "Durchschnittliche Scores"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Excel-Metriken inkl. BERT-Score und Tokens
        excel_metrics = ['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 
                         'context_entity_recall', 'answer_relevancy', 'bert_f1', 'bert_precision', 'bert_recall', 
                         'RR_at5', 'hit_at5', 'latency', 'prompt_tokens', 'completion_tokens', 'total_tokens', 'reranking_tokens']
        
        for metric in excel_metrics:
            if metric in results_df.columns and results_df[metric].notna().any():
                avg = results_df[metric].mean()
                # Anzeigename für Excel
                display_names = {'RR_at5': 'MRR@5', 'hit_at5': 'Hit@5', 'bert_f1': 'BERT-F1', 
                                 'bert_precision': 'BERT-Precision', 'bert_recall': 'BERT-Recall',
                                 'prompt_tokens': 'Prompt Tokens (avg)', 'completion_tokens': 'Completion Tokens (avg)',
                                 'total_tokens': 'Total Tokens (avg)', 'reranking_tokens': 'ReRanking Tokens (avg)'}
                display_name = display_names.get(metric, metric)
                ws_summary[f'A{row}'] = display_name
                ws_summary[f'B{row}'] = avg
                
                # Token-Metriken als Ganzzahlen formatieren
                if 'tokens' in metric:
                    ws_summary[f'B{row}'].number_format = '#,##0'
                else:
                    ws_summary[f'B{row}'].number_format = '0.000'
                
                # Farbe basierend auf Score (nicht für latency und tokens)
                if metric not in ['latency', 'prompt_tokens', 'completion_tokens', 'total_tokens', 'reranking_tokens']:
                    if avg >= 0.8:
                        ws_summary[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif avg >= 0.6:
                        ws_summary[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1
        
        # Nach Kategorie
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Kategorie"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header für Kategorie-Tabelle (mit allen 3 BERT-Metriken und Tokens)
        ws_summary[f'A{row}'] = "Kategorie"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        ws_summary[f'E{row}'] = "Semantic Similarity"
        ws_summary[f'F{row}'] = "Context Entity Recall"
        ws_summary[f'G{row}'] = "Answer Relevancy"
        ws_summary[f'H{row}'] = "BERT-F1"
        ws_summary[f'I{row}'] = "BERT-Precision"
        ws_summary[f'J{row}'] = "BERT-Recall"
        ws_summary[f'K{row}'] = "MRR@5"
        ws_summary[f'L{row}'] = "Hit@5"
        ws_summary[f'M{row}'] = "Latency"
        ws_summary[f'N{row}'] = "(Input) Tokens"
        ws_summary[f'O{row}'] = "(Output) Tokens"
        ws_summary[f'P{row}'] = "Gesamt Tokens"
        ws_summary[f'Q{row}'] = "ReRanking Tokens"
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        # NaN-Kategorien ausfiltern für Excel
        excel_categories = [c for c in results_df['category'].unique() if pd.notna(c)]
        for category in sorted(excel_categories):
            cat_df = results_df[results_df['category'] == category]
            ws_summary[f'A{row}'] = category
            
            for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'bert_f1', 'bert_precision', 'bert_recall', 'RR_at5', 'hit_at5', 'latency', 'prompt_tokens', 'completion_tokens', 'total_tokens', 'reranking_tokens'], 1):
                if metric in cat_df.columns:
                    avg = cat_df[metric].mean()
                    col_letter = chr(65 + idx)  # B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q
                    ws_summary[f'{col_letter}{row}'] = avg
                    if metric in ['prompt_tokens', 'completion_tokens', 'total_tokens', 'reranking_tokens']:
                        ws_summary[f'{col_letter}{row}'].number_format = '#,##0'
                    else:
                        ws_summary[f'{col_letter}{row}'].number_format = '0.000'
            
            row += 1
        
        # Nach Schwierigkeit
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Schwierigkeit"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header (mit allen 3 BERT-Metriken und Tokens)
        ws_summary[f'A{row}'] = "Schwierigkeit"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        ws_summary[f'E{row}'] = "Semantic Similarity"
        ws_summary[f'F{row}'] = "Context Entity Recall"
        ws_summary[f'G{row}'] = "Answer Relevancy"
        ws_summary[f'H{row}'] = "BERT-F1"
        ws_summary[f'I{row}'] = "BERT-Precision"
        ws_summary[f'J{row}'] = "BERT-Recall"
        ws_summary[f'K{row}'] = "MRR@5"
        ws_summary[f'L{row}'] = "Hit@5"
        ws_summary[f'M{row}'] = "Latency"
        ws_summary[f'N{row}'] = "(Input) Tokens"
        ws_summary[f'O{row}'] = "(Output) Tokens"
        ws_summary[f'P{row}'] = "Gesamt Tokens"
        ws_summary[f'Q{row}'] = "ReRanking Tokens"
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        for difficulty in ['easy', 'medium', 'hard']:
            diff_df = results_df[results_df['difficulty'] == difficulty]
            if len(diff_df) > 0:
                ws_summary[f'A{row}'] = difficulty.upper()
                
                for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision', 'semantic_similarity', 'context_entity_recall', 'answer_relevancy', 'bert_f1', 'bert_precision', 'bert_recall', 'RR_at5', 'hit_at5', 'latency', 'prompt_tokens', 'completion_tokens', 'total_tokens', 'reranking_tokens'], 1):
                    if metric in diff_df.columns:
                        avg = diff_df[metric].mean()
                        col_letter = chr(65 + idx)  # B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q
                        ws_summary[f'{col_letter}{row}'] = avg
                        if metric in ['prompt_tokens', 'completion_tokens', 'total_tokens', 'reranking_tokens']:
                            ws_summary[f'{col_letter}{row}'].number_format = '#,##0'
                        else:
                            ws_summary[f'{col_letter}{row}'].number_format = '0.000'
                
                row += 1
        
        # Spaltenbreiten für Zusammenfassung (erweitert um alle 3 BERT-Metriken und Tokens)
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 18
        ws_summary.column_dimensions['E'].width = 20
        ws_summary.column_dimensions['F'].width = 22
        ws_summary.column_dimensions['G'].width = 18
        ws_summary.column_dimensions['H'].width = 12  # BERT-F1
        ws_summary.column_dimensions['I'].width = 16  # BERT-Precision
        ws_summary.column_dimensions['J'].width = 14  # BERT-Recall
        ws_summary.column_dimensions['K'].width = 10  # MRR@5
        ws_summary.column_dimensions['L'].width = 10  # Hit@5
        ws_summary.column_dimensions['M'].width = 10  # Latency
        ws_summary.column_dimensions['N'].width = 16  # (Input) Tokens
        ws_summary.column_dimensions['O'].width = 17  # (Output) Tokens
        ws_summary.column_dimensions['P'].width = 14  # Gesamt Tokens
        ws_summary.column_dimensions['Q'].width = 18  # ReRanking Tokens
        
        # Speichern
        wb.save(output_path_excel)
        
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV:   {output_path_csv}")
        print(f"   Excel: {output_path_excel}")
        print("=" * 80 + "\n")
        
    except ImportError:
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV: {output_path_csv}")
        print(f"   ⚠️ Excel-Export nicht verfügbar (openpyxl nicht installiert)")
        print("=" * 80 + "\n")


def main():
    """Hauptfunktion"""
    global EVAL_TIMESTAMP
    
    print("\n" + "=" * 80)
    print(f"🎯 RAGAS-EVALUATION - WiSo-Chatbot ({len(EVAL_TIMESTAMPS)} Timestamps)")
    print("=" * 80 + "\n")
    
    # Zeige Evaluationsmodus
    if RUN_EVALUATION_LOCAL:
        print(f"⚙️ Evaluationsmodus: LOKAL (Ollama: {RAGAS_EVAL_MODEL})")
    else:
        print(f"⚙️ Evaluationsmodus: CLOUD (OpenAI: {OPENAI_EVAL_MODEL})")
    print(f"📁 Timestamps: {EVAL_TIMESTAMPS}")
    if EVAL_IDS:
        print(f"🎯 Selektive IDs: {EVAL_IDS}")
    else:
        print(f"🎯 IDs: Alle Fragen aus Testset")
    print()
    
    # Iteriere über alle Timestamps
    for ts_idx, timestamp in enumerate(EVAL_TIMESTAMPS, 1):
        EVAL_TIMESTAMP = timestamp
        
        print("\n" + "#" * 80)
        print(f"📋 [{ts_idx}/{len(EVAL_TIMESTAMPS)}] Timestamp: {EVAL_TIMESTAMP}")
        print("#" * 80)
        
        # Checkpoint-Pfad mit Timestamp (gleicher wie in generate_chatbot_responses)
        checkpoint_path = Path(__file__).parent / "data" / f"responses_checkpoint_{EVAL_TIMESTAMP}.pkl"
        
        # Variablen für Timing, URLs, Content-Types und Token-Usage initialisieren
        response_times = None
        urls_list = None
        content_types_list = None
        token_usage_list = None
        dataset = None
        checkpoint_complete = False
        
        try:
            # 1. LangSmith Client (immer initialisieren)
            print("🔗 Initialisiere LangSmith...")
            langsmith_client = Client(api_key=LANGSMITH_API_KEY)
            print(f"   ✅ Projekt: {LANGSMITH_PROJECT}\n")
            
            # 2. Testset laden (mit optionalem Limit und ID-Filter)
            print("📂 Lade Testset...")
            test_df = load_testset(limit=TEST_LIMIT, ids=EVAL_IDS)
            print()
            
            # 3. Prüfe ob Checkpoint existiert und vollständig ist
            if checkpoint_path.exists():
                print("📂 Prüfe Checkpoint...")
                import pickle
                with open(checkpoint_path, 'rb') as f:
                    checkpoint_data = pickle.load(f)
                
                # Checkpoint kann EvaluationDataset oder dict sein
                if isinstance(checkpoint_data, dict):
                    saved_dataset = checkpoint_data.get('dataset')
                    saved_df = checkpoint_data.get('test_df')
                    response_times = checkpoint_data.get('response_times', None)
                    urls_list = checkpoint_data.get('urls_list', None)
                    content_types_list = checkpoint_data.get('content_types_list', None)
                    token_usage_list = checkpoint_data.get('token_usage_list', None)
                    
                    # Prüfe ob Checkpoint vollständig ist
                    if saved_dataset and hasattr(saved_dataset, 'samples'):
                        # Ermittle IDs im Checkpoint
                        checkpoint_ids = set(saved_df['id'].tolist()) if saved_df is not None else set()
                        required_ids = set(test_df['id'].tolist())
                        
                        # Prüfe ob ALLE benötigten IDs im Checkpoint sind
                        if required_ids.issubset(checkpoint_ids):
                            # Alle IDs sind vorhanden → filtere auf die benötigten IDs
                            print(f"   ✅ Checkpoint vollständig für gefilterte IDs: {len(required_ids)} Antworten")
                            
                            # Filtere Dataset, response_times, urls, etc. auf die benötigten IDs
                            indices_to_keep = [i for i, row_id in enumerate(saved_df['id'].tolist()) if row_id in required_ids]
                            
                            filtered_samples = [s for i, s in enumerate(saved_dataset.samples) if i in indices_to_keep]
                            filtered_response_times = [response_times[i] for i in indices_to_keep] if response_times else None
                            filtered_urls_list = [urls_list[i] for i in indices_to_keep] if urls_list else None
                            filtered_content_types_list = [content_types_list[i] for i in indices_to_keep] if content_types_list else None
                            filtered_token_usage_list = [token_usage_list[i] for i in indices_to_keep] if token_usage_list else None
                            
                            dataset = EvaluationDataset(samples=filtered_samples)
                            response_times = filtered_response_times
                            urls_list = filtered_urls_list
                            content_types_list = filtered_content_types_list
                            token_usage_list = filtered_token_usage_list
                            checkpoint_complete = True
                        else:
                            # Nicht alle IDs vorhanden → Fortsetzung nötig
                            missing_ids = required_ids - checkpoint_ids
                            print(f"   ⏳ Checkpoint unvollständig: {len(checkpoint_ids)} von {len(required_ids)} Fragen")
                            print(f"   → Fehlende IDs: {sorted(missing_ids)}")
                            print(f"   → generate_chatbot_responses() wird fortsetzen\n")
                else:
                    # Alter Checkpoint-Format (nur Dataset)
                    dataset = checkpoint_data
                    if hasattr(dataset, 'samples'):
                        checkpoint_ids = getattr(dataset, '_saved_ids', set(range(len(dataset.samples))))
                        required_ids = set(test_df['id'].tolist())
                        if required_ids.issubset(checkpoint_ids):
                            checkpoint_complete = True
                            print(f"   ✅ Alter Checkpoint vollständig: {len(dataset.samples)} Antworten")
            
            # 4. Falls Checkpoint unvollständig oder nicht vorhanden → Antworten generieren
            if not checkpoint_complete:
                # Chatbot initialisieren
                print("🤖 Initialisiere Chatbot...")
                agent = create_react_agent()
                print()
                
                # Antworten generieren (setzt bei Checkpoint fort)
                dataset, response_times, urls_list, content_types_list, token_usage_list = generate_chatbot_responses(test_df, agent, langsmith_client)
                
                # ====================================================================
                # CHATBOT-MODELL ENTLADEN (GPU-Speicher freigeben vor RAGAS)
                # ====================================================================
                print("\n🧹 Räume GPU-Speicher auf (Chatbot + Embedding-Modell entladen)...")
                del agent  # Python-Referenz löschen
                gc.collect()  # Garbage Collection
                stop_ollama_model(OLLAMA_MODEL)  # LLM via CLI stoppen
                stop_embedding_model()  # Embedding-Modell (BGE-M3) freigeben
                
                # Bei lokaler Evaluation: Warte kurz damit GPU-Speicher freigegeben wird
                if RUN_EVALUATION_LOCAL:
                    print("   ⏳ Warte 2s für GPU-Speicherfreigabe...")
                    time.sleep(2)
                print()
            
            # 5. RAGAS-Evaluation (immer ausführen, Modus aus Config)
            results_df, evaluation_time = run_ragas_evaluation(dataset)
            
            # 6. Ergebnisse anzeigen und speichern (mit allen neuen Daten inkl. Token-Usage)
            display_and_save_results(results_df, test_df, response_times, urls_list, content_types_list, evaluation_time, token_usage_list)
            
            print(f"✅ Evaluation für {EVAL_TIMESTAMP} erfolgreich abgeschlossen!")
            
        except KeyboardInterrupt:
            print("\n\n⚠️ Evaluation abgebrochen!\n")
            sys.exit(1)
        except Exception as e:
            print(f"\n❌ Fehler bei {EVAL_TIMESTAMP}: {e}\n")
            import traceback
            traceback.print_exc()
            continue  # Fahre mit nächstem Timestamp fort


if __name__ == "__main__":
    main()
