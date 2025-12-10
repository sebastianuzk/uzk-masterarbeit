"""
Aktualisiere Excel-Dateien mit gold_doc_rank in Zusammenfassung
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def update_excel_with_gold_doc_rank(xlsx_path: Path):
    """
    Fügt gold_doc_rank zu einer Excel-Datei hinzu.
    """
    print(f"\n{'='*80}")
    print(f"Verarbeite: {xlsx_path.name}")
    print('='*80)
    
    # Lade Detaillierte Ergebnisse für gold_doc_rank Berechnung
    try:
        df_details = pd.read_excel(xlsx_path, sheet_name='Detaillierte Ergebnisse')
    except Exception as e:
        print(f"  ❌ Fehler beim Laden: {e}")
        return False
    
    # Prüfe ob gold_doc_rank existiert
    if 'gold_doc_rank' not in df_details.columns:
        print(f"  ⚠️  Keine 'gold_doc_rank' Spalte in Detaillierte Ergebnisse")
        return False
    
    # Berechne Durchschnitt
    avg_gold_doc_rank = df_details['gold_doc_rank'].mean()
    print(f"  📊 Durchschnitt gold_doc_rank: {avg_gold_doc_rank:.4f}")
    
    # Berechne pro Kategorie
    if 'category' in df_details.columns:
        cat_stats = df_details.groupby('category')['gold_doc_rank'].mean()
        print(f"  📊 Kategorien: {len(cat_stats)}")
    else:
        cat_stats = None
        print(f"  ⚠️  Keine 'category' Spalte")
    
    # Berechne pro Schwierigkeit
    if 'difficulty' in df_details.columns:
        diff_stats = df_details.groupby('difficulty')['gold_doc_rank'].mean()
        print(f"  📊 Schwierigkeiten: {len(diff_stats)}")
    else:
        diff_stats = None
        print(f"  ⚠️  Keine 'difficulty' Spalte")
    
    # Lade Workbook
    wb = load_workbook(xlsx_path)
    ws = wb['Zusammenfassung']
    
    # 1. Füge gold_doc_rank nach context_precision ein (Zeile 7, da 1-basiert)
    # Finde Zeile mit context_precision
    context_precision_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and 'context_precision' in str(cell_val):
            context_precision_row = row
            break
    
    if context_precision_row:
        # Füge neue Zeile ein
        ws.insert_rows(context_precision_row + 1)
        ws.cell(row=context_precision_row + 1, column=1, value='gold_doc_rank')
        ws.cell(row=context_precision_row + 1, column=2, value=avg_gold_doc_rank)
        print(f"  ✅ gold_doc_rank Durchschnitt in Zeile {context_precision_row + 1} eingefügt")
    else:
        print(f"  ⚠️  context_precision nicht gefunden")
    
    # 2. Füge Spalte für Kategorie-Tabelle hinzu
    kategorie_header_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and 'Kategorie' in str(cell_val) and ws.cell(row=row, column=2).value == 'Faithfulness':
            kategorie_header_row = row
            break
    
    if kategorie_header_row and cat_stats is not None:
        # Finde die Spalte nach Context Precision
        # Header ist: Kategorie | Faithfulness | Context Recall | Context Precision
        # Wir fügen Spalte E (5) hinzu: Gold Doc Rank
        header_col = 5
        ws.cell(row=kategorie_header_row, column=header_col, value='Gold Doc Rank')
        ws.cell(row=kategorie_header_row, column=header_col).font = Font(bold=True)
        
        # Füge Werte für jede Kategorie hinzu
        current_row = kategorie_header_row + 1
        while current_row <= ws.max_row:
            cat_name = ws.cell(row=current_row, column=1).value
            if cat_name is None or str(cat_name).strip() == '':
                break
            
            if cat_name in cat_stats.index:
                ws.cell(row=current_row, column=header_col, value=cat_stats[cat_name])
            current_row += 1
        
        print(f"  ✅ Gold Doc Rank für {current_row - kategorie_header_row - 1} Kategorien eingefügt")
    
    # 3. Füge Spalte für Schwierigkeit-Tabelle hinzu
    schwierigkeit_header_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and 'Schwierigkeit' in str(cell_val) and ws.cell(row=row, column=2).value == 'Faithfulness':
            schwierigkeit_header_row = row
            break
    
    if schwierigkeit_header_row and diff_stats is not None:
        header_col = 5
        ws.cell(row=schwierigkeit_header_row, column=header_col, value='Gold Doc Rank')
        ws.cell(row=schwierigkeit_header_row, column=header_col).font = Font(bold=True)
        
        # Füge Werte für jede Schwierigkeit hinzu
        current_row = schwierigkeit_header_row + 1
        while current_row <= ws.max_row:
            diff_name = ws.cell(row=current_row, column=1).value
            if diff_name is None or str(diff_name).strip() == '':
                break
            
            if diff_name in diff_stats.index:
                ws.cell(row=current_row, column=header_col, value=diff_stats[diff_name])
            current_row += 1
        
        print(f"  ✅ Gold Doc Rank für {current_row - schwierigkeit_header_row - 1} Schwierigkeiten eingefügt")
    
    # Speichere
    wb.save(xlsx_path)
    print(f"  💾 Gespeichert: {xlsx_path.name}")
    return True


# Hauptprogramm
if __name__ == "__main__":
    backup_dir = Path('src/evaluation/data/Backup')
    
    # Liste der zu verarbeitenden Dateien
    target_files = [
        'naive_1250_300.xlsx',
        'naive_1500_300.xlsx',
        'naive_1750_200.xlsx',
        'naive_1750_300.xlsx',
        'semantic_1750_400_200_07.xlsx',
        'semantic_1750_400_300_07.xlsx',
        'semantic_2000_500_300_08.xlsx',
        'semantic_70.xlsx',
    ]
    
    print("=" * 80)
    print("AKTUALISIERUNG: gold_doc_rank in Zusammenfassung-Blättern")
    print("=" * 80)
    
    success_count = 0
    for filename in target_files:
        xlsx_path = backup_dir / filename
        if xlsx_path.exists():
            if update_excel_with_gold_doc_rank(xlsx_path):
                success_count += 1
        else:
            print(f"\n⚠️  Datei nicht gefunden: {filename}")
    
    print("\n" + "=" * 80)
    print(f"ABGESCHLOSSEN: {success_count}/{len(target_files)} Dateien aktualisiert")
    print("=" * 80)
