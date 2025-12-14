"""Aktualisiert nur die Metrik-Werte in der Excel-Datei ohne Formatierung zu ändern."""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Pfade
csv_path = r'src/evaluation/data/ragas_results.csv'
excel_path = r'src/evaluation/data/ragas_results.xlsx'

# CSV laden
df = pd.read_csv(csv_path)

# Bestehende Excel öffnen
wb = load_workbook(excel_path)
ws = wb.active

# Spaltennamen aus Header lesen (Zeile 1)
header = {cell.value: cell.column for cell in ws[1]}

# Metriken-Spalten identifizieren (jetzt mit gold_doc_rank)
metric_cols = ['faithfulness', 'context_recall', 'context_precision', 'gold_doc_rank']

# Prüfe ob gold_doc_rank Spalte existiert, wenn nicht hinzufügen
if 'gold_doc_rank' not in header:
    # Finde die Position nach context_precision
    if 'context_precision' in header:
        new_col = header['context_precision'] + 1
    else:
        new_col = ws.max_column + 1
    
    # Spalte einfügen
    ws.insert_cols(new_col)
    ws.cell(row=1, column=new_col).value = 'gold_doc_rank'
    
    # Header neu lesen
    header = {cell.value: cell.column for cell in ws[1]}
    print(f'📊 Spalte gold_doc_rank an Position {new_col} hinzugefügt')

# Für jede Zeile in der CSV die Werte aktualisieren
updated_count = 0
for _, row in df.iterrows():
    # ID als int konvertieren für Vergleich
    try:
        row_id = int(row['id'])
    except (ValueError, TypeError):
        row_id = row['id']
    
    # Finde die passende Zeile in Excel (nach ID)
    for excel_row in range(2, ws.max_row + 1):
        id_col = header.get('id', 1)
        excel_id = ws.cell(row=excel_row, column=id_col).value
        # Vergleiche als int falls möglich
        try:
            excel_id_int = int(excel_id) if excel_id is not None else None
        except (ValueError, TypeError):
            excel_id_int = excel_id
            
        if excel_id_int == row_id:
            # Aktualisiere nur die Metrik-Werte
            for metric in metric_cols:
                if metric in header and metric in row:
                    col = header[metric]
                    new_value = row[metric]
                    ws.cell(row=excel_row, column=col).value = new_value
            updated_count += 1
            break

# Speichern
wb.save(excel_path)
wb.close()

print(f'✅ Excel aktualisiert: {excel_path}')
print(f'   {updated_count} Zeilen aktualisiert (nur Metrik-Werte)')

# Durchschnitte anzeigen
print(f'\n📊 Durchschnittswerte:')
print(f'   Faithfulness: {df["faithfulness"].mean():.3f}')
print(f'   Context Recall: {df["context_recall"].mean():.3f}')
print(f'   Context Precision: {df["context_precision"].mean():.3f}')
