"""
BERT-Score Evaluation für WiSo-Chatbot

Berechnet ausschließlich BERT-Scores (F1, Precision, Recall) basierend auf Checkpoint-PKL-Dateien.
Verwendet die gleiche Logik wie ragas_evaluation.py, aber ohne RAGAS-Metriken.

Verwendung:
    python src/evaluation/update_excel_trash.py <pfad_zur_pkl_datei>
"""

import sys
import warnings
import pickle
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import time

# BERT-Score für Token-Level semantische Ähnlichkeit
try:
    from bert_score import score as bert_score_fn
    BERT_SCORE_AVAILABLE = True
except ImportError:
    BERT_SCORE_AVAILABLE = False
    print("⚠️ bert-score nicht installiert. Installiere mit: pip install bert-score")

# Projekt-Root
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from config.settings import RANDOM_SEED, OLLAMA_MODEL, TEMPERATURE

# Setze Seeds für Reproduzierbarkeit
import random
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

# Timestamp für Output-Dateien
EVAL_TIMESTAMP = "validation"


def load_checkpoint(pkl_path: Path):
    """
    Lädt Checkpoint-PKL.
    Exakt gleiche Logik wie in ragas_evaluation.py main().
    """
    print("📂 Prüfe Checkpoint...")
    import pickle
    with open(pkl_path, 'rb') as f:
        checkpoint_data = pickle.load(f)
    
    # Checkpoint kann EvaluationDataset oder dict sein
    if isinstance(checkpoint_data, dict):
        saved_dataset = checkpoint_data.get('dataset')
        saved_df = checkpoint_data.get('test_df')
        response_times = checkpoint_data.get('response_times', None)
        urls_list = checkpoint_data.get('urls_list', None)
        content_types_list = checkpoint_data.get('content_types_list', None)
        token_usage_list = checkpoint_data.get('token_usage_list', None)
        
        # Prüfe ob Checkpoint vollständig ist
        if saved_dataset and hasattr(saved_dataset, 'samples'):
            num_saved = len(saved_dataset.samples)
            print(f"   ✅ Checkpoint vollständig: {num_saved} Antworten")
            if token_usage_list:
                total_tokens = sum(t.get('total_tokens', 0) for t in token_usage_list)
                print(f"   📊 Gesamt Tokens: {total_tokens}")
            dataset = saved_dataset
        else:
            raise ValueError("Checkpoint enthält kein gültiges Dataset!")
    else:
        # Alter Checkpoint-Format (nur Dataset)
        dataset = checkpoint_data
        if hasattr(dataset, 'samples'):
            print(f"   ✅ Alter Checkpoint vollständig: {len(dataset.samples)} Antworten")
        else:
            raise ValueError("Checkpoint enthält kein gültiges Dataset!")
        saved_df = None
        token_usage_list = None
    
    return dataset, saved_df, token_usage_list


def run_bert_evaluation(dataset, test_df, token_usage_list=None) -> pd.DataFrame:
    """
    Berechnet BERT-Scores für alle Samples.
    Exakt gleiche Logik wie in ragas_evaluation.py run_ragas_evaluation().
    """
    # Erstelle results_df aus dataset samples
    data = []
    for i, sample in enumerate(dataset.samples):
        row = {
            'user_input': sample.user_input,
            'response': sample.response,
            'reference': sample.reference,
            'retrieved_contexts': sample.retrieved_contexts,
        }
        # ID, category, difficulty aus test_df falls vorhanden
        if test_df is not None and i < len(test_df):
            row['id'] = test_df.iloc[i]['id'] if 'id' in test_df.columns else i+1
            row['category'] = test_df.iloc[i]['category'] if 'category' in test_df.columns else ''
            row['difficulty'] = test_df.iloc[i]['difficulty'] if 'difficulty' in test_df.columns else ''
        else:
            row['id'] = i + 1
            row['category'] = ''
            row['difficulty'] = ''
        
        # Token-Usage hinzufügen falls vorhanden
        if token_usage_list and i < len(token_usage_list):
            row['prompt_tokens'] = token_usage_list[i].get('prompt_tokens', 0)
            row['completion_tokens'] = token_usage_list[i].get('completion_tokens', 0)
            row['total_tokens'] = token_usage_list[i].get('total_tokens', 0)
        else:
            row['prompt_tokens'] = 0
            row['completion_tokens'] = 0
            row['total_tokens'] = 0
            
        data.append(row)
    results_df = pd.DataFrame(data)
    
    if BERT_SCORE_AVAILABLE:
        print("📊 Berechne BERT-Score...")
        bert_start = time.time()
        
        try:
            # Extrahiere responses und references aus dataset
            responses = [s.response for s in dataset.samples]
            references = [s.reference for s in dataset.samples]
            
            # BERT-Score berechnen (Multilingual: xlm-roberta-large)
            # Unterstützt Deutsch + Englisch gemischt
            # Unterdrücke Tokenizer-Warnungen
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore")
                P, R, F1 = bert_score_fn(
                    responses, 
                    references, 
                    model_type="xlm-roberta-large",  # Multilingual (100+ Sprachen, inkl. DE + EN)
                    lang="de",  # Sprache für Baseline-Rescaling (DE funktioniert auch für gemischte Texte)
                    verbose=False,
                    rescale_with_baseline=False  # Bessere Interpretierbarkeit
                )
            
            # Zu results_df hinzufügen
            results_df['bert_precision'] = P.numpy()
            results_df['bert_recall'] = R.numpy()
            results_df['bert_f1'] = F1.numpy()
            
            bert_time = time.time() - bert_start
            print(f"   ✅ BERT-Score berechnet in {bert_time:.2f}s")
            print(f"   📈 Durchschn. BERT-F1: {F1.mean():.3f}")
            
        except Exception as e:
            print(f"   ⚠️ BERT-Score Fehler: {e}")
            results_df['bert_precision'] = None
            results_df['bert_recall'] = None
            results_df['bert_f1'] = None
    else:
        print("⚠️ BERT-Score übersprungen (nicht installiert)")
        results_df['bert_precision'] = None
        results_df['bert_recall'] = None
        results_df['bert_f1'] = None
    
    return results_df


def display_and_save_results(results_df: pd.DataFrame, output_name: str = None):
    """
    Zeigt Ergebnisse an und speichert sie.
    Gleiche Logik wie in ragas_evaluation.py, aber nur für BERT-Scores.
    """
    if output_name is None:
        output_name = f"bert_scores_{EVAL_TIMESTAMP}"
    
    output_path_csv = Path(__file__).parent / "data" / f"{output_name}.csv"
    output_path_excel = Path(__file__).parent / "data" / f"{output_name}.xlsx"
    
    print("\n" + "=" * 80)
    print("📊 BERT-SCORE ERGEBNISSE")
    print("=" * 80)
    
    # BERT-Metriken und Token-Metriken
    bert_metrics = ['bert_f1', 'bert_precision', 'bert_recall']
    token_metrics = ['prompt_tokens', 'completion_tokens', 'total_tokens']
    
    # Gesamtscores
    print("\n📈 Durchschnittliche BERT-Scores:")
    print("-" * 80)
    for metric in bert_metrics:
        if metric in results_df.columns and results_df[metric].notna().any():
            avg = results_df[metric].mean()
            display_name = {'bert_f1': 'BERT-F1', 'bert_precision': 'BERT-Precision', 'bert_recall': 'BERT-Recall'}.get(metric, metric)
            print(f"   {display_name:20s}: {avg:.3f}")
    
    # Token-Statistik
    if 'total_tokens' in results_df.columns and results_df['total_tokens'].sum() > 0:
        print("\n📊 Token-Statistik:")
        print("-" * 80)
        # Gesamt
        total = results_df['total_tokens'].sum()
        total_prompt = results_df['prompt_tokens'].sum()
        total_completion = results_df['completion_tokens'].sum()
        print(f"   {'Gesamt Tokens':20s}: {total:,.0f}")
        print(f"   {'  - Input (Prompt)':20s}: {total_prompt:,.0f}")
        print(f"   {'  - Output (Compl.)':20s}: {total_completion:,.0f}")
        # Durchschnitt
        avg = results_df['total_tokens'].mean()
        avg_prompt = results_df['prompt_tokens'].mean()
        avg_completion = results_df['completion_tokens'].mean()
        print(f"   {'Durchschn. Tokens':20s}: {avg:,.0f}")
        print(f"   {'  - Input (Prompt)':20s}: {avg_prompt:,.0f}")
        print(f"   {'  - Output (Compl.)':20s}: {avg_completion:,.0f}")
    
    # Nach Kategorie
    if 'category' in results_df.columns:
        print("\n📁 BERT-Scores nach Kategorie:")
        print("-" * 80)
        categories = [c for c in results_df['category'].unique() if pd.notna(c)]
        for category in sorted(categories):
            cat_df = results_df[results_df['category'] == category]
            print(f"\n   {category}:")
            for metric in bert_metrics:
                if metric in cat_df.columns and cat_df[metric].notna().any():
                    avg = cat_df[metric].mean()
                    display_name = {'bert_f1': 'BERT-F1', 'bert_precision': 'BERT-Precision', 'bert_recall': 'BERT-Recall'}.get(metric, metric)
                    print(f"      {display_name:20s}: {avg:.3f}")
            # Token-Statistik pro Kategorie
            if 'total_tokens' in cat_df.columns and cat_df['total_tokens'].sum() > 0:
                print(f"      {'Tokens (Input)':20s}: {cat_df['prompt_tokens'].mean():,.0f}")
                print(f"      {'Tokens (Output)':20s}: {cat_df['completion_tokens'].mean():,.0f}")
                print(f"      {'Tokens (Gesamt)':20s}: {cat_df['total_tokens'].mean():,.0f}")
    
    # Nach Schwierigkeit
    if 'difficulty' in results_df.columns:
        print("\n⚡ BERT-Scores nach Schwierigkeit:")
        print("-" * 80)
        for difficulty in ['easy', 'medium', 'hard']:
            diff_df = results_df[results_df['difficulty'] == difficulty]
            if len(diff_df) > 0:
                print(f"\n   {difficulty.upper()}:")
                for metric in bert_metrics:
                    if metric in diff_df.columns and diff_df[metric].notna().any():
                        avg = diff_df[metric].mean()
                        display_name = {'bert_f1': 'BERT-F1', 'bert_precision': 'BERT-Precision', 'bert_recall': 'BERT-Recall'}.get(metric, metric)
                        print(f"      {display_name:20s}: {avg:.3f}")
                # Token-Statistik pro Schwierigkeit
                if 'total_tokens' in diff_df.columns and diff_df['total_tokens'].sum() > 0:
                    print(f"      {'Tokens (Input)':20s}: {diff_df['prompt_tokens'].mean():,.0f}")
                    print(f"      {'Tokens (Output)':20s}: {diff_df['completion_tokens'].mean():,.0f}")
                    print(f"      {'Tokens (Gesamt)':20s}: {diff_df['total_tokens'].mean():,.0f}")
    
    # CSV speichern
    # Entferne Zeilenumbrüche aus Textfeldern
    text_columns = ['user_input', 'response', 'reference']
    for col in text_columns:
        if col in results_df.columns:
            results_df[col] = results_df[col].apply(lambda x: x.replace('\n', ' ').replace('\r', ' ') if isinstance(x, str) else x)
    
    # CSV-Spalten (erweitert um Token-Spalten)
    csv_columns = ['id', 'category', 'difficulty', 'user_input', 'response', 'reference',
                   'bert_f1', 'bert_precision', 'bert_recall', 
                   'prompt_tokens', 'completion_tokens', 'total_tokens']
    csv_columns = [col for col in csv_columns if col in results_df.columns]
    csv_df = results_df[csv_columns].copy()
    
    # ============================================================================
    # METADATEN-ZEILEN: Durchschnitte
    # ============================================================================
    metadata_rows = []
    
    # Alle Metriken für Durchschnitte (BERT + Tokens)
    all_metrics = bert_metrics + token_metrics
    
    # Gesamtdurchschnitt
    avg_row = {col: '' for col in csv_columns}
    avg_row['id'] = 'AVG'
    avg_row['category'] = 'GESAMT'
    avg_row['difficulty'] = f'n={len(results_df)}'
    for metric in all_metrics:
        if metric in results_df.columns:
            avg_row[metric] = results_df[metric].mean()
    metadata_rows.append(avg_row)
    
    # Durchschnitte pro Kategorie
    if 'category' in results_df.columns:
        categories = [c for c in results_df['category'].unique() if pd.notna(c)]
        for category in sorted(categories):
            cat_df = results_df[results_df['category'] == category]
            cat_row = {col: '' for col in csv_columns}
            cat_row['id'] = 'AVG'
            cat_row['category'] = category
            cat_row['difficulty'] = f'n={len(cat_df)}'
            for metric in all_metrics:
                if metric in cat_df.columns:
                    cat_row[metric] = cat_df[metric].mean()
            metadata_rows.append(cat_row)
    
    # Durchschnitte pro Schwierigkeit
    if 'difficulty' in results_df.columns:
        for difficulty in ['easy', 'medium', 'hard']:
            diff_df = results_df[results_df['difficulty'] == difficulty]
            if len(diff_df) > 0:
                diff_row = {col: '' for col in csv_columns}
                diff_row['id'] = 'AVG'
                diff_row['category'] = f'Schwierigkeit: {difficulty.upper()}'
                diff_row['difficulty'] = f'n={len(diff_df)}'
                for metric in all_metrics:
                    if metric in diff_df.columns:
                        diff_row[metric] = diff_df[metric].mean()
                metadata_rows.append(diff_row)
    
    # Durchschnitte pro Kategorie + Schwierigkeit (kombiniert)
    if 'category' in results_df.columns and 'difficulty' in results_df.columns:
        categories = [c for c in results_df['category'].unique() if pd.notna(c)]
        for category in sorted(categories):
            for difficulty in ['easy', 'medium', 'hard']:
                combo_df = results_df[(results_df['category'] == category) & (results_df['difficulty'] == difficulty)]
                if len(combo_df) > 0:
                    combo_row = {col: '' for col in csv_columns}
                    combo_row['id'] = 'AVG'
                    combo_row['category'] = f'{category} / {difficulty.upper()}'
                    combo_row['difficulty'] = f'n={len(combo_df)}'
                    for metric in all_metrics:
                        if metric in combo_df.columns:
                            combo_row[metric] = combo_df[metric].mean()
                    metadata_rows.append(combo_row)
    
    # Metadaten-DataFrame erstellen und anhängen
    meta_df = pd.DataFrame(metadata_rows)
    csv_df = pd.concat([csv_df, meta_df], ignore_index=True)
    
    # Speichere mit UTF-8-BOM
    csv_df.to_csv(output_path_csv, index=False, encoding='utf-8-sig', sep=',', quoting=1)
    
    # ============================================================================
    # EXCEL MIT FORMATIERUNG
    # ============================================================================
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Sheet 1: Detaillierte Ergebnisse
        ws_details = wb.active
        ws_details.title = "Detaillierte Ergebnisse"
        
        # Header-Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Daten schreiben - NUR results_df (ohne AVG-Zeilen)
        excel_details_df = results_df[csv_columns].copy()
        for r_idx, row in enumerate(dataframe_to_rows(excel_details_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatieren
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # BERT-Score Spalten farbig
                    col_name = csv_columns[c_idx - 1] if c_idx <= len(csv_columns) else None
                    if col_name in ['bert_f1', 'bert_precision', 'bert_recall']:
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            if value >= 0.8:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value >= 0.6:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.number_format = '0.000'
                    
                    # Token-Spalten als Ganzzahlen formatieren
                    if col_name in ['prompt_tokens', 'completion_tokens', 'total_tokens']:
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '#,##0'
                    
                    # Text-Wrap für lange Texte
                    if col_name in ['user_input', 'response', 'reference']:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Spaltenbreiten anpassen
        ws_details.column_dimensions['A'].width = 8   # id
        ws_details.column_dimensions['B'].width = 20  # category
        ws_details.column_dimensions['C'].width = 12  # difficulty
        ws_details.column_dimensions['D'].width = 50  # user_input
        ws_details.column_dimensions['E'].width = 60  # response
        ws_details.column_dimensions['F'].width = 50  # reference
        ws_details.column_dimensions['G'].width = 12  # bert_f1
        ws_details.column_dimensions['H'].width = 15  # bert_precision
        ws_details.column_dimensions['I'].width = 13  # bert_recall
        ws_details.column_dimensions['J'].width = 14  # prompt_tokens
        ws_details.column_dimensions['K'].width = 16  # completion_tokens
        ws_details.column_dimensions['L'].width = 12  # total_tokens
        
        # Sheet 2: Zusammenfassung
        ws_summary = wb.create_sheet("Zusammenfassung")
        
        # Titel
        ws_summary['A1'] = "📊 BERT-Score Evaluation Zusammenfassung"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')
        
        # Durchschnittliche Scores
        row = 3
        ws_summary[f'A{row}'] = "Durchschnittliche BERT-Scores"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for metric in bert_metrics:
            if metric in results_df.columns and results_df[metric].notna().any():
                avg = results_df[metric].mean()
                display_name = {'bert_f1': 'BERT-F1', 'bert_precision': 'BERT-Precision', 'bert_recall': 'BERT-Recall'}.get(metric, metric)
                ws_summary[f'A{row}'] = display_name
                ws_summary[f'B{row}'] = avg
                ws_summary[f'B{row}'].number_format = '0.000'
                
                # Farbcodierung
                if avg >= 0.8:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif avg >= 0.6:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                row += 1
        
        # Token-Statistik in Zusammenfassung
        if 'total_tokens' in results_df.columns and results_df['total_tokens'].sum() > 0:
            row += 1
            ws_summary[f'A{row}'] = "Token-Statistik"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            # Gesamt
            ws_summary[f'A{row}'] = "Gesamt Tokens"
            ws_summary[f'B{row}'] = results_df['total_tokens'].sum()
            ws_summary[f'B{row}'].number_format = '#,##0'
            row += 1
            ws_summary[f'A{row}'] = "  - Input (Prompt)"
            ws_summary[f'B{row}'] = results_df['prompt_tokens'].sum()
            ws_summary[f'B{row}'].number_format = '#,##0'
            row += 1
            ws_summary[f'A{row}'] = "  - Output (Completion)"
            ws_summary[f'B{row}'] = results_df['completion_tokens'].sum()
            ws_summary[f'B{row}'].number_format = '#,##0'
            row += 1
            # Durchschnitt
            ws_summary[f'A{row}'] = "Durchschn. Tokens"
            ws_summary[f'B{row}'] = results_df['total_tokens'].mean()
            ws_summary[f'B{row}'].number_format = '#,##0'
            row += 1
            ws_summary[f'A{row}'] = "  - Input (Prompt)"
            ws_summary[f'B{row}'] = results_df['prompt_tokens'].mean()
            ws_summary[f'B{row}'].number_format = '#,##0'
            row += 1
            ws_summary[f'A{row}'] = "  - Output (Completion)"
            ws_summary[f'B{row}'] = results_df['completion_tokens'].mean()
            ws_summary[f'B{row}'].number_format = '#,##0'
            row += 1
        
        # Nach Kategorie
        if 'category' in results_df.columns:
            row += 2
            ws_summary[f'A{row}'] = "BERT-Scores nach Kategorie"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Header (mit Token-Spalten für Input/Output/Gesamt)
            ws_summary[f'A{row}'] = "Kategorie"
            ws_summary[f'B{row}'] = "BERT-F1"
            ws_summary[f'C{row}'] = "BERT-Precision"
            ws_summary[f'D{row}'] = "BERT-Recall"
            ws_summary[f'E{row}'] = "Tokens (Input)"
            ws_summary[f'F{row}'] = "Tokens (Output)"
            ws_summary[f'G{row}'] = "Tokens (Gesamt)"
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws_summary[f'{col}{row}'].font = Font(bold=True)
                ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            row += 1
            
            categories = [c for c in results_df['category'].unique() if pd.notna(c)]
            for category in sorted(categories):
                cat_df = results_df[results_df['category'] == category]
                ws_summary[f'A{row}'] = category
                
                for idx, metric in enumerate(['bert_f1', 'bert_precision', 'bert_recall'], 1):
                    if metric in cat_df.columns:
                        avg = cat_df[metric].mean()
                        col_letter = chr(65 + idx)  # B, C, D
                        ws_summary[f'{col_letter}{row}'] = avg
                        ws_summary[f'{col_letter}{row}'].number_format = '0.000'
                # Token-Durchschnitte (Input, Output, Total)
                if 'prompt_tokens' in cat_df.columns:
                    ws_summary[f'E{row}'] = cat_df['prompt_tokens'].mean()
                    ws_summary[f'E{row}'].number_format = '#,##0'
                if 'completion_tokens' in cat_df.columns:
                    ws_summary[f'F{row}'] = cat_df['completion_tokens'].mean()
                    ws_summary[f'F{row}'].number_format = '#,##0'
                if 'total_tokens' in cat_df.columns:
                    ws_summary[f'G{row}'] = cat_df['total_tokens'].mean()
                    ws_summary[f'G{row}'].number_format = '#,##0'
                row += 1
        
        # Nach Schwierigkeit
        if 'difficulty' in results_df.columns:
            row += 2
            ws_summary[f'A{row}'] = "BERT-Scores nach Schwierigkeit"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Header (mit Token-Spalten für Input/Output/Gesamt)
            ws_summary[f'A{row}'] = "Schwierigkeit"
            ws_summary[f'B{row}'] = "BERT-F1"
            ws_summary[f'C{row}'] = "BERT-Precision"
            ws_summary[f'D{row}'] = "BERT-Recall"
            ws_summary[f'E{row}'] = "Tokens (Input)"
            ws_summary[f'F{row}'] = "Tokens (Output)"
            ws_summary[f'G{row}'] = "Tokens (Gesamt)"
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws_summary[f'{col}{row}'].font = Font(bold=True)
                ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            row += 1
            
            for difficulty in ['easy', 'medium', 'hard']:
                diff_df = results_df[results_df['difficulty'] == difficulty]
                if len(diff_df) > 0:
                    ws_summary[f'A{row}'] = difficulty.upper()
                    
                    for idx, metric in enumerate(['bert_f1', 'bert_precision', 'bert_recall'], 1):
                        if metric in diff_df.columns:
                            avg = diff_df[metric].mean()
                            col_letter = chr(65 + idx)  # B, C, D
                            ws_summary[f'{col_letter}{row}'] = avg
                            ws_summary[f'{col_letter}{row}'].number_format = '0.000'
                    # Token-Durchschnitte (Input, Output, Total)
                    if 'prompt_tokens' in diff_df.columns:
                        ws_summary[f'E{row}'] = diff_df['prompt_tokens'].mean()
                        ws_summary[f'E{row}'].number_format = '#,##0'
                    if 'completion_tokens' in diff_df.columns:
                        ws_summary[f'F{row}'] = diff_df['completion_tokens'].mean()
                        ws_summary[f'F{row}'].number_format = '#,##0'
                    if 'total_tokens' in diff_df.columns:
                        ws_summary[f'G{row}'] = diff_df['total_tokens'].mean()
                        ws_summary[f'G{row}'].number_format = '#,##0'
                    row += 1
        
        # Spaltenbreiten für Zusammenfassung
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 15
        ws_summary.column_dimensions['E'].width = 14  # Input Tokens
        ws_summary.column_dimensions['F'].width = 14  # Output Tokens
        ws_summary.column_dimensions['G'].width = 14  # Total Tokens
        
        # Speichern
        wb.save(output_path_excel)
        
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV:   {output_path_csv}")
        print(f"   Excel: {output_path_excel}")
        print("=" * 80 + "\n")
        
    except ImportError:
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV: {output_path_csv}")
        print(f"   ⚠️ Excel-Export nicht verfügbar (openpyxl nicht installiert)")
        print("=" * 80 + "\n")


def main():
    """Hauptfunktion"""
    
    print("\n" + "=" * 80)
    print("🎯 BERT-SCORE EVALUATION - WiSo-Chatbot")
    print("=" * 80 + "\n")
    
    if not BERT_SCORE_AVAILABLE:
        print("❌ bert-score ist nicht installiert!")
        print("   Installiere mit: pip install bert-score")
        sys.exit(1)
    
    print(f"📁 Timestamp: {EVAL_TIMESTAMP}")
    print()
    
    # Checkpoint-Pfad mit Timestamp (gleicher wie in generate_chatbot_responses)
    checkpoint_path = Path(__file__).parent / "data" / f"responses_checkpoint_{EVAL_TIMESTAMP}.pkl"
    
    # PKL-Pfad aus Argumenten oder Standard-Checkpoint
    if len(sys.argv) >= 2:
        pkl_path = Path(sys.argv[1])
    else:
        pkl_path = checkpoint_path
    
    if not pkl_path.exists():
        print(f"❌ Datei nicht gefunden: {pkl_path}")
        sys.exit(1)
    
    # Output-Name basierend auf PKL-Name
    output_name = f"bert_{pkl_path.stem}"
    
    print(f"📁 PKL-Datei: {pkl_path}")
    print(f"📁 Output: {output_name}")
    print()
    
    # 1. Lade Checkpoint (inkl. Token-Usage falls vorhanden)
    dataset, test_df, token_usage_list = load_checkpoint(pkl_path)
    
    # 2. Berechne BERT-Scores (mit Token-Usage)
    results_df = run_bert_evaluation(dataset, test_df, token_usage_list)
    
    # 3. Speichere Ergebnisse
    display_and_save_results(results_df, output_name)
    
    print("✅ BERT-Score Evaluation abgeschlossen!")


if __name__ == "__main__":
    main()
