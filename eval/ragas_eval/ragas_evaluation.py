"""
RAGAS-Evaluation für WiSo-Chatbot

Evaluiert den Chatbot mit RAGAS-Framework:
- Verwendet Ollama (qwen3:8b) als LLM-Judge
- Lädt Testfragen aus Testset.CSV
- Generiert Antworten mit dem Chatbot
- Extrahiert RAG-Kontexte aus LangSmith
- Berechnet RAGAS-Metriken (Faithfulness, Context Recall)
"""

import sys
import os
import pickle
import tempfile
import pandas as pd
from pathlib import Path
from typing import List, Optional
import time

# Import RAGAS library FIRST (before adding project_root to avoid shadowing)
from ragas import evaluate
from ragas.metrics import faithfulness, context_recall, context_precision  # answer_relevancy benötigt Embeddings
from ragas.dataset_schema import SingleTurnSample, EvaluationDataset
from ragas.run_config import RunConfig

# Projekt-Root (add AFTER RAGAS imports)
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))
from langchain_ollama import ChatOllama, OllamaEmbeddings
from langsmith import Client
from config.settings import (
    OLLAMA_MODEL,
    OLLAMA_BASE_URL,
    OLLAMA_EMBEDDING_MODEL,
    LANGSMITH_API_KEY,
    LANGSMITH_PROJECT,
    RAGAS_JUDGE_MODEL,
    TEMPERATURE,
    settings
)
from src.agent.react_agent import create_react_agent


def _save_checkpoint_atomic(checkpoint_path: Path, checkpoint_data: dict) -> None:
    """Persist checkpoint atomically.

    Writes to a temp file in the same directory, fsyncs, then os.replace()s
    onto the target. Prevents partial/corrupted .pkl files if the process is
    killed mid-write (Ctrl+C, OOM, container stop).
    """
    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(
        prefix=checkpoint_path.name + ".",
        suffix=".tmp",
        dir=str(checkpoint_path.parent),
    )
    try:
        with os.fdopen(fd, "wb") as f:
            pickle.dump(checkpoint_data, f)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, checkpoint_path)
    except Exception:
        # Best-effort cleanup of the temp file on failure
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _find_run_by_session_id(
    langsmith_client: Client,
    session_id: str,
    max_attempts: int = 6,
    initial_wait: float = 1.0,
    list_limit: int = 30,
) -> Optional[object]:
    """Poll LangSmith for a root run whose metadata.session_id matches.

    LangSmith trace ingestion is eventually consistent, so a single list_runs()
    immediately after agent.chat() may miss the run. We poll with exponential
    backoff (1s, 2s, 4s, ...) up to ``max_attempts`` times. Returns the matching
    run, or None if no match was found within the budget.

    The caller is responsible for any fallback behavior; this function does NOT
    fall back to "some other recent run" because that would attribute contexts
    to the wrong question and silently corrupt RAGAS metrics.
    """
    wait = initial_wait
    for attempt in range(1, max_attempts + 1):
        time.sleep(wait)
        try:
            recent_runs = list(langsmith_client.list_runs(
                project_name=LANGSMITH_PROJECT,
                is_root=True,
                limit=list_limit,
            ))
        except Exception as e:
            print(f"   ⚠️ LangSmith list_runs failed (attempt {attempt}/{max_attempts}): {e}")
            wait = min(wait * 2, 16.0)
            continue

        for run in recent_runs:
            run_name = str(run.name).lower() if run.name else ""
            if "self_reflection" in run_name or "chatollama" in run_name:
                continue
            run_metadata = run.extra.get("metadata", {}) if run.extra else {}
            if run_metadata.get("session_id") == session_id:
                return run

        # No match yet — back off and retry
        wait = min(wait * 2, 16.0)

    return None


def load_testset(csv_path: str = "data/Testset.CSV", limit: int = None) -> pd.DataFrame:
    """Lädt Testset.CSV"""
    full_path = Path(__file__).parent / csv_path
    df = pd.read_csv(full_path, sep=';', encoding='utf-8')
    
    if limit:
        df = df.head(limit)
    
    print(f"✅ {len(df)} Testfragen geladen")
    
    return df


def get_rag_context_from_langsmith(client: Client, trace_id: str, debug: bool = False) -> List[str]:
    """
    Holt RAG-Kontext aus LangSmith für eine spezifische Trace-ID.
    
    Die Documents befinden sich im Retriever-Output unter dem Key 'output' (nicht 'documents'!).
    Jedes Document hat 'page_content' und 'metadata'.
    
    Args:
        client: LangSmith Client
        trace_id: Die Trace-ID der Session
        debug: Wenn True, gebe Debug-Informationen aus
        
    Returns:
        Liste von RAG-Context-Chunks aus den Retriever-Documents
    """
    try:
        # Hole alle Child-Runs für diese Trace
        child_runs = list(client.list_runs(
            project_name=LANGSMITH_PROJECT,
            trace_id=trace_id,
            is_root=False
        ))
        
        if debug:
            print(f"      🔍 DEBUG: {len(child_runs)} child runs gefunden")
            for i, child in enumerate(child_runs[:10]):  # Nur erste 10 zeigen
                print(f"         [{i}] Type: {child.run_type}, Name: {child.name}")
        
        # Suche nach Retriever-Run oder Tool-Run mit RAG
        contexts = []
        for child in child_runs:
            # Prüfe verschiedene Möglichkeiten
            if child.run_type == "retriever":
                if child.outputs and isinstance(child.outputs, dict):
                    # Documents sind unter 'output' Key (nicht 'documents')!
                    documents = child.outputs.get('output', [])
                    for doc in documents:
                        if isinstance(doc, dict) and 'page_content' in doc:
                            contexts.append(doc['page_content'])
            
            # Auch Tool-Runs prüfen (university_knowledge_search)
            elif child.run_type == "tool" and "university_knowledge" in str(child.name).lower():
                if child.outputs:
                    if debug:
                        print(f"      🔍 DEBUG: Found university_knowledge tool run")
                        print(f"         Output type: {type(child.outputs)}")
                        print(f"         Output (first 200 chars): {str(child.outputs)[:200]}")
                    
                    # Outputs könnte ein String mit RAG-Ergebnis sein
                    # oder ein Dict mit weiteren Infos
                    if isinstance(child.outputs, dict) and 'output' in child.outputs:
                        contexts.append(str(child.outputs['output']))
                    elif isinstance(child.outputs, str):
                        contexts.append(child.outputs)
        
        if contexts:
            if debug:
                print(f"      ✅ DEBUG: {len(contexts)} Kontexte gefunden")
            return contexts  # Liste von Chunks zurückgeben

        if debug:
            print(f"      ⚠️ DEBUG: Keine Kontexte gefunden")
        # WICHTIG: Leere Liste statt Platzhalter-String, damit RAGAS-Metriken NaN liefern
        # und diese Zeilen in der Aggregation explizit ausgeschlossen werden können
        # (siehe display_and_save_results / _metric_stats).
        return []

    except Exception as e:
        print(f"      ⚠️ LangSmith-Fehler: {str(e)[:100]}")
        if debug:
            import traceback
            traceback.print_exc()
        # Leere Liste statt Fehler-String → Zeile wird in der Aggregation ausgeschlossen
        return []


def generate_chatbot_responses(
    df: pd.DataFrame, 
    agent, 
    langsmith_client: Client, 
    model_name: str = None, 
    resume: bool = True,
    retry_questions: List[int] = None,
    agent_type: str = "single",
    provider: str = "ollama"
) -> EvaluationDataset:
    """
    Generiert Chatbot-Antworten für alle Fragen und sammelt RAG-Kontexte.
    Speichert nach jeder Frage einen Checkpoint für Resume-Fähigkeit.
    
    Args:
        df: DataFrame mit Fragen
        agent: Agent-Instanz
        langsmith_client: LangSmith Client
        model_name: Name des verwendeten Modells (für Checkpoint-Validierung)
        resume: Wenn True, versuche von Checkpoint fortzusetzen; wenn False, starte neu
        retry_questions: Liste von Frage-Nummern (1-basiert) die wiederholt werden sollen
        agent_type: Agent-Typ (single, multi, constrained, confirmation) für agent-spezifischen Checkpoint
    """
    print("\n🤖 Generiere Chatbot-Antworten...")
    print("=" * 80)
    
    # Checkpoint-Pfad (agent-spezifisch)
    checkpoint_filename = f"responses_checkpoint_{agent_type}.pkl"
    checkpoint_path = Path(__file__).parent / "data" / checkpoint_filename
    checkpoint_path.parent.mkdir(exist_ok=True, parents=True)
    print(f"📁 Checkpoint-Pfad: {checkpoint_path}")
    print(f"   Verzeichnis existiert: {checkpoint_path.parent.exists()}")
    print(f"   Verzeichnis beschreibbar: {checkpoint_path.parent.is_dir()}\n")
    
    # Versuche vorhandenen Checkpoint zu laden
    samples = []
    start_idx = 0
    
    if checkpoint_path.exists() and resume:
        import pickle
        try:
            with open(checkpoint_path, 'rb') as f:
                checkpoint_data = pickle.load(f)
                
                # Validiere Checkpoint
                checkpoint_valid = True
                checkpoint_model = checkpoint_data.get('model_name')
                
                # Prüfe ob Checkpoint für gleiches Modell
                if model_name and checkpoint_model and checkpoint_model != model_name:
                    print(f"⚠️ Checkpoint ist für anderes Modell ({checkpoint_model} vs {model_name})")
                    print(f"   Starte frisch...")
                    checkpoint_valid = False
                
                # Prüfe ob Checkpoint vollständige Daten hat
                if checkpoint_valid and 'samples' in checkpoint_data:
                    samples = checkpoint_data['samples']
                    start_idx = len(samples)
                    print(f"📂 Checkpoint gefunden: {start_idx} Fragen bereits beantwortet")
                    print(f"   Modell: {checkpoint_model or 'unbekannt'}")
                    print(f"   Fortsetzung ab Frage {start_idx + 1}...\n")
                else:
                    samples = []
                    start_idx = 0
                    
        except Exception as e:
            print(f"⚠️ Checkpoint konnte nicht geladen werden: {e}")
            print(f"   Starte frisch...\n")
            samples = []
            start_idx = 0
    elif checkpoint_path.exists() and not resume:
        print(f"🗑️  Ignoriere vorhandenen Checkpoint (--no-resume gesetzt)")
        print(f"   Starte frisch...\n")
    
    # Retry-Questions verarbeiten: Fragen wiederholen während bestehender Progress beibehalten wird
    #
    # WICHTIG (B1/B2-Fix): retry_questions wird AUSSCHLIESSLICH auf einen vollständigen
    # Checkpoint angewendet (len(samples) == len(df)). Ein gemischter Modus
    # "retry + Fortsetzung des Tails" führte zu falscher Index-Ausrichtung
    # zwischen samples und test_df (Antworten landeten an falschen Positionen).
    #
    # Workflow:
    #   1. Erst eval normal zu Ende laufen lassen (alle Fragen beantwortet).
    #   2. Dann mit retry_questions=[...] gezielt einzelne Fragen überschreiben.
    questions_to_process = set()

    if retry_questions:
        # Konvertiere 1-basierte Indizes zu 0-basierten
        retry_indices = {q - 1 for q in retry_questions if 0 <= q - 1 < len(df)}

        if not retry_indices:
            print("⚠️  retry_questions enthielt keine gültigen Indizes → nichts zu tun.")
        else:
            if len(samples) != len(df):
                raise ValueError(
                    f"retry_questions benötigt einen vollständigen Checkpoint "
                    f"(len(samples)={len(samples)}, len(df)={len(df)}). "
                    f"Bitte zuerst die Evaluation ohne retry_questions vollständig durchlaufen lassen, "
                    f"dann gezielt Fragen mit retry_questions wiederholen."
                )
            print(f"\n🔄 Wiederhole {len(retry_indices)} Fragen: {sorted(q + 1 for q in retry_indices)}\n")
            questions_to_process = retry_indices

    # Wenn keine Retry-Questions: Normale Fortsetzung ab start_idx
    if not questions_to_process:
        questions_to_process = set(range(start_idx, len(df)))
    
    # Beantworte Fragen (entweder retry oder continuation)
    for idx in sorted(questions_to_process):
        if idx >= len(df):
            continue
            
        row = df.iloc[idx]
        question = row['question']
        expected_answer = row['expected_answer']
        
        print(f"\n[{idx + 1}/{len(df)}] {question[:70]}...")
        
        # Memory löschen für isolierte Evaluation
        agent.clear_memory()
        
        # Chatbot fragen - mit Session-ID für LangSmith-Tracking
        print(f"   💬 Chatbot fragen...")
        import uuid
        session_id = str(uuid.uuid4())
        
        # Agent.chat() mit session_id aufrufen
        answer = agent.chat(question, session_id=session_id)
        print(f"   ✅ Antwort: {answer[:80]}...")

        # RAG-Kontext aus LangSmith holen
        # WICHTIG: Wir müssen den LangGraph-Run finden, NICHT den self-reflection-Run!
        # Self-reflection Runs sind separate LLM calls die NACH dem LangGraph-Agent laufen.
        # Wir polling auf eine session_id-Übereinstimmung statt blind zu schlafen,
        # weil LangSmith-Ingestion eventually consistent ist und ein Fallback auf einen
        # "irgendeinen kürzlichen" Run die Kontexte einer ANDEREN Frage zuordnen würde.
        print(f"   🔍 Hole RAG-Kontext aus LangSmith (polling auf session_id)...")

        # Initial-Wait pro Provider:
        # - Anthropic: 5s wegen strikterer Rate-Limits.
        # - Lokale/OpenAI Modelle: 3s. Ein zu kleiner Initial-Wait führte
        #   dazu, dass schnelle lokale Modelle häufiger das Polling-Budget
        #   erschöpften, bevor LangSmith ihren Run sichtbar macht. Das Ergebnis
        #   waren pro Modell unterschiedlich große "no-context"-Anteile und
        #   damit nicht-vergleichbare RAGAS-Mittelwerte (siehe Audit B3).
        initial_wait = 5.0 if provider == "anthropic" else 3.0
        matching_run = _find_run_by_session_id(
            langsmith_client,
            session_id,
            max_attempts=6,
            initial_wait=initial_wait,
        )

        # Leere Liste = kein RAG-Kontext. Wird in display_and_save_results / _metric_stats
        # als invalid behandelt und aus den Mittelwerten entfernt (verhindert systematische
        # Verzerrung von Faithfulness / Context-Recall / Context-Precision).
        contexts: List[str] = []

        if matching_run is not None:
            trace_id = matching_run.trace_id
            contexts = get_rag_context_from_langsmith(langsmith_client, trace_id, debug=False)
            print(f"   ✅ Run gefunden (session_id match): {matching_run.name} (type: {matching_run.run_type})")
        else:
            # HARTE Warnung: Wir haben den Trace nicht gefunden. KEIN Fallback auf einen
            # anderen Run, weil das die RAGAS-Metriken stillschweigend kontaminieren würde.
            print(
                f"   ❌ HARD WARNING: Kein LangSmith-Run mit session_id={session_id} "
                f"innerhalb des Polling-Budgets gefunden. "
                f"Frage {idx + 1} wird OHNE RAG-Kontext gespeichert."
            )
        
        total_chars = sum(len(c) for c in contexts)
        print(f"   📄 Kontext: {len(contexts)} chunks, {total_chars} Zeichen")
        
        # RAGAS-Sample erstellen
        sample = SingleTurnSample(
            user_input=question,
            response=answer,
            retrieved_contexts=contexts,  # Jetzt bereits eine Liste von Chunks
            reference=expected_answer
        )
        
        # Wenn retry: Sample an gleicher Position überschreiben (samples ist vollständig);
        # sonst: append (samples wächst sequenziell von start_idx auf len(df)).
        if retry_questions:
            samples[idx] = sample
        else:
            samples.append(sample)
        
        # Checkpoint nach jeder Frage atomar speichern (~ms; tempfile + os.replace)
        try:
            checkpoint_data = {
                'samples': samples,
                'test_df': df,
                'last_idx': idx,
                'model_name': model_name  # Speichere Modellname
            }
            _save_checkpoint_atomic(checkpoint_path, checkpoint_data)
            print(f"   💾 Checkpoint gespeichert ({len(samples)}/{len(df)} Fragen)")
            print(f"      Pfad: {checkpoint_path}")
        except Exception as e:
            print(f"   ⚠️ FEHLER beim Checkpoint-Speichern: {e}")
    
    print("\n" + "=" * 80)
    print(f"✅ {len(samples)} Antworten generiert\n")
    
    # Finale Dataset-Konvertierung
    dataset = EvaluationDataset(samples=samples)
    
    # Finales Checkpoint mit vollständigem Dataset (atomar)
    try:
        checkpoint_data = {
            'dataset': dataset,
            'samples': samples,
            'test_df': df,
            'model_name': model_name  # Speichere Modellname
        }
        _save_checkpoint_atomic(checkpoint_path, checkpoint_data)
        print(f"💾 Finaler Checkpoint gespeichert: {checkpoint_path}")
        print(f"   (Antworten + Kontexte für alle {len(samples)} Fragen)\n")
    except Exception as e:
        print(f"⚠️ FEHLER beim finalen Checkpoint: {e}\n")
    
    return dataset


def run_ragas_evaluation(
    dataset: EvaluationDataset, 
    model: str = None,
    judge_provider: str = None,
    judge_model: str = None,
    max_workers: int = 8
) -> pd.DataFrame:
    """
    Führt RAGAS-Evaluation durch.
    Verwendet 3 Standard-RAGAS-Metriken: faithfulness, context_recall, context_precision.
    (answer_relevancy auskommentiert - benötigt qwen3-embedding:8b)
    
    Args:
        dataset: RAGAS EvaluationDataset mit Samples
        model: Agent-Modell (nur für Dokumentation)
        judge_provider: 'openai' oder 'ollama' (None = auto-detect)
        judge_model: Judge-Modell (None = use RAGAS_JUDGE_MODEL from settings)
        max_workers: Anzahl paralleler Workers (default: 8)
    """
    print("🚀 Starte RAGAS-Evaluation...")
    print("=" * 80)
    
    # Verwende immer den festen RAGAS Judge für faire Vergleiche
    print(f"   Agent-Modell:  {model if model else 'N/A'}")
    
    # Bestimme Judge-Provider und -Modell
    final_judge_provider = judge_provider
    final_judge_model = judge_model or RAGAS_JUDGE_MODEL
    
    # Auto-detect provider if not specified
    if final_judge_provider is None:
        if final_judge_model.startswith('gpt-'):
            final_judge_provider = 'openai'
        else:
            final_judge_provider = 'ollama'
    
    print(f"   RAGAS-Judge:   {final_judge_model} ({final_judge_provider})")
    print(f"   Max Workers:   {max_workers}")
    
    # LLM konfigurieren basierend auf Provider
    # WICHTIG: Judge MUSS deterministisch sein (temperature=0.0 + seed),
    # sonst sind RAGAS-Scores zwischen wiederholten Läufen nicht reproduzierbar
    # und Vergleiche zwischen Agent-Varianten werden verrauscht.
    JUDGE_TEMPERATURE = 0.0
    JUDGE_SEED = 42
    if final_judge_provider == 'openai':
        from langchain_openai import ChatOpenAI
        llm = ChatOpenAI(
            model=final_judge_model,
            temperature=JUDGE_TEMPERATURE,
            seed=JUDGE_SEED,
            api_key=settings.OPENAI_API_KEY,
            max_retries=3
        )
    else:
        # Ollama LLM konfigurieren (seed via model_kwargs, da ChatOllama
        # ihn nicht als Top-Level-Parameter exponiert)
        llm = ChatOllama(
            model=final_judge_model,
            base_url=OLLAMA_BASE_URL,
            temperature=JUDGE_TEMPERATURE,
            seed=JUDGE_SEED,
        )
    
    # Ollama Embeddings für answer_relevancy (später aktivieren)
    # embeddings = OllamaEmbeddings(
    #     model=OLLAMA_EMBEDDING_MODEL,
    #     base_url=OLLAMA_BASE_URL
    # )
    # print(f"   Embeddings: {OLLAMA_EMBEDDING_MODEL} @ {OLLAMA_BASE_URL}")
    
    # Standard RAGAS-Metriken
    metrics = [
        faithfulness,       # Ist Antwort treu zum Kontext?
        context_recall,     # Wurden alle relevanten Infos abgerufen?
        context_precision   # Sind relevante Chunks höher gerankt?
        # answer_relevancy  # Ist Antwort relevant zur Frage? (benötigt Embeddings)
    ]
    print(f"   Metriken: {[m.name for m in metrics]}")
    print(f"\n   ⏳ Evaluiere {len(dataset.samples)} Samples...")
    if final_judge_provider == 'openai' and max_workers > 50:
        print(f"   💡 Mit {max_workers} Workern - OpenAI kann sehr schnell sein!\n")
    else:
        print(f"   💡 Dies kann mehrere Minuten dauern (ca. 1-2 Min pro Sample)\n")
    
    # RunConfig für parallele Requests (seed für Reproduzierbarkeit der Sampling-Reihenfolge)
    run_config = RunConfig(max_workers=max_workers, seed=JUDGE_SEED)
    
    # Evaluation durchführen
    results = evaluate(
        dataset=dataset,
        metrics=metrics,
        llm=llm,
        run_config=run_config,
        raise_exceptions=False  # Weiter bei Fehlern
    )
    
    return results.to_pandas()


def _metric_stats(df: pd.DataFrame, metric: str) -> dict:
    """NaN- und Empty-Context-bewusste Aggregation einer RAGAS-Metrik.

    Schließt Zeilen ohne RAG-Kontext (context_count == 0) explizit aus, weil dort
    die Metrik systematisch 0 oder NaN ist und sonst die Vergleichbarkeit zwischen
    Agent-Varianten verzerren würde (Agenten, die seltener auf RAG routen, würden
    fälschlich schlechter aussehen).

    Reports:
      mean / std — über Zeilen mit Kontext UND nicht-NaN-Wert
      n_valid    — Zähler des Mittelwerts (Zeilen mit Kontext, Wert nicht NaN)
      n_total    — Gesamtzeilen in df
      n_no_ctx   — Zeilen mit context_count == 0 (ausgeschlossen)
      n_nan      — Zeilen mit Kontext aber NaN-Metrik (ausgeschlossen)
    """
    n_total = len(df)
    if metric not in df.columns or n_total == 0:
        return {"mean": float("nan"), "std": float("nan"),
                "n_valid": 0, "n_total": n_total, "n_no_ctx": 0, "n_nan": 0}

    if "context_count" in df.columns:
        has_ctx = df["context_count"] > 0
    else:
        # Fallback: Kontextliste direkt prüfen
        has_ctx = df["retrieved_contexts"].apply(
            lambda x: isinstance(x, list) and len(x) > 0
            and not (len(x) == 1 and isinstance(x[0], str) and x[0] in (
                "Kein RAG-Kontext gefunden", "LangSmith-Fehler"))
        ) if "retrieved_contexts" in df.columns else pd.Series([True] * n_total, index=df.index)

    n_no_ctx = int((~has_ctx).sum())
    series_with_ctx = df.loc[has_ctx, metric]
    n_nan = int(series_with_ctx.isna().sum())
    valid = series_with_ctx.dropna()
    n_valid = int(len(valid))

    return {
        "mean": float(valid.mean()) if n_valid > 0 else float("nan"),
        "std": float(valid.std()) if n_valid > 1 else float("nan"),
        "n_valid": n_valid,
        "n_total": n_total,
        "n_no_ctx": n_no_ctx,
        "n_nan": n_nan,
    }


def _format_metric_stats(stats: dict) -> str:
    """Kompakte einzeilige Darstellung: '0.647 (n=92/100, no_ctx=5, nan=3)'."""
    excl = stats["n_no_ctx"] + stats["n_nan"]
    if excl == 0:
        return f"{stats['mean']:.3f} (n={stats['n_valid']}/{stats['n_total']})"
    return (
        f"{stats['mean']:.3f} "
        f"(n={stats['n_valid']}/{stats['n_total']}, "
        f"no_ctx={stats['n_no_ctx']}, nan={stats['n_nan']})"
    )


def display_and_save_results(results_df: pd.DataFrame, test_df: pd.DataFrame):
    """Zeigt Ergebnisse an und speichert sie."""

    # IDs, Kategorien und Schwierigkeiten hinzufügen
    results_df['id'] = test_df['id'].values[:len(results_df)]
    results_df['category'] = test_df['category'].values[:len(results_df)]
    results_df['difficulty'] = test_df['difficulty'].values[:len(results_df)]

    # Context-Count FRÜH berechnen, damit _metric_stats sie für alle Aggregationen nutzen kann.
    # Zusätzlich: alte Platzhalter-Strings (aus historischen CSVs) als 0 zählen.
    def _ctx_count(x):
        if not isinstance(x, list):
            return 0
        if len(x) == 1 and isinstance(x[0], str) and x[0] in (
                "Kein RAG-Kontext gefunden", "LangSmith-Fehler"):
            return 0
        return len(x)
    results_df['context_count'] = results_df['retrieved_contexts'].apply(_ctx_count)

    print("\n" + "=" * 80)
    print("📊 RAGAS-EVALUATION ERGEBNISSE")
    print("=" * 80)

    # Gesamtscores (NaN- und Empty-Context-aware)
    print("\n📈 Durchschnittliche Scores:")
    print("-" * 80)
    for metric in ['faithfulness', 'context_recall', 'context_precision']:
        stats = _metric_stats(results_df, metric)
        print(f"   {metric:20s}: {_format_metric_stats(stats)}")

    # Nach Kategorie
    print("\n📁 Scores nach Kategorie:")
    print("-" * 80)
    for category in results_df['category'].unique():
        cat_df = results_df[results_df['category'] == category]
        print(f"\n   {category}:")
        for metric in ['faithfulness', 'context_recall', 'context_precision']:
            stats = _metric_stats(cat_df, metric)
            print(f"      {metric:20s}: {_format_metric_stats(stats)}")

    # Nach Schwierigkeit
    print("\n⚡ Scores nach Schwierigkeit:")
    print("-" * 80)
    for difficulty in ['easy', 'medium', 'hard']:
        diff_df = results_df[results_df['difficulty'] == difficulty]
        if len(diff_df) > 0:
            print(f"\n   {difficulty.upper()}:")
            for metric in ['faithfulness', 'context_recall', 'context_precision']:
                stats = _metric_stats(diff_df, metric)
                print(f"      {metric:20s}: {_format_metric_stats(stats)}")

    # Speichern in CSV (alle Spalten)
    output_path_csv = Path(__file__).parent / "data" / "ragas_results.csv"
    
    # Entferne Zeilenumbrüche aus Textfeldern für saubere CSV
    text_columns = ['user_input', 'response', 'reference']
    for col in text_columns:
        if col in results_df.columns:
            results_df[col] = results_df[col].apply(lambda x: x.replace('\n', ' ').replace('\r', ' ') if isinstance(x, str) else x)
    
    # Konvertiere retrieved_contexts zu String ohne Zeilenumbrüche
    results_df['retrieved_contexts'] = results_df['retrieved_contexts'].apply(
        lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if isinstance(x, list) else str(x)
    )
    
    # CSV mit allen wichtigen Spalten
    csv_df = results_df[['id', 'category', 'difficulty', 'user_input', 'response', 
                          'reference', 'retrieved_contexts', 'faithfulness', 
                          'context_recall', 'context_precision', 'context_count']].copy()
    
    # Speichere mit UTF-8-BOM für korrekte Umlaut-Darstellung
    csv_df.to_csv(output_path_csv, index=False, encoding='utf-8-sig', sep=',', quoting=1)
    
    # Excel mit Formatierung erstellen
    output_path_excel = Path(__file__).parent / "data" / "ragas_results.xlsx"
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Sheet 1: Detaillierte Ergebnisse
        ws_details = wb.active
        ws_details.title = "Detaillierte Ergebnisse"
        
        # Header-Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Daten schreiben
        for r_idx, row in enumerate(dataframe_to_rows(csv_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatieren
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Metriken-Spalten (faithfulness, context_recall, context_precision) farbig
                    if c_idx in [8, 9, 10]:  # Metrik-Spalten
                        if isinstance(value, (int, float)):
                            if value >= 0.8:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value >= 0.6:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.number_format = '0.000'
                    
                    # Text-Wrap für lange Texte
                    if c_idx in [4, 5, 6, 7]:  # user_input, response, reference, contexts
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Spaltenbreiten anpassen
        ws_details.column_dimensions['A'].width = 8   # id
        ws_details.column_dimensions['B'].width = 20  # category
        ws_details.column_dimensions['C'].width = 12  # difficulty
        ws_details.column_dimensions['D'].width = 50  # user_input
        ws_details.column_dimensions['E'].width = 60  # response
        ws_details.column_dimensions['F'].width = 50  # reference
        ws_details.column_dimensions['G'].width = 40  # contexts
        ws_details.column_dimensions['H'].width = 15  # faithfulness
        ws_details.column_dimensions['I'].width = 15  # context_recall
        ws_details.column_dimensions['J'].width = 17  # context_precision
        ws_details.column_dimensions['K'].width = 15  # context_count
        
        # Sheet 2: Zusammenfassung
        ws_summary = wb.create_sheet("Zusammenfassung")
        
        # Titel
        ws_summary['A1'] = "📊 RAGAS-Evaluation Zusammenfassung"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')
        
        # Durchschnittliche Scores (NaN- und Empty-Context-aware)
        row = 3
        ws_summary[f'A{row}'] = "Durchschnittliche Scores"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for metric in ['faithfulness', 'context_recall', 'context_precision']:
            stats = _metric_stats(results_df, metric)
            avg = stats["mean"]
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = avg if avg == avg else None  # NaN → leere Zelle
            ws_summary[f'C{row}'] = (
                f"n={stats['n_valid']}/{stats['n_total']} "
                f"(no_ctx={stats['n_no_ctx']}, nan={stats['n_nan']})"
            )
            ws_summary[f'B{row}'].number_format = '0.000'

            # Farbe basierend auf Score (NaN → keine Einfärbung)
            if avg == avg:
                if avg >= 0.8:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif avg >= 0.6:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1
        
        # Nach Kategorie
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Kategorie"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header für Kategorie-Tabelle
        ws_summary[f'A{row}'] = "Kategorie"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        for category in sorted(results_df['category'].unique()):
            cat_df = results_df[results_df['category'] == category]
            ws_summary[f'A{row}'] = category

            for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision'], 2):
                stats = _metric_stats(cat_df, metric)
                avg = stats["mean"]
                col_letter = chr(65 + idx)  # B, C, D
                ws_summary[f'{col_letter}{row}'] = avg if avg == avg else None
                ws_summary[f'{col_letter}{row}'].number_format = '0.000'

            row += 1
        
        # Nach Schwierigkeit
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Schwierigkeit"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header
        ws_summary[f'A{row}'] = "Schwierigkeit"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        for difficulty in ['easy', 'medium', 'hard']:
            diff_df = results_df[results_df['difficulty'] == difficulty]
            if len(diff_df) > 0:
                ws_summary[f'A{row}'] = difficulty.upper()

                for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision'], 2):
                    stats = _metric_stats(diff_df, metric)
                    avg = stats["mean"]
                    col_letter = chr(65 + idx)
                    ws_summary[f'{col_letter}{row}'] = avg if avg == avg else None
                    ws_summary[f'{col_letter}{row}'].number_format = '0.000'

                row += 1
        
        # Spaltenbreiten für Zusammenfassung
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 18
        
        # Speichern
        wb.save(output_path_excel)
        
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV:   {output_path_csv}")
        print(f"   Excel: {output_path_excel}")
        print("=" * 80 + "\n")
        
    except ImportError:
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV: {output_path_csv}")
        print(f"   ⚠️ Excel-Export nicht verfügbar (openpyxl nicht installiert)")
        print("=" * 80 + "\n")


def main():
    """Hauptfunktion"""
    
    print("\n" + "=" * 80)
    print("🎯 RAGAS-EVALUATION - WiSo-Chatbot")
    print("=" * 80 + "\n")
    
    # Checkpoint-Pfad
    checkpoint_path = Path(__file__).parent / "data" / "responses_checkpoint.pkl"
    
    try:
        # Prüfe ob Checkpoint existiert
        if checkpoint_path.exists():
            print("📂 Lade Checkpoint...")
            import pickle
            with open(checkpoint_path, 'rb') as f:
                checkpoint_data = pickle.load(f)
            
            # Checkpoint kann EvaluationDataset oder dict sein
            if isinstance(checkpoint_data, dict):
                dataset = checkpoint_data['dataset']
                test_df = checkpoint_data['test_df']
            else:
                # Alter Checkpoint-Format (nur Dataset)
                dataset = checkpoint_data
                # test_df muss neu geladen werden
                test_df = load_testset()  # Alle Fragen laden
            
            print(f"   ✅ {len(dataset.samples)} Antworten aus Checkpoint geladen\n")
            
        else:
            # Kein Checkpoint → Vollständiger Durchlauf
            # 1. LangSmith Client
            print("🔗 Initialisiere LangSmith...")
            langsmith_client = Client(api_key=LANGSMITH_API_KEY)
            print(f"   ✅ Projekt: {LANGSMITH_PROJECT}\n")
            
            # 2. Testset laden (alle Fragen)
            print("📂 Lade Testset...")
            test_df = load_testset()  # Alle Fragen laden
            print()
            
            # 3. Chatbot initialisieren
            print("🤖 Initialisiere Chatbot...")
            agent = create_react_agent()
            print()
            
            # 4. Antworten generieren
            dataset = generate_chatbot_responses(test_df, agent, langsmith_client)
        
        # 5. RAGAS-Evaluation (immer ausführen)
        results_df = run_ragas_evaluation(dataset, model=OLLAMA_MODEL)
        
        # 6. Ergebnisse anzeigen und speichern
        display_and_save_results(results_df, test_df)
        
        print("✅ Evaluation erfolgreich abgeschlossen!")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Evaluation abgebrochen!\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Fehler: {e}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
