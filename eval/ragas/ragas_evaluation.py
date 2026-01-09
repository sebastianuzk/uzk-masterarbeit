"""
RAGAS-Evaluation für WiSo-Chatbot

Evaluiert den Chatbot mit RAGAS-Framework:
- Verwendet Ollama (qwen3:8b) als LLM-Judge
- Lädt Testfragen aus Testset.CSV
- Generiert Antworten mit dem Chatbot
- Extrahiert RAG-Kontexte aus LangSmith
- Berechnet RAGAS-Metriken (Faithfulness, Context Recall)
"""

import sys
import pandas as pd
from pathlib import Path
from typing import List
import time

# Projekt-Root
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from ragas import evaluate
from ragas.metrics import faithfulness, context_recall, context_precision  # answer_relevancy benötigt Embeddings
from ragas.dataset_schema import SingleTurnSample, EvaluationDataset
from ragas.run_config import RunConfig
from langchain_ollama import ChatOllama, OllamaEmbeddings
from langsmith import Client
from config.settings import (
    OLLAMA_MODEL,
    OLLAMA_BASE_URL,
    OLLAMA_EMBEDDING_MODEL,
    LANGSMITH_API_KEY,
    LANGSMITH_PROJECT,
    RAGAS_JUDGE_MODEL
)
from src.agent.react_agent import create_react_agent


def load_testset(csv_path: str = "data/Testset.CSV", limit: int = None) -> pd.DataFrame:
    """Lädt Testset.CSV"""
    full_path = Path(__file__).parent / csv_path
    df = pd.read_csv(full_path, sep=';', encoding='utf-8')
    
    if limit:
        df = df.head(limit)
    
    print(f"✅ {len(df)} Testfragen geladen")
    
    return df


def get_rag_context_from_langsmith(client: Client, trace_id: str) -> List[str]:
    """
    Holt RAG-Kontext aus LangSmith für eine spezifische Trace-ID.
    
    Die Documents befinden sich im Retriever-Output unter dem Key 'output' (nicht 'documents'!).
    Jedes Document hat 'page_content' und 'metadata'.
    
    Args:
        client: LangSmith Client
        trace_id: Die Trace-ID der Session
        
    Returns:
        Liste von RAG-Context-Chunks aus den Retriever-Documents
    """
    try:
        # Hole alle Child-Runs für diese Trace
        child_runs = list(client.list_runs(
            project_name=LANGSMITH_PROJECT,
            trace_id=trace_id,
            is_root=False
        ))
        
        # Suche nach Retriever-Run
        contexts = []
        for child in child_runs:
            if child.run_type == "retriever":
                if child.outputs and isinstance(child.outputs, dict):
                    # Documents sind unter 'output' Key (nicht 'documents')!
                    documents = child.outputs.get('output', [])
                    for doc in documents:
                        if isinstance(doc, dict) and 'page_content' in doc:
                            contexts.append(doc['page_content'])
        
        if contexts:
            return contexts  # Liste von Chunks zurückgeben
        
        return ["Kein RAG-Kontext gefunden"]  # Als Liste
    
    except Exception as e:
        print(f"      ⚠️ LangSmith-Fehler: {str(e)[:100]}")
        return ["LangSmith-Fehler"]  # Als Liste


def generate_chatbot_responses(df: pd.DataFrame, agent, langsmith_client: Client) -> EvaluationDataset:
    """
    Generiert Chatbot-Antworten für alle Fragen und sammelt RAG-Kontexte.
    """
    print("\n🤖 Generiere Chatbot-Antworten...")
    print("=" * 80)
    
    samples = []
    
    for idx, row in df.iterrows():
        question = row['question']
        expected_answer = row['expected_answer']
        
        print(f"\n[{idx + 1}/{len(df)}] {question[:70]}...")
        
        # Memory löschen für isolierte Evaluation
        agent.clear_memory()
        
        # Chatbot fragen - mit Session-ID für LangSmith-Tracking
        print(f"   💬 Chatbot fragen...")
        import uuid
        session_id = str(uuid.uuid4())
        
        # Agent.chat() mit session_id aufrufen
        answer = agent.chat(question, session_id=session_id)
        print(f"   ✅ Antwort: {answer[:80]}...")
        
        # Warten damit LangSmith Trace vollständig ist
        time.sleep(1)  # Reduziert von 3s auf 1s
        
        # RAG-Kontext aus LangSmith holen - nur den letzten Run abrufen
        print(f"   🔍 Hole RAG-Kontext aus LangSmith...")
        
        # Optimiert: Nur den letzten Run holen (statt alle)
        recent_runs = list(langsmith_client.list_runs(
            project_name=LANGSMITH_PROJECT,
            is_root=True,
            limit=1  # Nur den letzten Run
        ))
        
        contexts = ["Kein RAG-Kontext gefunden"]  # Default als Liste
        matching_run = None
        
        # Der letzte Run sollte unser Run sein
        if recent_runs:
            matching_run = recent_runs[0]
        
        if matching_run:
            trace_id = matching_run.trace_id
            contexts = get_rag_context_from_langsmith(langsmith_client, trace_id)
            print(f"   ✅ Run gefunden mit Session-ID: {session_id[:8]}...")
        else:
            print(f"   ⚠️ Kein Run mit Session-ID {session_id[:8]}... gefunden")
        
        total_chars = sum(len(c) for c in contexts)
        print(f"   📄 Kontext: {len(contexts)} chunks, {total_chars} Zeichen")
        
        # RAGAS-Sample erstellen
        sample = SingleTurnSample(
            user_input=question,
            response=answer,
            retrieved_contexts=contexts,  # Jetzt bereits eine Liste von Chunks
            reference=expected_answer
        )
        samples.append(sample)
    
    print("\n" + "=" * 80)
    print(f"✅ {len(samples)} Antworten generiert\n")
    
    # Zwischenspeicherung der Antworten und Kontexte
    dataset = EvaluationDataset(samples=samples)
    checkpoint_path = Path(__file__).parent / "data" / "responses_checkpoint.pkl"
    checkpoint_path.parent.mkdir(exist_ok=True)
    
    import pickle
    checkpoint_data = {
        'dataset': dataset,
        'test_df': df
    }
    with open(checkpoint_path, 'wb') as f:
        pickle.dump(checkpoint_data, f)
    print(f"💾 Checkpoint gespeichert: {checkpoint_path}")
    print(f"   (Antworten + Kontexte für alle {len(samples)} Fragen)\n")
    
    return dataset


def run_ragas_evaluation(dataset: EvaluationDataset, model: str = None) -> pd.DataFrame:
    """
    Führt RAGAS-Evaluation durch.
    Verwendet 3 Standard-RAGAS-Metriken: faithfulness, context_recall, context_precision.
    (answer_relevancy auskommentiert - benötigt qwen3-embedding:8b)
    
    Args:
        dataset: RAGAS EvaluationDataset mit Samples
        model: Agent-Modell (nur für Dokumentation, Judge ist immer RAGAS_JUDGE_MODEL)
    """
    print("🚀 Starte RAGAS-Evaluation...")
    print("=" * 80)
    
    # Verwende immer den festen RAGAS Judge für faire Vergleiche
    print(f"   Agent-Modell:  {model if model else 'N/A'}")
    print(f"   RAGAS-Judge:   {RAGAS_JUDGE_MODEL}")
    
    # Ollama LLM konfigurieren mit festem Judge
    llm = ChatOllama(
        model=RAGAS_JUDGE_MODEL,
        base_url=OLLAMA_BASE_URL,
        temperature=0.0
    )
    
    # Ollama Embeddings für answer_relevancy (später aktivieren)
    # embeddings = OllamaEmbeddings(
    #     model=OLLAMA_EMBEDDING_MODEL,
    #     base_url=OLLAMA_BASE_URL
    # )
    # print(f"   Embeddings: {OLLAMA_EMBEDDING_MODEL} @ {OLLAMA_BASE_URL}")
    
    # Standard RAGAS-Metriken
    metrics = [
        faithfulness,       # Ist Antwort treu zum Kontext?
        context_recall,     # Wurden alle relevanten Infos abgerufen?
        context_precision   # Sind relevante Chunks höher gerankt?
        # answer_relevancy  # Ist Antwort relevant zur Frage? (benötigt Embeddings)
    ]
    print(f"   Metriken: {[m.name for m in metrics]}")
    print(f"\n   ⏳ Evaluiere {len(dataset.samples)} Samples...")
    print(f"   💡 Dies kann mehrere Minuten dauern (ca. 1-2 Min pro Sample)\n")
    
    # RunConfig für parallele Requests an Ollama
    run_config = RunConfig(max_workers=8)
    
    # Evaluation durchführen
    results = evaluate(
        dataset=dataset,
        metrics=metrics,
        llm=llm,
        run_config=run_config,
        raise_exceptions=False  # Weiter bei Fehlern
    )
    
    return results.to_pandas()


def display_and_save_results(results_df: pd.DataFrame, test_df: pd.DataFrame):
    """Zeigt Ergebnisse an und speichert sie."""
    
    # IDs, Kategorien und Schwierigkeiten hinzufügen
    results_df['id'] = test_df['id'].values[:len(results_df)]
    results_df['category'] = test_df['category'].values[:len(results_df)]
    results_df['difficulty'] = test_df['difficulty'].values[:len(results_df)]
    
    print("\n" + "=" * 80)
    print("📊 RAGAS-EVALUATION ERGEBNISSE")
    print("=" * 80)
    
    # Gesamtscores
    print("\n📈 Durchschnittliche Scores:")
    print("-" * 80)
    for metric in ['faithfulness', 'context_recall', 'context_precision']:
        if metric in results_df.columns:
            avg = results_df[metric].mean()
            print(f"   {metric:20s}: {avg:.3f}")
    
    # Nach Kategorie
    print("\n📁 Scores nach Kategorie:")
    print("-" * 80)
    for category in results_df['category'].unique():
        cat_df = results_df[results_df['category'] == category]
        print(f"\n   {category}:")
        for metric in ['faithfulness', 'context_recall', 'context_precision']:
            if metric in cat_df.columns:
                avg = cat_df[metric].mean()
                print(f"      {metric:20s}: {avg:.3f}")
    
    # Nach Schwierigkeit
    print("\n⚡ Scores nach Schwierigkeit:")
    print("-" * 80)
    for difficulty in ['easy', 'medium', 'hard']:
        diff_df = results_df[results_df['difficulty'] == difficulty]
        if len(diff_df) > 0:
            print(f"\n   {difficulty.upper()}:")
            for metric in ['faithfulness', 'context_recall', 'context_precision']:
                if metric in diff_df.columns:
                    avg = diff_df[metric].mean()
                    print(f"      {metric:20s}: {avg:.3f}")
    
    # Speichern in CSV (alle Spalten)
    output_path_csv = Path(__file__).parent / "data" / "ragas_results.csv"
    
    # Berechne Anzahl der Context-Chunks
    results_df['context_count'] = results_df['retrieved_contexts'].apply(lambda x: len(x) if isinstance(x, list) else 0)
    
    # Entferne Zeilenumbrüche aus Textfeldern für saubere CSV
    text_columns = ['user_input', 'response', 'reference']
    for col in text_columns:
        if col in results_df.columns:
            results_df[col] = results_df[col].apply(lambda x: x.replace('\n', ' ').replace('\r', ' ') if isinstance(x, str) else x)
    
    # Konvertiere retrieved_contexts zu String ohne Zeilenumbrüche
    results_df['retrieved_contexts'] = results_df['retrieved_contexts'].apply(
        lambda x: str(x).replace('\n', ' ').replace('\r', ' ') if isinstance(x, list) else str(x)
    )
    
    # CSV mit allen wichtigen Spalten
    csv_df = results_df[['id', 'category', 'difficulty', 'user_input', 'response', 
                          'reference', 'retrieved_contexts', 'faithfulness', 
                          'context_recall', 'context_precision', 'context_count']].copy()
    
    # Speichere mit UTF-8-BOM für korrekte Umlaut-Darstellung
    csv_df.to_csv(output_path_csv, index=False, encoding='utf-8-sig', sep=',', quoting=1)
    
    # Excel mit Formatierung erstellen
    output_path_excel = Path(__file__).parent / "data" / "ragas_results.xlsx"
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Sheet 1: Detaillierte Ergebnisse
        ws_details = wb.active
        ws_details.title = "Detaillierte Ergebnisse"
        
        # Header-Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Daten schreiben
        for r_idx, row in enumerate(dataframe_to_rows(csv_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatieren
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Metriken-Spalten (faithfulness, context_recall, context_precision) farbig
                    if c_idx in [8, 9, 10]:  # Metrik-Spalten
                        if isinstance(value, (int, float)):
                            if value >= 0.8:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value >= 0.6:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.number_format = '0.000'
                    
                    # Text-Wrap für lange Texte
                    if c_idx in [4, 5, 6, 7]:  # user_input, response, reference, contexts
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Spaltenbreiten anpassen
        ws_details.column_dimensions['A'].width = 8   # id
        ws_details.column_dimensions['B'].width = 20  # category
        ws_details.column_dimensions['C'].width = 12  # difficulty
        ws_details.column_dimensions['D'].width = 50  # user_input
        ws_details.column_dimensions['E'].width = 60  # response
        ws_details.column_dimensions['F'].width = 50  # reference
        ws_details.column_dimensions['G'].width = 40  # contexts
        ws_details.column_dimensions['H'].width = 15  # faithfulness
        ws_details.column_dimensions['I'].width = 15  # context_recall
        ws_details.column_dimensions['J'].width = 17  # context_precision
        ws_details.column_dimensions['K'].width = 15  # context_count
        
        # Sheet 2: Zusammenfassung
        ws_summary = wb.create_sheet("Zusammenfassung")
        
        # Titel
        ws_summary['A1'] = "📊 RAGAS-Evaluation Zusammenfassung"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')
        
        # Durchschnittliche Scores
        row = 3
        ws_summary[f'A{row}'] = "Durchschnittliche Scores"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for metric in ['faithfulness', 'context_recall', 'context_precision']:
            if metric in results_df.columns:
                avg = results_df[metric].mean()
                ws_summary[f'A{row}'] = metric
                ws_summary[f'B{row}'] = avg
                ws_summary[f'B{row}'].number_format = '0.000'
                
                # Farbe basierend auf Score
                if avg >= 0.8:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif avg >= 0.6:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1
        
        # Nach Kategorie
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Kategorie"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header für Kategorie-Tabelle
        ws_summary[f'A{row}'] = "Kategorie"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        for category in sorted(results_df['category'].unique()):
            cat_df = results_df[results_df['category'] == category]
            ws_summary[f'A{row}'] = category
            
            for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision'], 2):
                if metric in cat_df.columns:
                    avg = cat_df[metric].mean()
                    col_letter = chr(65 + idx)  # B, C, D
                    ws_summary[f'{col_letter}{row}'] = avg
                    ws_summary[f'{col_letter}{row}'].number_format = '0.000'
            
            row += 1
        
        # Nach Schwierigkeit
        row += 2
        ws_summary[f'A{row}'] = "Scores nach Schwierigkeit"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header
        ws_summary[f'A{row}'] = "Schwierigkeit"
        ws_summary[f'B{row}'] = "Faithfulness"
        ws_summary[f'C{row}'] = "Context Recall"
        ws_summary[f'D{row}'] = "Context Precision"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1
        
        for difficulty in ['easy', 'medium', 'hard']:
            diff_df = results_df[results_df['difficulty'] == difficulty]
            if len(diff_df) > 0:
                ws_summary[f'A{row}'] = difficulty.upper()
                
                for idx, metric in enumerate(['faithfulness', 'context_recall', 'context_precision'], 2):
                    if metric in diff_df.columns:
                        avg = diff_df[metric].mean()
                        col_letter = chr(65 + idx)
                        ws_summary[f'{col_letter}{row}'] = avg
                        ws_summary[f'{col_letter}{row}'].number_format = '0.000'
                
                row += 1
        
        # Spaltenbreiten für Zusammenfassung
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 18
        
        # Speichern
        wb.save(output_path_excel)
        
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV:   {output_path_csv}")
        print(f"   Excel: {output_path_excel}")
        print("=" * 80 + "\n")
        
    except ImportError:
        print("\n" + "=" * 80)
        print(f"💾 Ergebnisse gespeichert:")
        print(f"   CSV: {output_path_csv}")
        print(f"   ⚠️ Excel-Export nicht verfügbar (openpyxl nicht installiert)")
        print("=" * 80 + "\n")


def main():
    """Hauptfunktion"""
    
    print("\n" + "=" * 80)
    print("🎯 RAGAS-EVALUATION - WiSo-Chatbot")
    print("=" * 80 + "\n")
    
    # Checkpoint-Pfad
    checkpoint_path = Path(__file__).parent / "data" / "responses_checkpoint.pkl"
    
    try:
        # Prüfe ob Checkpoint existiert
        if checkpoint_path.exists():
            print("📂 Lade Checkpoint...")
            import pickle
            with open(checkpoint_path, 'rb') as f:
                checkpoint_data = pickle.load(f)
            
            # Checkpoint kann EvaluationDataset oder dict sein
            if isinstance(checkpoint_data, dict):
                dataset = checkpoint_data['dataset']
                test_df = checkpoint_data['test_df']
            else:
                # Alter Checkpoint-Format (nur Dataset)
                dataset = checkpoint_data
                # test_df muss neu geladen werden
                test_df = load_testset()  # Alle Fragen laden
            
            print(f"   ✅ {len(dataset.samples)} Antworten aus Checkpoint geladen\n")
            
        else:
            # Kein Checkpoint → Vollständiger Durchlauf
            # 1. LangSmith Client
            print("🔗 Initialisiere LangSmith...")
            langsmith_client = Client(api_key=LANGSMITH_API_KEY)
            print(f"   ✅ Projekt: {LANGSMITH_PROJECT}\n")
            
            # 2. Testset laden (alle Fragen)
            print("📂 Lade Testset...")
            test_df = load_testset()  # Alle Fragen laden
            print()
            
            # 3. Chatbot initialisieren
            print("🤖 Initialisiere Chatbot...")
            agent = create_react_agent()
            print()
            
            # 4. Antworten generieren
            dataset = generate_chatbot_responses(test_df, agent, langsmith_client)
        
        # 5. RAGAS-Evaluation (immer ausführen)
        results_df = run_ragas_evaluation(dataset, model=OLLAMA_MODEL)
        
        # 6. Ergebnisse anzeigen und speichern
        display_and_save_results(results_df, test_df)
        
        print("✅ Evaluation erfolgreich abgeschlossen!")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Evaluation abgebrochen!\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Fehler: {e}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
